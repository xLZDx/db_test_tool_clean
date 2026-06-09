#!/usr/bin/env python3
"""
api_first_report_builder.py

v18.0 API-first report builder.

Key design:
- Builds native JSON/API report models directly from raw compare/insert CSV/JSON outputs.
- Excel workbooks are optional renderings FROM the API model.
- API mode does not create or read Excel workbooks.
- No reverse Excel-to-JSON conversion exists in this API path.

Outputs under final_reports/:
  api/manifest.json
  api/step1_compare_report.json
  api/step4_full_cycle_report.json
  api/step1/tabs/*.json
  api/step4/tabs/*.json
  api/openapi_contract.json

Optional Excel outputs:
  step1_compare_report.xlsx
  step4_full_cycle_report.xlsx
  final_full_cycle_report.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Sequence


DRD_ODI_HEADERS = [
    "Area / Columns",
    "Conclusion",
    "Difference Type",
    "Mapping Logic",
    "ODI XML Logic",
    "Recommended Action",
]

THREE_WAY_HEADERS = [
    "Area / Columns",
    "Conclusion",
    "Difference Type",
    "DRD Mapping Logic",
    "ODI File 1 XML Logic",
    "ODI File 2 XML Logic",
    "Recommended Action",
]

ODI_ODI_HEADERS = [
    "Area / Columns",
    "Conclusion",
    "Difference Type",
    "DRD Mapping Logic",
    "ODI File 1 XML Logic",
    "ODI File 2 XML Logic",
    "Recommended Action",
]

INSERT_HEADERS = [
    "Area / Columns",
    "Conclusion",
    "Difference Type",
    "DRD Mapping Logic",
    "ODI XML Logic",
    "Generated INSERT Logic",
    "Recommended Action",
]


def clean(x) -> str:
    return (x or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def short_text(x: str, limit: int = 12000) -> str:
    s = clean(x)
    return s if len(s) <= limit else s[:limit - 40] + "\n...[truncated]..."


def read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def read_json(path: Path) -> Any:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return {"_error": str(exc), "_path": str(path)}


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def parse_area_columns(area: str) -> List[str]:
    area = clean(area)
    cols = re.findall(r"`([^`]+)`", area)
    if not cols and re.fullmatch(r"[A-Z0-9_#$]+", area):
        cols = [area]
    return [c.strip().upper() for c in cols if c.strip()]


def safe_name(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9_]+", "_", clean(name)).strip("_")
    return s or "tab"


def build_tab(name: str, headers: Sequence[str], rows: Sequence[Dict[str, str]]) -> Dict[str, Any]:
    return {
        "name": name,
        "safe_name": safe_name(name),
        "headers": list(headers),
        "rows": [{h: r.get(h, "") for h in headers} for r in rows],
        "row_count": len(rows),
        "column_count": len(headers),
    }


def independent_conclusion(row: Dict[str, str]) -> str:
    conclusion = clean(row.get("Conclusion", ""))
    diff = clean(row.get("Difference Type", ""))
    area = clean(row.get("Area / Columns", ""))

    if "missing" in diff.lower() or "not present" in conclusion.lower():
        return "Mismatch. Mapping logic is not implemented in the ODI XML final/resolved lineage."
    if "source drift" in diff.lower():
        return "Review required. ODI source or lineage differs from the mapping rule."
    if "case" in diff.lower():
        return "Review required. CASE/transformation logic differs from the mapping rule."
    if "structural" in diff.lower() or "column count" in diff.lower() or "final target" in area.lower():
        return "Review required. Structural target-load behavior differs from the mapping contract."
    return "Review required. ODI XML logic should be validated against the mapping rule."


def recommended_action(row: Dict[str, str]) -> str:
    diff = clean(row.get("Difference Type", ""))
    if "missing" in diff.lower():
        return "Implement the required mapping logic in ODI or document an approved exception."
    if "source drift" in diff.lower():
        return "Validate source lineage against the mapping rule. Update ODI or document an approved exception."
    if "structural" in diff.lower() or "column count" in diff.lower():
        return "Validate target-load contract and column list. Update ODI/DRD or document an approved exception."
    return "Review and approve, or update ODI to match the mapping rule."


def normalize_drd_vs_odi_rows(rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    out = []
    for r in rows:
        area = clean(r.get("Area / Columns", ""))
        if not area:
            continue
        out.append({
            "Area / Columns": area,
            "Conclusion": independent_conclusion(r),
            "Difference Type": clean(r.get("Difference Type", "")),
            "Mapping Logic": clean(r.get("Mapping Logic", "")),
            "ODI XML Logic": clean(r.get("ODI XML Logic", "")),
            "Recommended Action": recommended_action(r),
        })
    return out


def statuses_for_area(area: str, delta_by_col: Dict[str, Dict[str, str]]) -> List[str]:
    statuses = []
    for col in parse_area_columns(area):
        if col in delta_by_col:
            statuses.append(clean(delta_by_col[col].get("delta_status", "")))
    return statuses


def status_label_for_area(area: str, delta_by_col: Dict[str, Dict[str, str]]) -> str:
    statuses = statuses_for_area(area, delta_by_col)
    if not statuses:
        return ""
    uniq = list(dict.fromkeys(statuses))
    return uniq[0] if len(uniq) == 1 else "MIXED: " + " | ".join(uniq)


def filter_file2_rows(rows: Sequence[Dict[str, str]], delta_rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    delta_by_col = {clean(r.get("target_column", "")).upper(): r for r in delta_rows if clean(r.get("target_column", ""))}
    omit_statuses = {"FIXED_BY_RESOLVED_RULE_PROOF", "UNCHANGED_NO_ACTIVE_MISMATCH_BY_V16_6"}
    kept = []
    for r in rows:
        cols = parse_area_columns(r.get("Area / Columns", ""))
        statuses = statuses_for_area(r.get("Area / Columns", ""), delta_by_col)
        if cols and statuses and all(s in omit_statuses for s in statuses):
            continue
        kept.append(r)
    return kept


def three_way_conclusion(area: str, delta_by_col: Dict[str, Dict[str, str]]) -> str:
    statuses = statuses_for_area(area, delta_by_col)
    status_set = set(statuses)
    if not statuses:
        return "Review required. This grouped/structural item cannot be resolved to a single target-column delta."
    if status_set <= {"FIXED_BY_RESOLVED_RULE_PROOF", "UNCHANGED_NO_ACTIVE_MISMATCH_BY_V16_6"}:
        return "ODI File 2 appears aligned for this mapping item; ODI File 1 shows a difference against the mapping."
    if "STILL_OPEN" in status_set:
        return "Still open. ODI File 2 still does not fully implement this mapping item."
    if "FIX_CANDIDATE_UPSTREAM_CHANGED" in status_set:
        return "Review required. ODI File 2 has related upstream changes, but equivalence to the mapping rule is not fully proven."
    return "Review required. Compare DRD logic against both ODI implementations."


def three_way_action(area: str, delta_by_col: Dict[str, Dict[str, str]]) -> str:
    statuses = statuses_for_area(area, delta_by_col)
    status_set = set(statuses)
    if not statuses:
        return "Manually validate structural/grouped behavior against both ODI files and the DRD mapping."
    if status_set <= {"FIXED_BY_RESOLVED_RULE_PROOF", "UNCHANGED_NO_ACTIVE_MISMATCH_BY_V16_6"}:
        return "Use ODI File 2 behavior as the candidate implementation; keep File 1 as historical comparison evidence."
    if "STILL_OPEN" in status_set:
        return "Update ODI File 2 or document an approved exception."
    if "FIX_CANDIDATE_UPSTREAM_CHANGED" in status_set:
        return "Manually validate File 2 upstream logic against the DRD rule; if equivalent, mark accepted."
    return "Review and approve the intended ODI implementation."


def build_three_way_report(original_rows: Sequence[Dict[str, str]], fixed_rows: Sequence[Dict[str, str]], delta_rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    fixed_by_area = {clean(r.get("Area / Columns", "")): r for r in fixed_rows}
    delta_by_col = {clean(r.get("target_column", "")).upper(): r for r in delta_rows if clean(r.get("target_column", ""))}
    out = []
    for r in original_rows:
        area = clean(r.get("Area / Columns", ""))
        if not area:
            continue
        fixed = fixed_by_area.get(area, {})
        label = status_label_for_area(area, delta_by_col)
        base_diff = clean(r.get("Difference Type", ""))
        out.append({
            "Area / Columns": area,
            "Conclusion": three_way_conclusion(area, delta_by_col),
            "Difference Type": f"{base_diff} | {label}" if label else base_diff,
            "DRD Mapping Logic": clean(r.get("Mapping Logic", "")),
            "ODI File 1 XML Logic": clean(r.get("ODI XML Logic", "")),
            "ODI File 2 XML Logic": clean(fixed.get("ODI XML Logic", "")),
            "Recommended Action": three_way_action(area, delta_by_col),
        })
    return out


def build_drd_logic_map(rows: Sequence[Dict[str, str]]) -> Dict[str, str]:
    out = {}
    for r in rows:
        area = clean(r.get("Area / Columns", ""))
        logic = clean(r.get("Mapping Logic", "") or r.get("DRD Mapping Logic", ""))
        if not area or not logic:
            continue
        for col in parse_area_columns(area):
            out[col] = logic
    return out


def odi_column_conclusion(status: str) -> str:
    if status == "CHANGED":
        return "Different ODI implementation detected for the same target column."
    if status == "ADDED_IN_FIXED":
        return "Column exists in ODI File 2 but not in ODI File 1."
    if status == "REMOVED_IN_FIXED":
        return "Column exists in ODI File 1 but not in ODI File 2."
    if status == "NO_RESOLVED_LINEAGE":
        return "Resolved lineage was not available for one or both ODI files."
    return "No difference detected."


def odi_sql_conclusion(status: str) -> str:
    if status == "CHANGED":
        return "SQL block differs between ODI File 1 and ODI File 2."
    if status == "ADDED_IN_FIXED":
        return "SQL block exists in ODI File 2 but not in ODI File 1."
    if status == "REMOVED_IN_FIXED":
        return "SQL block exists in ODI File 1 but not in ODI File 2."
    return "No difference detected."


def odi_action(status: str) -> str:
    if status in {"CHANGED", "ADDED_IN_FIXED", "REMOVED_IN_FIXED", "NO_RESOLVED_LINEAGE"}:
        return "Review whether this ODI change is expected. Validate target output impact and document approval."
    return "No action."


def build_odi_vs_odi_columns(rows: Sequence[Dict[str, str]], drd_logic_by_col: Dict[str, str]) -> List[Dict[str, str]]:
    out = []
    for r in rows:
        status = clean(r.get("xml_delta_status", ""))
        if status == "UNCHANGED":
            continue
        col = clean(r.get("target_column", ""))
        old_parts, new_parts = [], []
        if r.get("original_final_expression"):
            old_parts.append("Final expression:\n" + clean(r.get("original_final_expression")))
        if r.get("original_resolved_expression"):
            old_parts.append("Resolved expression:\n" + clean(r.get("original_resolved_expression")))
        if r.get("original_lineage_path"):
            old_parts.append("Lineage path:\n" + clean(r.get("original_lineage_path")))
        if r.get("fixed_final_expression"):
            new_parts.append("Final expression:\n" + clean(r.get("fixed_final_expression")))
        if r.get("fixed_resolved_expression"):
            new_parts.append("Resolved expression:\n" + clean(r.get("fixed_resolved_expression")))
        if r.get("fixed_lineage_path"):
            new_parts.append("Lineage path:\n" + clean(r.get("fixed_lineage_path")))
        out.append({
            "Area / Columns": col,
            "Conclusion": odi_column_conclusion(status),
            "Difference Type": "Resolved multi-step lineage " + status.lower().replace("_", " "),
            "DRD Mapping Logic": drd_logic_by_col.get(col.upper(), ""),
            "ODI File 1 XML Logic": short_text("\n\n".join(old_parts)),
            "ODI File 2 XML Logic": short_text("\n\n".join(new_parts)),
            "Recommended Action": odi_action(status),
        })
    return out


def build_odi_vs_odi_sql(rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    out = []
    for r in rows:
        status = clean(r.get("sql_delta_status", ""))
        if status == "UNCHANGED":
            continue
        area = f"Step {r.get('step_no','')} / Task {r.get('task_no','')}"
        if r.get("task_name"):
            area += f" / {r.get('task_name')}"
        out.append({
            "Area / Columns": clean(area),
            "Conclusion": odi_sql_conclusion(status),
            "Difference Type": "SQL block " + status.lower().replace("_", " "),
            "DRD Mapping Logic": "Not applicable: SQL block-level ODI1 vs ODI2 comparison.",
            "ODI File 1 XML Logic": short_text(r.get("original_sql_excerpt", ""), 12000),
            "ODI File 2 XML Logic": short_text(r.get("fixed_sql_excerpt", ""), 12000),
            "Recommended Action": odi_action(status),
        })
    return out


def build_mismatch_area_map(rows: Sequence[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    out = {}
    for r in rows:
        for c in parse_area_columns(r.get("Area / Columns", "")):
            out[c] = r
    return out


def expression_text_from_impl(r: Dict[str, str]) -> str:
    parts = []
    src = ".".join([x for x in [clean(r.get("source_schema", "")), clean(r.get("source_table", "")), clean(r.get("source_attribute", ""))] if x])
    if src:
        parts.append("DRD source: " + src)
    if clean(r.get("drd_expression", "")):
        parts.append("DRD expression:\n" + clean(r.get("drd_expression", "")))
    if clean(r.get("drd_rule", "")):
        parts.append("DRD transformation rule:\n" + clean(r.get("drd_rule", "")))
    return "\n\n".join(parts)


def generated_text_from_impl(r: Dict[str, str], tri: Dict[str, str]) -> str:
    parts = []
    if clean(r.get("generated_expression", "")):
        parts.append("Generated expression:\n" + clean(r.get("generated_expression", "")))
    if clean(r.get("implementation_status", "")):
        parts.append("Implementation status: " + clean(r.get("implementation_status", "")))
    if clean(r.get("implementation_source", "")):
        parts.append("Implementation source: " + clean(r.get("implementation_source", "")))
    if tri:
        parts.append("Generated vs DRD: " + clean(tri.get("generated_vs_drd", "")))
        parts.append("Generated vs ODI: " + clean(tri.get("generated_vs_odi", "")))
        parts.append("Same mismatch as DRD/ODI: " + clean(tri.get("same_mismatch_as_drd_odi", "")))
    if clean(r.get("notes", "")):
        parts.append("Notes:\n" + clean(r.get("notes", "")))
    return "\n\n".join(parts)


def odi_text(col: str, impl: Dict[str, str], delta: Dict[str, str], fixed_mismatch: Dict[str, Dict[str, str]]) -> str:
    parts = []
    fm = fixed_mismatch.get(col.upper(), {})
    if clean(fm.get("ODI XML Logic", "")):
        parts.append(clean(fm.get("ODI XML Logic", "")))
    if delta:
        if clean(delta.get("fixed_resolved_expression", "")):
            parts.append("ODI File 2 resolved expression:\n" + clean(delta.get("fixed_resolved_expression", "")))
        if clean(delta.get("fixed_lineage_path", "")):
            parts.append("ODI File 2 lineage path:\n" + clean(delta.get("fixed_lineage_path", "")))
        if clean(delta.get("delta_status", "")):
            parts.append("ODI delta status: " + clean(delta.get("delta_status", "")))
    if not parts and clean(impl.get("odi_expression", "")):
        parts.append(clean(impl.get("odi_expression", "")))
    return "\n\n".join(parts)


def is_generated_drd_match_status(status: str) -> bool:
    s = clean(status).upper()
    return s == "DRD_SOURCE" or s.startswith("MATCH") or "DRD_ODI_EQUIVALENCE" in s or "EQUIVALENT" in s


def classify_insert_row(impl: Dict[str, str], tri: Dict[str, str], delta: Dict[str, str]) -> Dict[str, str]:
    gen_vs_drd = clean(tri.get("generated_vs_drd", "")) or "UNKNOWN"
    same = clean(tri.get("same_mismatch_as_drd_odi", "")) or ""
    odi_status = clean(delta.get("delta_status", "")) or clean(impl.get("comparison_class", ""))
    generated_matches = is_generated_drd_match_status(gen_vs_drd) and same in {"Y", ""}

    if not generated_matches:
        conclusion = "Blocker. Generated INSERT does not prove equivalence to the DRD mapping logic."
        action = "Fix insert builder output before using generated SQL as a control implementation."
    elif odi_status == "STILL_OPEN":
        conclusion = "Generated INSERT matches DRD; ODI File 2 still differs from DRD for this mapping item."
        action = "Use generated INSERT as DRD-control candidate. Update ODI File 2 or document an approved exception."
    elif odi_status == "FIX_CANDIDATE_UPSTREAM_CHANGED":
        conclusion = "Generated INSERT matches DRD; ODI File 2 has related upstream changes but still requires validation."
        action = "Validate ODI File 2 upstream logic against DRD and generated INSERT; approve or update ODI."
    elif odi_status == "FIXED_BY_RESOLVED_RULE_PROOF":
        conclusion = "DRD, ODI File 2 resolved logic, and generated INSERT are aligned by rule proof."
        action = "No insert-builder action. Keep this row as positive proof in regression evidence."
    elif odi_status in {"UNCHANGED", "IN_BOTH_NO_REVIEW"}:
        conclusion = "Generated INSERT matches DRD; no active ODI mismatch detected for this column."
        action = "No action."
    else:
        conclusion = "Generated INSERT matches DRD; ODI status requires review."
        action = "Review ODI logic and approve or update ODI as needed."

    return {
        "conclusion": conclusion,
        "difference_type": f"Generated vs DRD: {gen_vs_drd} | ODI status: {odi_status}",
        "recommended_action": action,
    }


def build_insert_report(compare_out: Path, insert_out: Path) -> List[Dict[str, str]]:
    impl_rows = read_csv(insert_out / "implementation_map.csv")
    tri_rows = read_csv(insert_out / "tri_compare_report.csv")
    delta_rows = read_csv(compare_out / "delta_report_v16_6_generic_rule_proof.csv") or read_csv(compare_out / "delta_report_fixed_still_open_regression.csv")
    fixed_mismatch_rows = read_csv(compare_out / "fixed_v15_mismatch_rows.csv")

    impl_by_col = {clean(r.get("target_column", "")).upper(): r for r in impl_rows if clean(r.get("target_column", ""))}
    tri_by_col = {clean(r.get("target_column", "")).upper(): r for r in tri_rows if clean(r.get("target_column", ""))}
    delta_by_col = {clean(r.get("target_column", "")).upper(): r for r in delta_rows if clean(r.get("target_column", ""))}
    fixed_mismatch = build_mismatch_area_map(fixed_mismatch_rows)

    candidate_cols = set()
    for c, r in impl_by_col.items():
        if clean(r.get("comparison_class", "")) == "REVIEW_REQUIRED":
            candidate_cols.add(c)
    for c, d in delta_by_col.items():
        if clean(d.get("delta_status", "")) not in {"", "UNCHANGED"}:
            candidate_cols.add(c)
    for c, t in tri_by_col.items():
        if not is_generated_drd_match_status(clean(t.get("generated_vs_drd", ""))) or clean(t.get("same_mismatch_as_drd_odi", "")) not in {"Y", ""}:
            candidate_cols.add(c)

    out = []
    for col in sorted(candidate_cols):
        impl = impl_by_col.get(col, {})
        tri = tri_by_col.get(col, {})
        delta = delta_by_col.get(col, {})
        cls = classify_insert_row(impl, tri, delta)
        out.append({
            "Area / Columns": col,
            "Conclusion": cls["conclusion"],
            "Difference Type": cls["difference_type"],
            "DRD Mapping Logic": short_text(expression_text_from_impl(impl)),
            "ODI XML Logic": short_text(odi_text(col, impl, delta, fixed_mismatch)),
            "Generated INSERT Logic": short_text(generated_text_from_impl(impl, tri)),
            "Recommended Action": cls["recommended_action"],
        })
    return out


def business_status_from_insert_rows(rows: Sequence[Dict[str, str]]) -> Dict[str, Any]:
    conclusions = Counter(r.get("Conclusion", "") for r in rows)
    has_insert_blocker = any("Blocker" in r.get("Conclusion", "") for r in rows)
    has_review = any(
        phrase in r.get("Conclusion", "")
        for r in rows
        for phrase in ["still differs", "requires validation", "requires review"]
    )
    if has_insert_blocker:
        status = "INSERT_BLOCKER"
    elif has_review:
        status = "REVIEW_REQUIRED"
    else:
        status = "SOLVED_OR_NO_ACTIVE_REVIEW_ROWS"
    return {
        "status": status,
        "meaning": "process success means artifacts were generated; business_status tells whether review items remain",
        "conclusion_counts": dict(conclusions),
    }


def build_step1_tabs(compare_out: Path) -> List[Dict[str, Any]]:
    original_rows = read_csv(compare_out / "original_v15_mismatch_rows.csv")
    fixed_rows = read_csv(compare_out / "fixed_v15_mismatch_rows.csv")
    delta_rows = read_csv(compare_out / "delta_report_v16_6_generic_rule_proof.csv") or read_csv(compare_out / "delta_report_fixed_still_open_regression.csv")
    resolved_delta_rows = read_csv(compare_out / "original_vs_fixed_resolved_xml_delta.csv")
    sql_delta_rows = read_csv(compare_out / "sql_block_differences.csv")

    report1 = normalize_drd_vs_odi_rows(original_rows)
    report2 = normalize_drd_vs_odi_rows(filter_file2_rows(fixed_rows, delta_rows))
    report3 = build_three_way_report(original_rows, fixed_rows, delta_rows)
    drd_logic_by_col = build_drd_logic_map(original_rows + fixed_rows)
    report4 = build_odi_vs_odi_columns(resolved_delta_rows, drd_logic_by_col)
    report5 = build_odi_vs_odi_sql(sql_delta_rows)

    return [
        build_tab("Summary", ["Metric", "Value"], [
            {"Metric": "Mode", "Value": "v18.0 API-first native model"},
            {"Metric": "DRD vs ODI File 1 rows", "Value": str(len(report1))},
            {"Metric": "DRD vs ODI File 2 rows", "Value": str(len(report2))},
            {"Metric": "DRD vs ODI1 vs ODI2 rows", "Value": str(len(report3))},
            {"Metric": "ODI1 vs ODI2 column rows", "Value": str(len(report4))},
            {"Metric": "ODI1 vs ODI2 SQL block rows", "Value": str(len(report5))},
        ]),
        build_tab("DRD vs ODI File 1", DRD_ODI_HEADERS, report1),
        build_tab("DRD vs ODI File 2", DRD_ODI_HEADERS, report2),
        build_tab("DRD vs ODI1 vs ODI2", THREE_WAY_HEADERS, report3),
        build_tab("ODI1 vs ODI2 Columns", ODI_ODI_HEADERS, report4),
        build_tab("ODI1 vs ODI2 SQL Blocks", ODI_ODI_HEADERS, report5),
    ]


def build_step4_tabs(compare_out: Path, insert_out: Path) -> tuple[List[Dict[str, Any]], Dict[str, Any]]:
    tabs = build_step1_tabs(compare_out)
    insert_rows = build_insert_report(compare_out, insert_out)
    tabs.append(build_tab("DRD vs ODI vs INSERT", INSERT_HEADERS, insert_rows))
    business = business_status_from_insert_rows(insert_rows)
    # update summary tab in-place
    summary = tabs[0]
    summary["rows"].extend([
        {"Metric": "DRD vs ODI vs INSERT rows", "Value": str(len(insert_rows))},
        {"Metric": "Business status", "Value": business["status"]},
    ])
    summary["row_count"] = len(summary["rows"])
    return tabs, business


def build_report_payload(report_id: str, report_type: str, tabs: Sequence[Dict[str, Any]], metadata: Dict[str, Any], business_status: Dict[str, Any] | None = None) -> Dict[str, Any]:
    return {
        "report_id": report_id,
        "report_type": report_type,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "engine_version": "18.0",
        "pipeline_mode": "api_first",
        "process_status": "ARTIFACTS_GENERATED",
        "business_status": business_status or {
            "status": "NOT_EVALUATED",
            "meaning": "Step 1 compare only; final business status is produced after INSERT comparison.",
        },
        "metadata": metadata,
        "tab_count": len(tabs),
        "tabs": list(tabs),
    }


def write_payload_and_tabs(api_dir: Path, phase: str, payload: Dict[str, Any]) -> List[str]:
    api_dir.mkdir(parents=True, exist_ok=True)
    files = []
    report_file = api_dir / ("step1_compare_report.json" if phase == "step1" else "step4_full_cycle_report.json")
    write_json(report_file, payload)
    files.append(str(report_file))
    tab_dir = api_dir / phase / "tabs"
    tab_dir.mkdir(parents=True, exist_ok=True)
    for tab in payload.get("tabs", []):
        tab_payload = {
            "report_id": payload["report_id"],
            "report_type": payload["report_type"],
            "phase": phase,
            "tab": tab,
        }
        tab_file = tab_dir / f"{tab['safe_name']}.json"
        write_json(tab_file, tab_payload)
        files.append(str(tab_file))
    return files


def render_excel_from_payload(payload: Dict[str, Any], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Remove default after first sheet handled.
    default = wb.active
    wb.remove(default)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9E2F3")

    for tab in payload.get("tabs", []):
        ws = wb.create_sheet(tab["name"][:31])
        headers = tab.get("headers", [])
        rows = tab.get("rows", [])
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        for cell in ws[1]:
            cell.fill = dark_fill if tab["name"] == "Summary" else header_fill
            cell.font = Font(bold=True, color="FFFFFF" if tab["name"] == "Summary" else "000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        widths = [30, 44, 40, 78, 98, 98, 58]
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = widths[i-1] if i <= len(widths) else 40
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30
        for idx in range(2, min(len(rows) + 2, 302)):
            ws.row_dimensions[idx].height = 105
        if rows and headers:
            ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
            table_name = re.sub(r"[^A-Za-z0-9_]", "_", tab["safe_name"])[:20] + "_T"
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            try:
                ws.add_table(table)
            except Exception:
                pass
    wb.save(path)


def build_openapi_contract() -> Dict[str, Any]:
    return {
        "openapi": "3.0.3",
        "info": {
            "title": "DRD/ODI/Generated INSERT API-first Report API",
            "version": "18.0",
            "description": "Native JSON/API reports. Excel is optional rendering from the same JSON model.",
        },
        "paths": {
            "/manifest": {"get": {"summary": "Get report manifest"}},
            "/step1": {"get": {"summary": "Get Step 1 report JSON"}},
            "/step1/tabs/{tab}": {"get": {"summary": "Get Step 1 tab JSON"}},
            "/step4": {"get": {"summary": "Get Step 4 report JSON"}},
            "/step4/tabs/{tab}": {"get": {"summary": "Get Step 4 tab JSON"}},
            "/health": {"get": {"summary": "Health check"}},
        },
    }


def build_reports(compare_out: Path, insert_out: Path | None, final_dir: Path, report_mode: str, phase: str) -> Dict[str, Any]:
    api_dir = final_dir / "api"
    api_dir.mkdir(parents=True, exist_ok=True)
    files = []

    metadata = {
        "compare_out": str(compare_out),
        "insert_out": str(insert_out) if insert_out else "",
        "final_dir": str(final_dir),
        "report_mode": report_mode,
        "api_first_note": "JSON reports are built directly from raw compare/insert outputs; Excel is optional rendering from JSON.",
    }
    if insert_out:
        metadata["insert_final_consistency_summary"] = read_json(insert_out / "final_consistency_summary.json")
        metadata["insert_hardcode_gate_report"] = read_json(insert_out / "hardcode_gate_report.json")

    step1_tabs = build_step1_tabs(compare_out)
    step1_payload = build_report_payload(
        "step1_compare_report",
        "DRD/ODI comparison",
        step1_tabs,
        metadata,
    )
    files.extend(write_payload_and_tabs(api_dir, "step1", step1_payload))

    step4_payload = None
    if phase in {"step4", "all"} and insert_out:
        step4_tabs, business = build_step4_tabs(compare_out, insert_out)
        step4_payload = build_report_payload(
            "step4_full_cycle_report",
            "DRD/ODI/generated INSERT full-cycle comparison",
            step4_tabs,
            metadata,
            business_status=business,
        )
        files.extend(write_payload_and_tabs(api_dir, "step4", step4_payload))
        # Full-cycle summary is native JSON, not workbook-derived.
        write_json(final_dir / "full_cycle_summary.json", {
            "final_report_json": str(api_dir / "step4_full_cycle_report.json"),
            "insert_tab": "DRD vs ODI vs INSERT",
            "insert_tab_rows": next((t["row_count"] for t in step4_tabs if t["name"] == "DRD vs ODI vs INSERT"), 0),
            "conclusion_counts": business.get("conclusion_counts", {}),
            "business_status": business,
            "pipeline_mode": "api_first",
        })

    if report_mode in {"excel", "both"}:
        render_excel_from_payload(step1_payload, final_dir / "step1_compare_report.xlsx")
        files.append(str(final_dir / "step1_compare_report.xlsx"))
        if step4_payload:
            render_excel_from_payload(step4_payload, final_dir / "step4_full_cycle_report.xlsx")
            render_excel_from_payload(step4_payload, final_dir / "final_full_cycle_report.xlsx")
            files.append(str(final_dir / "step4_full_cycle_report.xlsx"))
            files.append(str(final_dir / "final_full_cycle_report.xlsx"))

    write_json(api_dir / "openapi_contract.json", build_openapi_contract())
    files.append(str(api_dir / "openapi_contract.json"))

    manifest = {
        "version": "18.0",
        "pipeline_mode": "api_first",
        "report_mode": report_mode,
        "phase": phase,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "process_status": "ARTIFACTS_GENERATED",
        "business_status": (step4_payload or step1_payload).get("business_status"),
        "api_dir": str(api_dir),
        "reports": {
            "step1": {
                "json": str(api_dir / "step1_compare_report.json"),
                "tab_count": step1_payload["tab_count"],
                "tabs": [t["safe_name"] for t in step1_payload["tabs"]],
            },
            "step4": {
                "json": str(api_dir / "step4_full_cycle_report.json") if step4_payload else "",
                "tab_count": step4_payload["tab_count"] if step4_payload else 0,
                "tabs": [t["safe_name"] for t in step4_payload["tabs"]] if step4_payload else [],
            },
        },
        "excel": {
            "step1_compare_report": str(final_dir / "step1_compare_report.xlsx") if report_mode in {"excel", "both"} else "",
            "step4_full_cycle_report": str(final_dir / "step4_full_cycle_report.xlsx") if step4_payload and report_mode in {"excel", "both"} else "",
            "final_full_cycle_report": str(final_dir / "final_full_cycle_report.xlsx") if step4_payload and report_mode in {"excel", "both"} else "",
        },
        "openapi_contract": str(api_dir / "openapi_contract.json"),
        "generated_files": files,
    }
    write_json(api_dir / "manifest.json", manifest)
    return manifest


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description="Build v18.0 API-first reports")
    p.add_argument("--compare-out", required=True)
    p.add_argument("--insert-out", default="")
    p.add_argument("--final-dir", required=True)
    p.add_argument("--report-mode", default="both", choices=["excel", "api", "both"])
    p.add_argument("--phase", default="all", choices=["step1", "step4", "all"])
    args = p.parse_args(argv)

    manifest = build_reports(
        compare_out=Path(args.compare_out).expanduser().resolve(),
        insert_out=Path(args.insert_out).expanduser().resolve() if args.insert_out else None,
        final_dir=Path(args.final_dir).expanduser().resolve(),
        report_mode=args.report_mode,
        phase=args.phase,
    )
    print(json.dumps(manifest, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
