#!/usr/bin/env python3
"""
add_generated_insert_tab.py

Adds/refreshes the final full-cycle workbook tab:
  DRD vs ODI vs INSERT

Inputs:
- v16.6.x compare output directory
- v6.2 config-driven insert builder output directory

Output:
- <out>/independent_reports/final_full_cycle_report.xlsx
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from pathlib import Path
from typing import Dict, List, Sequence
from collections import Counter


def clean(x) -> str:
    return (x or '').replace('\r\n','\n').replace('\r','\n').strip()


def short_text(x: str, limit: int = 12000) -> str:
    s = clean(x)
    return s if len(s) <= limit else s[:limit-40] + '\n...[truncated]...'


def read_csv(path: Path) -> List[Dict[str,str]]:
    if not path.exists(): return []
    with path.open(newline='', encoding='utf-8-sig') as f:
        return list(csv.DictReader(f))


def is_generated_drd_match_status(status: str) -> bool:
    s = clean(status).upper()
    return (
        s == 'DRD_SOURCE'
        or s.startswith('MATCH')
        or 'DRD_ODI_EQUIVALENCE' in s
        or 'EQUIVALENT' in s
    )


def parse_area_columns(area: str) -> List[str]:
    area = clean(area)
    cols = re.findall(r'`([^`]+)`', area)
    if not cols and re.fullmatch(r'[A-Z0-9_#$]+', area):
        cols = [area]
    return [c.strip().upper() for c in cols if c.strip()]


def build_mismatch_area_map(rows: Sequence[Dict[str,str]]) -> Dict[str, Dict[str,str]]:
    out = {}
    for r in rows:
        for c in parse_area_columns(r.get('Area / Columns','')):
            out[c] = r
    return out


def expression_text_from_impl(r: Dict[str,str]) -> str:
    parts = []
    src = '.'.join([x for x in [clean(r.get('source_schema','')), clean(r.get('source_table','')), clean(r.get('source_attribute',''))] if x])
    if src:
        parts.append('DRD source: ' + src)
    if clean(r.get('drd_expression','')):
        parts.append('DRD expression:\n' + clean(r.get('drd_expression','')))
    if clean(r.get('drd_rule','')):
        parts.append('DRD transformation rule:\n' + clean(r.get('drd_rule','')))
    return '\n\n'.join(parts)


def generated_text_from_impl(r: Dict[str,str], tri: Dict[str,str]) -> str:
    parts = []
    if clean(r.get('generated_expression','')):
        parts.append('Generated expression:\n' + clean(r.get('generated_expression','')))
    if clean(r.get('implementation_status','')):
        parts.append('Implementation status: ' + clean(r.get('implementation_status','')))
    if clean(r.get('implementation_source','')):
        parts.append('Implementation source: ' + clean(r.get('implementation_source','')))
    if tri:
        parts.append('Generated vs DRD: ' + clean(tri.get('generated_vs_drd','')))
        parts.append('Generated vs ODI: ' + clean(tri.get('generated_vs_odi','')))
        parts.append('Same mismatch as DRD/ODI: ' + clean(tri.get('same_mismatch_as_drd_odi','')))
    if clean(r.get('notes','')):
        parts.append('Notes:\n' + clean(r.get('notes','')))
    return '\n\n'.join(parts)


def odi_text(col: str, impl: Dict[str,str], delta: Dict[str,str], fixed_mismatch: Dict[str,Dict[str,str]]) -> str:
    parts = []
    fm = fixed_mismatch.get(col.upper(), {})
    if clean(fm.get('ODI XML Logic','')):
        parts.append(clean(fm.get('ODI XML Logic','')))
    if delta:
        if clean(delta.get('fixed_resolved_expression','')):
            parts.append('ODI File 2 resolved expression:\n' + clean(delta.get('fixed_resolved_expression','')))
        if clean(delta.get('fixed_lineage_path','')):
            parts.append('ODI File 2 lineage path:\n' + clean(delta.get('fixed_lineage_path','')))
        if clean(delta.get('delta_status','')):
            parts.append('ODI delta status: ' + clean(delta.get('delta_status','')))
    if not parts and clean(impl.get('odi_expression','')):
        parts.append(clean(impl.get('odi_expression','')))
    return '\n\n'.join(parts)


def classify_row(col: str, impl: Dict[str,str], tri: Dict[str,str], delta: Dict[str,str]) -> Dict[str,str]:
    gen_vs_drd = clean(tri.get('generated_vs_drd','')) or 'UNKNOWN'
    same = clean(tri.get('same_mismatch_as_drd_odi','')) or ''
    odi_status = clean(delta.get('delta_status','')) or clean(impl.get('comparison_class',''))

    generated_matches = (is_generated_drd_match_status(gen_vs_drd) and (same in {'Y',''}))

    if not generated_matches:
        conclusion = 'Blocker. Generated INSERT does not prove equivalence to the DRD mapping logic.'
        action = 'Fix insert builder output before using generated SQL as a control implementation.'
    elif odi_status == 'STILL_OPEN':
        conclusion = 'Generated INSERT matches DRD; ODI File 2 still differs from DRD for this mapping item.'
        action = 'Use generated INSERT as DRD-control candidate. Update ODI File 2 or document an approved exception.'
    elif odi_status == 'FIX_CANDIDATE_UPSTREAM_CHANGED':
        conclusion = 'Generated INSERT matches DRD; ODI File 2 has related upstream changes but still requires validation.'
        action = 'Validate ODI File 2 upstream logic against DRD and generated INSERT; approve or update ODI.'
    elif odi_status == 'FIXED_BY_RESOLVED_RULE_PROOF':
        conclusion = 'DRD, ODI File 2 resolved logic, and generated INSERT are aligned by rule proof.'
        action = 'No insert-builder action. Keep this row as positive proof in regression evidence.'
    elif odi_status in {'UNCHANGED', 'IN_BOTH_NO_REVIEW'}:
        conclusion = 'Generated INSERT matches DRD; no active ODI mismatch detected for this column.'
        action = 'No action.'
    else:
        conclusion = 'Generated INSERT matches DRD; ODI status requires review.'
        action = 'Review ODI logic and approve or update ODI as needed.'

    diff = f'Generated vs DRD: {gen_vs_drd} | ODI status: {odi_status}'
    return {'conclusion': conclusion, 'difference_type': diff, 'recommended_action': action}


def build_rows(compare_out: Path, insert_out: Path) -> List[Dict[str,str]]:
    impl_rows = read_csv(insert_out/'implementation_map.csv')
    tri_rows = read_csv(insert_out/'tri_compare_report.csv')
    delta_rows = read_csv(compare_out/'delta_report_v16_6_generic_rule_proof.csv') or read_csv(compare_out/'delta_report_fixed_still_open_regression.csv')
    fixed_mismatch_rows = read_csv(compare_out/'fixed_v15_mismatch_rows.csv')

    impl_by_col = {clean(r.get('target_column','')).upper(): r for r in impl_rows if clean(r.get('target_column',''))}
    tri_by_col = {clean(r.get('target_column','')).upper(): r for r in tri_rows if clean(r.get('target_column',''))}
    delta_by_col = {clean(r.get('target_column','')).upper(): r for r in delta_rows if clean(r.get('target_column',''))}
    fixed_mismatch = build_mismatch_area_map(fixed_mismatch_rows)

    candidate_cols = set()
    for c, r in impl_by_col.items():
        if clean(r.get('comparison_class','')) == 'REVIEW_REQUIRED':
            candidate_cols.add(c)
    for c, d in delta_by_col.items():
        if clean(d.get('delta_status','')) not in {'', 'UNCHANGED'}:
            candidate_cols.add(c)
    for c, t in tri_by_col.items():
        if not is_generated_drd_match_status(clean(t.get('generated_vs_drd',''))) or clean(t.get('same_mismatch_as_drd_odi','')) not in {'Y',''}:
            candidate_cols.add(c)

    rows = []
    for col in sorted(candidate_cols):
        impl = impl_by_col.get(col, {})
        tri = tri_by_col.get(col, {})
        delta = delta_by_col.get(col, {})
        cls = classify_row(col, impl, tri, delta)
        rows.append({
            'Area / Columns': col,
            'Conclusion': cls['conclusion'],
            'Difference Type': cls['difference_type'],
            'DRD Mapping Logic': short_text(expression_text_from_impl(impl)),
            'ODI XML Logic': short_text(odi_text(col, impl, delta, fixed_mismatch)),
            'Generated INSERT Logic': short_text(generated_text_from_impl(impl, tri)),
            'Recommended Action': cls['recommended_action'],
        })
    return rows


def create_or_update_workbook(compare_out: Path, insert_out: Path, final_path: Path) -> Dict[str, object]:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    source_wb = compare_out/'independent_reports'/'independent_reports.xlsx'
    if not source_wb.exists():
        raise FileNotFoundError(f'Base compare workbook not found: {source_wb}')

    final_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_wb, final_path)
    wb = load_workbook(final_path)

    tab_name = 'DRD vs ODI vs INSERT'
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)

    rows = build_rows(compare_out, insert_out)
    headers = ['Area / Columns','Conclusion','Difference Type','DRD Mapping Logic','ODI XML Logic','Generated INSERT Logic','Recommended Action']
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h,'') for h in headers])

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    thin = Side(style='thin', color='D9E2F3')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [30, 44, 44, 88, 98, 98, 58]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30
    for idx in range(2, min(len(rows)+2, 302)):
        ws.row_dimensions[idx].height = 112
    if rows:
        ref = f'A1:G{len(rows)+1}'
        tab = Table(displayName='DRD_ODI_INSERT_T', ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)

    # Update Summary sheet if present.
    if 'Summary' in wb.sheetnames:
        s = wb['Summary']
        start = s.max_row + 2
        s.cell(start, 1, 'Generated INSERT full-cycle rows')
        s.cell(start, 2, len(rows))
        s.cell(start+1, 1, 'Generated INSERT validation tab')
        s.cell(start+1, 2, tab_name)
        for row in s.iter_rows(min_row=start, max_row=start+1, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.save(final_path)

    summary = {
        'final_workbook': str(final_path),
        'insert_tab': tab_name,
        'insert_tab_rows': len(rows),
        'conclusion_counts': dict(Counter(r['Conclusion'] for r in rows)),
    }
    (final_path.parent/'full_cycle_summary.json').write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding='utf-8')
    return summary


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description='Add DRD/ODI/Generated INSERT tab to v16.6 workbook')
    p.add_argument('--compare-out', required=True)
    p.add_argument('--insert-out', required=True)
    p.add_argument('--final-workbook', required=True)
    args = p.parse_args(argv)
    summary = create_or_update_workbook(Path(args.compare_out).resolve(), Path(args.insert_out).resolve(), Path(args.final_workbook).resolve())
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0

if __name__ == '__main__':
    raise SystemExit(main())
