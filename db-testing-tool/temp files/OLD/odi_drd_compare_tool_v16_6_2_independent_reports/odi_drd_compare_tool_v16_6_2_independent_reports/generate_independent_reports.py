#!/usr/bin/env python3
"""
generate_independent_reports.py

Post-processor for v16.6.2.

Always generates three independent reports from a v16.6 compare output directory:

1. DRD vs ODI File 1
2. DRD vs ODI File 2
3. ODI File 1 vs ODI File 2, with DRD Mapping Logic added where available

Outputs:
- independent_reports/01_DRD_vs_ODI_File_1.csv
- independent_reports/02_DRD_vs_ODI_File_2.csv
- independent_reports/03_ODI1_vs_ODI2_Columns_with_DRD_logic.csv
- independent_reports/04_ODI1_vs_ODI2_SQL_Blocks_with_DRD_logic.csv
- independent_reports/summary.json
- independent_reports/AVY_v16_6_2_independent_reports.xlsx (if openpyxl is available)

This script does not compare to v15 semantically. It only reuses legacy-named raw files
emitted by the engine as inputs, then normalizes output wording to independent reports.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Sequence


REPORT_COLUMNS_DRD_ODI = [
    "Area / Columns",
    "Conclusion",
    "Difference Type",
    "Mapping Logic",
    "ODI XML Logic",
    "Recommended Action",
]

REPORT_COLUMNS_ODI_ODI = [
    "Area / Columns",
    "Conclusion",
    "Difference Type",
    "DRD Mapping Logic",
    "ODI File 1 XML Logic",
    "ODI File 2 XML Logic",
    "Recommended Action",
]


def clean(x) -> str:
    return (x or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def short_text(x: str, limit: int = 9000) -> str:
    s = clean(x)
    return s if len(s) <= limit else s[:limit - 40] + "\n...[truncated]..."


def read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: Sequence[Dict[str, str]], headers: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(headers), extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def parse_area_columns(area: str) -> List[str]:
    area = clean(area)
    cols = re.findall(r"`([^`]+)`", area)
    if not cols and re.fullmatch(r"[A-Z0-9_#$]+", area):
        cols = [area]
    return [c.strip().upper() for c in cols if c.strip()]


def build_drd_logic_map(mismatch_rows: Sequence[Dict[str, str]]) -> Dict[str, str]:
    out = {}
    for r in mismatch_rows:
        area = clean(r.get("Area / Columns", ""))
        logic = clean(r.get("Mapping Logic", "") or r.get("DRD Mapping Logic", ""))
        if not area or not logic:
            continue
        for col in parse_area_columns(area):
            out[col] = logic
    return out


def independent_conclusion(row: Dict[str, str], side: str = "file") -> str:
    conclusion = clean(row.get("Conclusion", ""))
    diff = clean(row.get("Difference Type", ""))

    if "missing" in diff.lower() or "not present" in conclusion.lower():
        return "Mismatch. Mapping logic is not implemented in the ODI XML final/resolved lineage."
    if "source drift" in diff.lower():
        return "Review required. ODI source or lineage differs from the mapping rule."
    if "case" in diff.lower():
        return "Review required. CASE/transformation logic differs from the mapping rule."
    if "structural" in diff.lower() or "column count" in diff.lower() or "final target" in clean(row.get("Area / Columns", "")).lower():
        return "Review required. Structural target-load behavior differs from the mapping contract."
    return "Review required. ODI XML logic should be validated against the mapping rule."


def recommended_action(row: Dict[str, str]) -> str:
    diff = clean(row.get("Difference Type", ""))
    if "missing" in diff.lower():
        return "Implement the required mapping logic in ODI or document an approved exception."
    if "source drift" in diff.lower():
        return "Validate source lineage against the mapping rule. Update ODI or document an approved exception."
    if "structural" in diff.lower() or "column count" in diff.lower():
        return "Validate target-load contract and column list. Update ODI/DRD or document an approved exception."
    return "Review and approve, or update ODI to match the mapping rule."


def normalize_drd_vs_odi_rows(rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    out = []
    for r in rows:
        area = clean(r.get("Area / Columns", ""))
        if not area:
            continue
        out.append({
            "Area / Columns": area,
            "Conclusion": independent_conclusion(r),
            "Difference Type": clean(r.get("Difference Type", "")),
            "Mapping Logic": clean(r.get("Mapping Logic", "")),
            "ODI XML Logic": clean(r.get("ODI XML Logic", "")),
            "Recommended Action": recommended_action(r),
        })
    return out


def group_delta_statuses(area: str, delta_by_col: Dict[str, Dict[str, str]]) -> List[str]:
    statuses = []
    for col in parse_area_columns(area):
        if col in delta_by_col:
            statuses.append(clean(delta_by_col[col].get("delta_status", "")))
    return statuses


def filter_file2_rows(rows: Sequence[Dict[str, str]], delta_rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    File 2 standalone: omit groups where v16.6 proved every referenced column fixed/equivalent.
    Keep groups with still-open / candidates / structural review.
    """
    delta_by_col = {clean(r.get("target_column", "")).upper(): r for r in delta_rows if clean(r.get("target_column", ""))}
    omit_statuses = {"FIXED_BY_RESOLVED_RULE_PROOF", "UNCHANGED_NO_ACTIVE_MISMATCH_BY_V16_6"}

    kept = []
    for r in rows:
        cols = parse_area_columns(r.get("Area / Columns", ""))
        statuses = group_delta_statuses(r.get("Area / Columns", ""), delta_by_col)
        if cols and statuses and all(s in omit_statuses for s in statuses):
            continue
        # If no per-column statuses exist, keep structural/grouped row.
        kept.append(r)
    return kept


def odi_column_conclusion(status: str) -> str:
    if status == "CHANGED":
        return "Different ODI implementation detected for the same target column."
    if status == "ADDED_IN_FIXED":
        return "Column exists in ODI File 2 but not in ODI File 1."
    if status == "REMOVED_IN_FIXED":
        return "Column exists in ODI File 1 but not in ODI File 2."
    if status == "NO_RESOLVED_LINEAGE":
        return "Resolved lineage was not available for one or both ODI files."
    return "No difference detected."


def odi_sql_conclusion(status: str) -> str:
    if status == "CHANGED":
        return "SQL block differs between ODI File 1 and ODI File 2."
    if status == "ADDED_IN_FIXED":
        return "SQL block exists in ODI File 2 but not in ODI File 1."
    if status == "REMOVED_IN_FIXED":
        return "SQL block exists in ODI File 1 but not in ODI File 2."
    return "No difference detected."


def odi_action(status: str) -> str:
    if status in {"CHANGED", "ADDED_IN_FIXED", "REMOVED_IN_FIXED", "NO_RESOLVED_LINEAGE"}:
        return "Review whether this ODI change is expected. Validate target output impact and document approval."
    return "No action."


def build_odi_vs_odi_column_report(delta_rows: Sequence[Dict[str, str]], drd_logic_by_col: Dict[str, str]) -> List[Dict[str, str]]:
    out = []
    for r in delta_rows:
        status = clean(r.get("xml_delta_status", ""))
        if status == "UNCHANGED":
            continue
        col = clean(r.get("target_column", ""))
        old_parts, new_parts = [], []
        if r.get("original_final_expression"):
            old_parts.append("Final expression:\n" + clean(r.get("original_final_expression")))
        if r.get("original_resolved_expression"):
            old_parts.append("Resolved expression:\n" + clean(r.get("original_resolved_expression")))
        if r.get("original_lineage_path"):
            old_parts.append("Lineage path:\n" + clean(r.get("original_lineage_path")))
        if r.get("fixed_final_expression"):
            new_parts.append("Final expression:\n" + clean(r.get("fixed_final_expression")))
        if r.get("fixed_resolved_expression"):
            new_parts.append("Resolved expression:\n" + clean(r.get("fixed_resolved_expression")))
        if r.get("fixed_lineage_path"):
            new_parts.append("Lineage path:\n" + clean(r.get("fixed_lineage_path")))
        out.append({
            "Area / Columns": col,
            "Conclusion": odi_column_conclusion(status),
            "Difference Type": "Resolved multi-step lineage " + status.lower().replace("_", " "),
            "DRD Mapping Logic": drd_logic_by_col.get(col.upper(), ""),
            "ODI File 1 XML Logic": short_text("\n\n".join(old_parts)),
            "ODI File 2 XML Logic": short_text("\n\n".join(new_parts)),
            "Recommended Action": odi_action(status),
        })
    return out


def build_odi_vs_odi_sql_report(sql_rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    out = []
    for r in sql_rows:
        status = clean(r.get("sql_delta_status", ""))
        if status == "UNCHANGED":
            continue
        area = f"Step {r.get('step_no','')} / Task {r.get('task_no','')}"
        if r.get("task_name"):
            area += f" / {r.get('task_name')}"
        out.append({
            "Area / Columns": clean(area),
            "Conclusion": odi_sql_conclusion(status),
            "Difference Type": "SQL block " + status.lower().replace("_", " "),
            "DRD Mapping Logic": "Not applicable: SQL block-level ODI1 vs ODI2 comparison.",
            "ODI File 1 XML Logic": short_text(r.get("original_sql_excerpt", ""), 12000),
            "ODI File 2 XML Logic": short_text(r.get("fixed_sql_excerpt", ""), 12000),
            "Recommended Action": odi_action(status),
        })
    return out


def create_xlsx(path: Path, summary: Dict[str, object], report1, report2, report3, report4) -> bool:
    """
    Optional .xlsx generation. Uses openpyxl to avoid requiring artifact_tool for end users.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    def write_table(ws, rows, headers):
        ws.append(list(headers))
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = [30, 42, 32, 70, 90, 90, 48]
        for i, width in enumerate(widths[:len(headers)], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        if rows:
            ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
            table_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:20] + "_T"
            tab = Table(displayName=table_name, ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(tab)

    summary_rows = [
        ["Mode", "v16.6.2 always-generated independent reports"],
        ["DRD vs ODI File 1 rows", len(report1)],
        ["DRD vs ODI File 2 rows", len(report2)],
        ["ODI1 vs ODI2 column rows", len(report3)],
        ["ODI1 vs ODI2 SQL block rows", len(report4)],
        ["Note", "Generated automatically by compare_drd_odi_v16_6_generic_rule_proof.py"],
    ]
    ws.append(["Metric", "Value"])
    for r in summary_rows:
        ws.append(r)
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].font = ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws1 = wb.create_sheet("DRD vs ODI File 1")
    write_table(ws1, report1, REPORT_COLUMNS_DRD_ODI)

    ws2 = wb.create_sheet("DRD vs ODI File 2")
    write_table(ws2, report2, REPORT_COLUMNS_DRD_ODI)

    ws3 = wb.create_sheet("ODI1 vs ODI2 Columns")
    write_table(ws3, report3, REPORT_COLUMNS_ODI_ODI)

    ws4 = wb.create_sheet("ODI1 vs ODI2 SQL Blocks")
    write_table(ws4, report4, REPORT_COLUMNS_ODI_ODI)

    wb.save(path)
    return True


def generate_reports(out_dir: Path) -> Dict[str, object]:
    original_rows = read_csv(out_dir / "original_v15_mismatch_rows.csv")
    fixed_rows = read_csv(out_dir / "fixed_v15_mismatch_rows.csv")
    delta_rows = read_csv(out_dir / "delta_report_v16_6_generic_rule_proof.csv")
    if not delta_rows:
        delta_rows = read_csv(out_dir / "delta_report_fixed_still_open_regression.csv")
    resolved_delta_rows = read_csv(out_dir / "original_vs_fixed_resolved_xml_delta.csv")
    sql_delta_rows = read_csv(out_dir / "sql_block_differences.csv")

    reports_dir = out_dir / "independent_reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    report1 = normalize_drd_vs_odi_rows(original_rows)
    report2 = normalize_drd_vs_odi_rows(filter_file2_rows(fixed_rows, delta_rows))
    drd_logic_by_col = build_drd_logic_map(original_rows + fixed_rows)
    report3 = build_odi_vs_odi_column_report(resolved_delta_rows, drd_logic_by_col)
    report4 = build_odi_vs_odi_sql_report(sql_delta_rows)

    write_csv(reports_dir / "01_DRD_vs_ODI_File_1.csv", report1, REPORT_COLUMNS_DRD_ODI)
    write_csv(reports_dir / "02_DRD_vs_ODI_File_2.csv", report2, REPORT_COLUMNS_DRD_ODI)
    write_csv(reports_dir / "03_ODI1_vs_ODI2_Columns_with_DRD_logic.csv", report3, REPORT_COLUMNS_ODI_ODI)
    write_csv(reports_dir / "04_ODI1_vs_ODI2_SQL_Blocks_with_DRD_logic.csv", report4, REPORT_COLUMNS_ODI_ODI)

    summary = {
        "mode": "v16.6.2 always-generated independent reports",
        "drd_vs_odi_file_1_rows": len(report1),
        "drd_vs_odi_file_2_rows": len(report2),
        "odi1_vs_odi2_column_rows": len(report3),
        "odi1_vs_odi2_sql_block_rows": len(report4),
        "column_lineage_status_counts": dict(Counter(r.get("xml_delta_status", "") for r in resolved_delta_rows)),
        "sql_block_status_counts": dict(Counter(r.get("sql_delta_status", "") for r in sql_delta_rows)),
        "files": [
            "01_DRD_vs_ODI_File_1.csv",
            "02_DRD_vs_ODI_File_2.csv",
            "03_ODI1_vs_ODI2_Columns_with_DRD_logic.csv",
            "04_ODI1_vs_ODI2_SQL_Blocks_with_DRD_logic.csv",
            "independent_reports.xlsx",
            "summary.json",
        ],
    }
    (reports_dir / "summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")

    xlsx_created = create_xlsx(reports_dir / "independent_reports.xlsx", summary, report1, report2, report3, report4)
    summary["xlsx_created"] = xlsx_created
    (reports_dir / "summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    return summary


def main(argv=None):
    p = argparse.ArgumentParser(description="Generate v16.6.2 independent reports from a compare output directory")
    p.add_argument("--out-dir", required=True, help="Output directory produced by the v16.6 compare engine")
    args = p.parse_args(argv)
    summary = generate_reports(Path(args.out_dir).expanduser().resolve())
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
