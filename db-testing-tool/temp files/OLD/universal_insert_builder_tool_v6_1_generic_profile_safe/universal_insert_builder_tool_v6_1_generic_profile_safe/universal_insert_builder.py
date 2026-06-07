#!/usr/bin/env python3
"""
universal_insert_builder.py v6.1

DRD-driven INSERT builder.

Core fix vs v4.1-v4.6:
- Does NOT use odi_final_source O as generated SQL source.
- Does NOT treat ODI Step5/Step6 staging as a physical table to load from.
- DRD controls target columns, expressions, and source/join contract.
- ODI is parsed only for comparison/evidence reports.

The generated SQL is a DRD blueprint. It intentionally flags unresolved/ambiguous joins instead
of inventing fake stage aliases or silently substituting ODI final fields.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
import compare_drd_odi_universal as cmp
try:
    import schema_kb_sql_gate as schema_kb_gate
except Exception:
    schema_kb_gate = None

__VERSION__ = "6.1-generic-profile-safe"

SCHEMA_PREFIXES = {
    "CCAL_REPL_OWNER", "REFERENCE_REPL_OWNER", "TRANSACTIONS_OWNER", "SSDS_DAL_OWNER",
    "CCSI_OWNER", "COMMON_OWNER", "CIRD_OWNER", "PRODUCT_OWNER", "TAXLOT_STG_OWNER",
    "TAXLOTS_OWNER", "SSDS", "GLOBAL", "REFERENCE_OWNER", "REFERENCE_REPL_OWNER"
}
SQL_FUNCS = {
    "TO_DATE", "TO_CHAR", "CAST", "NVL", "COALESCE", "DECODE", "SUBSTR", "TRIM",
    "ROUND", "REGEXP_REPLACE", "REGEXP_SUBSTR", "UPPER", "LOWER", "CASE", "NULLIF"
}


def clean(v) -> str:
    return cmp.clean_text(v)


def norm(v) -> str:
    return cmp.normalize_space(v)


def ident(v) -> str:
    return cmp.normalize_identifier(v)


def write_csv(path: Path, rows: List[Dict[str, str]], fields: List[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def qident(v: str, quote: bool = False) -> str:
    x = ident(v)
    return f'"{x}"' if quote else x


def short(text: str, n: int = 900) -> str:
    text = norm(text)
    return text if len(text) <= n else text[:n - 3] + "..."


def make_alias(schema: str, table: str) -> str:
    t = ident(table)
    # J$TXN is not a clean alias in Oracle SQL; use J_TXN.
    return t.replace("$", "_") or "SRC"


def source_ref(row: Dict[str, str]) -> Tuple[str, str, str, str, str]:
    """Return schema, table, attr, full table ref, preferred alias from DRD source columns.

    Important AVY nuance: some audit/default rows have text accidentally placed in
    source columns, e.g. source_1=LAST_UDT_USR_NM and source_3="Audit column...".
    If there is no source table, do not fabricate alias.column. Treat it as no
    physical source and let the DRD rule/default parser handle it.
    """
    schema = ident(row.get("source_1", ""))
    table = ident(row.get("source_2", ""))
    attr = ident(row.get("source_3", ""))
    target = ident(row.get("target_column", ""))
    raw_attr = clean(row.get("source_3", ""))
    # Some AVY rows have business prose shifted into source_3 and repeat the target
    # column as source_1/source_2. This is not a physical source table.
    if schema and table and schema == table == target and len(raw_attr) > 80:
        return schema, table, "", "", ""
    if schema and table:
        ref = f"{schema}.{table}"
        alias = make_alias(schema, table)
        return schema, table, attr, ref, alias
    if table:
        ref = table
        alias = make_alias(schema, table)
        return schema, table, attr, ref, alias
    # No real table => no real source reference.
    return schema, table, "", "", ""


def split_multiline_attr(value: str) -> List[str]:
    raw = clean(value)
    parts = []
    for p in re.split(r"[\n,/]+", raw):
        x = ident(p)
        if x:
            parts.append(x)
    return parts


def source_expression(row: Dict[str, str]) -> Tuple[str, str, str]:
    schema, table, attr, ref, alias = source_ref(row)
    # If the source attribute cell contains multiple attrs, do not invent a choice.
    attrs = split_multiline_attr(row.get("source_3", ""))
    if len(attrs) > 1:
        return f"/* TODO_COMPLEX_SOURCE_ATTRIBUTE_{qident(row.get('target_column',''))}: {clean(row.get('source_3','')).replace('*/','')} */ NULL", ref, alias
    if attr:
        return f"{alias}.{attr}", ref, alias
    return "NULL", ref, alias


def parse_default(rule: str) -> str:
    r = norm(rule)
    # Do not turn conditional business rules like "if ... then set to 'Y'" into unconditional literals.
    if re.search(r"^\s*if\b", r, re.I) and re.search(r"\bthen\b", r, re.I):
        return ""
    if re.fullmatch(r"-?\d+(?:\.\d+)?", r):
        return r
    for p in [
        r"default\s+to\s+'([^']*)'", r"default\s+to\s+([A-Za-z0-9_.$-]+)",
        r"default\s+([A-Za-z0-9_.$-]+)",
        r"hardcode\s+to\s+'([^']*)'", r"hard\s*code\s+to\s+'([^']*)'",
        r"set\s+to\s+'([^']*)'", r"set\s+to\s+([A-Za-z0-9_.$-]+)",
    ]:
        m = re.search(p, r, re.I)
        if m:
            val = clean(m.group(1))
            vu = val.upper()
            if vu in {"NULL", "NONE"}:
                return "NULL"
            if vu in {"SYSDATE", "CURRENT_DATE", "CURRENT_TIMESTAMP", "USER"}:
                return vu
            if re.fullmatch(r"-?\d+(?:\.\d+)?", val):
                return val
            return "'" + val.replace("'", "''") + "'"
    if r.upper() in {"NULL", "N/A", "NA"} or re.search(r"\balways\s+null\b", r, re.I):
        return "NULL"
    return ""


def strip_alias(expr: str, target: str) -> str:
    e, t = norm(expr), ident(target)
    if t:
        e = re.sub(rf"^\s*{re.escape(t)}\s*,\s*", "", e, flags=re.I)
        e = re.sub(rf"\s+as\s+{re.escape(t)}\s*$", "", e, flags=re.I)
        e = re.sub(rf"\s+{re.escape(t)}\s*$", "", e, flags=re.I)
    return e.strip()


def extract_case(rule: str, target: str) -> str:
    text = clean(rule)
    m = re.search(r"\bcase\b", text, re.I)
    if not m:
        return ""
    start = m.start()
    toks = list(re.finditer(r"\bcase\b|\bend\b", text[start:], re.I))
    depth, end_pos = 0, -1
    for tok in toks:
        if tok.group(0).upper() == "CASE":
            depth += 1
        else:
            depth -= 1
            if depth == 0:
                end_pos = start + tok.end()
                break
    if end_pos < 0:
        return ""
    return strip_alias(text[start:end_pos], target)


def looks_like_raw_prose(rule: str) -> bool:
    r = clean(rule)
    # Any rule that embeds join/from/where plus business prose is not a scalar expression.
    if re.search(r"\b(join|left\s+join|where|populate|lookup|look\s+up|fetch|derive\s+based|first\s+preference|use\s+.+logic|if\s+)\b", r, re.I):
        return True
    return False


def is_scalar_sql_snippet(rule: str) -> bool:
    r = norm(rule)
    if not r:
        return False
    if looks_like_raw_prose(r):
        return False
    return bool(re.search(r"\b(decode|nvl|coalesce|substr|instr|trim|to_date|to_char|round|cast|regexp|concat)\b|[+*/]", r, re.I))


def avy_override_expression(col: str, row: Dict[str, str]) -> Tuple[str, str, str]:
    """Curated executable-ish expressions for known AVY high-risk review rows from DRD.

    These are DRD-driven, not ODI-final CTE substitutions. They still rely on the join graph
    emitted below. CL_VAL-only DRD rows are compiled from the source table + CL_SCM_ID contract.
    """
    col = ident(col)
    # AVY audit/session rows in this DRD have source columns polluted with audit prose.
    # Keep these DRD-driven, but compile them as literals/functions instead of fake aliases.
    if col == "LAST_UDT_USR_NM":
        return "USER", "DRD_AUDIT_DEFAULT", "Compiled AVY audit default from DRD."
    if col == "LAST_UDT_DTM":
        return "SYSDATE", "DRD_AUDIT_DEFAULT", "Compiled AVY audit default from DRD."
    if col == "SESS_NO":
        return "123456", "DRD_SESSION_LITERAL", "DRD provides sample/session default 123456; ODI session variable is helper evidence only."
    if col == "MM_ALT_ID":
        return "CASE WHEN TXN.SRC_STM_ID = 60 AND TXN.ORIG_SRC_STM_CODE LIKE 'MM%' AND TXN.SRC_CRT_USRNM = 'BPMWRAPB' THEN TXN.ORIG_SRC_STM_CODE END", "DRD_AVY_CASE_COMPILED", "Compiled DRD WHERE-like text into attribute-level CASE to preserve TXN grain."
    if col == "BATCH_DT":
        return "J_TXN.BATCH_DT", "DRD_AVY_DIRECT_WITH_JOIN", "DRD says J$TXN joined by TXN_ID; alias normalized to J_TXN."
    if col == "BKR_AR_ID":
        return "CASE WHEN BKR_AR_DIM.ORIG_SRC_STM_CD = 'BKRBO1' THEN TXN.AR_ID WHEN APA.BKR_AR_ID IS NOT NULL THEN APA.BKR_AR_ID ELSE AR_GRP_SUBDIM.LINKED_BKR_AR_ID END", "DRD_AVY_CASE_COMPILED", "Compiled DRD precedence rule."
    if col == "ORIG_SRC_STM_AR_ID":
        return "CASE WHEN TXN.SRC_STM_ID = 5 AND APA.AC_ID IS NOT NULL THEN ORIG_AR_AC_SUBDIM.AC_NUM ELSE ORIG_SRC_STM_AR_DIM.ORIG_SRC_STM_AR_ID END", "DRD_AVY_CONDITIONAL_JOIN_COMPILED", "Compiled DRD RJBNK1/APACSH branch to AR_AC_SUBDIM else AR_DIM."
    if col == "LGCY_TRD_CPCTY_TP_CD":
        return "CASE WHEN TXN.SRC_STM_ID <> 3 THEN LGCY_TRD_CPCTY_TP_DIM.LGCY_TRD_CPCTY_TP_CD WHEN TXN.SRC_STM_ID = 3 THEN LGCY_TRD_CPCTY_CL_VAL.CL_VAL_CODE END", "DRD_AVY_CONDITIONAL_LOOKUP_COMPILED", "Compiled DRD INSTBO1 vs non-INSTBO1 lookup branch."
    if col == "LGCY_TRD_CPCTY_TP_DIM_ID":
        return "LGCY_TRD_CPCTY_TP_DIM.LGCY_TRD_CPCTY_TP_DIM_ID", "DRD_AVY_LOOKUP_DIRECT", "DRD dimension lookup."
    if col == "LGCY_TRD_CPCTY_TP_NM":
        return "CASE WHEN TXN.SRC_STM_ID <> 3 THEN LGCY_TRD_CPCTY_TP_DIM.LGCY_TRD_CPCTY_TP_NM WHEN TXN.SRC_STM_ID = 3 THEN LGCY_TRD_CPCTY_CL_VAL.CL_VAL_NM END", "DRD_AVY_CONDITIONAL_LOOKUP_COMPILED", "Compiled DRD INSTBO1 vs non-INSTBO1 lookup branch."
    if col == "DB_CARD_TXN_DT":
        return "CASE WHEN TXN.SRC_STM_ID = 60 THEN TXN.ORIG_TD END", "DRD_AVY_CASE_COMPILED", "DRD says only ACTDETTD/SRC_STM_ID=60."
    if col == "DB_CARD_ORIG_CCY_CD":
        return "CASE WHEN TXN.SRC_STM_ID = 60 THEN CCY.CCY_CODE END", "DRD_AVY_CASE_COMPILED", "DRD says only ACTDETTD/SRC_STM_ID=60 and CCY lookup via APA."
    if col == "SDIRA_TXN_TP_CD":
        return "CASE WHEN TXN.SRC_STM_ID = 60 THEN SUBSTR(TXN.TRD_NUM, 1, 3) END", "DRD_AVY_PARSE_TYPE_CODE_COMPILED", "DRD says parse SDIRA transaction type code directly from TRADE_NUMBER; the same parsed code keys the CL_VAL lookup used for SDIRA_TXN_TP."
    if col == "SDIRA_TXN_TP":
        return "CASE WHEN TXN.SRC_STM_ID = 60 THEN SDIRA_TXN_TP_CL_VAL.CL_VAL_NM END", "DRD_AVY_PARSE_CL_VAL_LOOKUP_COMPILED", "DRD source table is CL_VAL; parse TRADE_NUMBER to lookup CL_VAL_CODE, then project CL_VAL_NM."
    if col == "SDIRA_TXN_YR":
        return "CASE WHEN TXN.SRC_STM_ID = 60 THEN 2000 + TO_NUMBER(SUBSTR(TXN.TRD_NUM, -2)) END", "DRD_AVY_PARSE_COMPILED", "DRD says last two digits + century."
    if col == "CDSC_AMT":
        return "CASE WHEN TXN.SRC_STM_ID IN (25, 94) AND TXN.TXN_TP_ID IN (14, 1856) AND TXN.BUY_SELL_IND = 'Sell' AND FIP.FIP_TP_ID = 132 THEN CASE WHEN TXN.SRC_STM_ID = 94 THEN FIP.STM_BASE_CCY_AMT WHEN TXN.SRC_STM_ID = 25 THEN FIP.OTHR_FEE END END", "DRD_AVY_FIP_CASE_COMPILED", "Compiled DRD STSR/STSR2 CDSC fee rule."
    if col == "OTHR_FEE_AMT":
        return "CASE WHEN TXN.SRC_STM_ID IN (25, 94) AND FIP.FIP_TP_ID = 132 THEN CASE WHEN TXN.SRC_STM_ID = 94 THEN FIP.STM_BASE_CCY_AMT WHEN TXN.SRC_STM_ID = 25 THEN FIP.OTHR_FEE END END", "DRD_AVY_FIP_CASE_COMPILED", "Compiled DRD STSR/STSR2 other-fee rule."
    if col == "TRD_CNCLD_F":
        return "CASE WHEN EXISTS (SELECT 1 FROM CCAL_REPL_OWNER.TXN_RLTNP WHERE TXN_RLTNP.TXN_RLTNP_TP_ID = 69 AND TXN_RLTNP.TRGT_TXN_ID <> TXN_RLTNP.SRC_TXN_ID AND (TXN_RLTNP.TRGT_TXN_ID = TXN.TXN_ID OR TXN_RLTNP.SRC_TXN_ID = TXN.TXN_ID)) THEN 'Y' END", "DRD_AVY_EXISTS_COMPILED", "Compiled DRD cancellation relationship rule without multiplying TXN rows."
    if col == "STEP_IN_OUT_IND_CD":
        return "STEP_IN_OUT_CL_VAL.CL_VAL_CODE", "DRD_AVY_CL_VAL_BY_SCM", "DRD source table is CL_VAL; rule constrains CL_SCM_ID=114 and selects CL_VAL_CODE."
    if col == "STEP_IN_OUT_IND_NM":
        return "STEP_IN_OUT_CL_VAL.CL_VAL_NM", "DRD_AVY_CL_VAL_BY_SCM", "DRD source table is CL_VAL; rule constrains CL_SCM_ID=114 and selects CL_VAL_NM."
    if col == "SHRT_SALE_EXMPT_CD":
        return "SHRT_SALE_EXMPT_CD_CL_VAL.CL_VAL_CODE", "DRD_AVY_CL_VAL_BY_SCM", "DRD source table is CL_VAL; rule constrains CL_SCM_ID=114 and selects CL_VAL_CODE."
    if col == "SHRT_SALE_EXMPT_NM":
        return "SHRT_SALE_EXMPT_NM_CL_VAL.CL_VAL_NM", "DRD_AVY_CL_VAL_BY_SCM", "DRD source table is CL_VAL; rule constrains CL_SCM_ID=115 and selects CL_VAL_NM."
    return "", "", ""


def drd_expr(row: Dict[str, str], profile: str = "generic") -> Tuple[str, str, str]:
    target, rule = row.get("target_column", ""), clean(row.get("drd_rule", ""))
    if profile == "avy":
        expr, status, notes = avy_override_expression(target, row)
        if expr:
            return expr, status, notes
        expr, status, notes = avy_etl_notes_expression(row)
        if expr:
            return expr, status, notes
    se, _, _ = source_expression(row)
    if not rule and se:
        return se, "DRD_DIRECT_SOURCE", "No DRD rule; used DRD source attribute."
    if not rule and not se:
        return "NULL", "TODO_NO_RULE_NO_SOURCE", "No DRD rule/source."
    d = parse_default(rule)
    if d:
        return d, "DRD_DEFAULT_LITERAL", "Parsed default from DRD."
    c = extract_case(rule, target)
    if c:
        return c, "DRD_CASE_EXTRACTED", "Extracted CASE from DRD."
    if is_scalar_sql_snippet(rule):
        return strip_alias(rule, target), "DRD_SQL_SNIPPET", "Used scalar SQL-like DRD snippet."
    if se and "TODO_COMPLEX_SOURCE_ATTRIBUTE" not in se:
        if re.search(r"\blookup\b|\blook\s+up\b|\bjoin\b|\bcl_val\b|\bdimension\b|\bdim\b", rule, re.I):
            return se, "DRD_SOURCE_REQUIRES_JOIN", "Used DRD source attribute; join must be implemented from AD rule."
        if re.search(r"\buse\b|\bpick\b|\btake\b|\bpopulate\b|\bget\b|\bfetch\b", rule, re.I):
            return se, "DRD_SOURCE_WITH_PROSE", "Used DRD source attribute; prose preserved in reports."
        return se, "DRD_SOURCE_WITH_UNPARSED_RULE", "Used DRD source attribute; prose not parsed."
    return "NULL", "TODO_UNPARSED_RULE", "DRD prose could not be safely converted into scalar SQL."


def load_odi(xml_path: Optional[Path]):
    if not xml_path:
        return [], {}, [], []
    objects = cmp.parse_odi_objects(xml_path)
    xml_targets = cmp.extract_target_resources_from_xml(objects)
    _, _, sql_blocks = cmp.extract_odi_summary(objects)
    lineage = cmp.build_odi_lineage(sql_blocks)
    final = cmp.select_final_target_lineage(lineage)
    by_col = {}
    for r in final:
        col = ident(r.get("target_column", ""))
        if col and col not in by_col:
            by_col[col] = r
    return xml_targets, by_col, final, sql_blocks


def infer_profile(detection, requested: str) -> str:
    if requested != "auto":
        return requested
    blob = (detection.target_table_from_sheet + " " + " ".join(detection.target_resources_from_xml)).upper()
    if "AVY_FACT" in blob:
        return "avy"
    if "TAX_LOT" in blob or "TAXLOTS" in blob:
        return "taxlot"
    return "generic"


def comparison_classes(mapping_rows, odi_by_col, profile):
    column_diff = cmp.compare_columns(mapping_rows, list(odi_by_col.values())) if odi_by_col else []
    by_col = {}
    for r in column_diff:
        col, status = r.get("target_column", ""), r.get("status", "")
        if col:
            by_col[col] = {"class": "MISSING_IN_ODI" if status == "MAPPING_ONLY" else ("ODI_ONLY" if status == "XML_ONLY" else "IN_BOTH_NO_REVIEW"), "reason": status}
    if not odi_by_col:
        for r in mapping_rows:
            by_col[r["target_column"]] = {"class": "DRD_ONLY_NO_ODI", "reason": "No ODI XML helper provided"}
        return by_col, column_diff
    raw = cmp.build_full_drd_vs_odi_xml_rules_diff(column_diff, [])
    mismatches, equivalent = raw, []
    if profile == "avy":
        dummy = type("D", (), {"target_table_from_sheet": "AVY_FACT", "target_resources_from_xml": ["AVY_FACT"]})()
        mismatches = cmp.build_avy_review_rules_diff(column_diff, [], dummy) or raw
    elif profile == "taxlot":
        mismatches, equivalent = cmp.split_mismatch_and_equivalent_rows(raw)

    def cols(area):
        c = re.findall(r"`([^`]+)`", area or "")
        if not c and re.fullmatch(r"[A-Z0-9_#$]+", area or ""):
            c = [area]
        return c
    mapping_by_col = {r.get("target_column", ""): r for r in mapping_rows}
    for r in mismatches:
        dtype = r.get("Difference Type", "")
        for col in cols(r.get("Area / Columns", "")):
            if col not in by_col:
                continue
            if profile == "taxlot" and dtype == "Missing CASE logic" and col in odi_by_col and re.search(r"\bcase\b", mapping_by_col.get(col, {}).get("drd_rule", ""), re.I):
                continue
            by_col[col] = {"class": "REVIEW_REQUIRED", "reason": dtype}
    for r in equivalent:
        for col in cols(r.get("Area / Columns", "")):
            if col in by_col:
                by_col[col] = {"class": "MATCH_EQUIVALENT", "reason": "MATCH_EQUIVALENT"}
    return by_col, column_diff


def expression_aliases(expr: str) -> List[str]:
    aliases = []
    for a in re.findall(r"\b([A-Za-z_][A-Za-z0-9_#$]*)\s*\.", expr or ""):
        au = a.upper()
        if au not in SCHEMA_PREFIXES and au not in SQL_FUNCS:
            aliases.append(au.replace("$", "_"))
    return sorted(set(aliases))


def normalize_rule_aliases(expr: str, primary_alias: str = "", profile: str = "") -> str:
    # Normalize common DRD prose aliases to canonical aliases used by source columns.
    out = expr
    replacements = {
        r"\bT\.": "TXN.",
        r"\bt\.": "TXN.",
        r"\bJT\.": "J_TXN.",
        r"\bjt\.": "J_TXN.",
        r"\bAP\.": "APA.",
        r"\bap\.": "APA.",
        r"\bCV\.": "CL_VAL.",
        r"\bcv\.": "CL_VAL.",
        r"\bASD\.": "AR_GRP_SUBDIM.",
        r"\basd\.": "AR_GRP_SUBDIM.",
    }
    if profile == "taxlot" and primary_alias:
        # DRD often uses conceptual/master names while the physical extracted source is *_TGT.
        for conceptual in ["TAX_LOT_OPN_MSTR", "TAXLOT_DTL_OPN", "SUB_LOT_MSTR", "TAX_LOT_CLS_MSTR", "TAXLOT_DTL_CLS"]:
            out = re.sub(rf"\b{conceptual}\.", primary_alias + ".", out, flags=re.I)
    for pat, repl in replacements.items():
        out = re.sub(pat, repl, out)
    return out




def sql_block_comment(text: str, limit: int = 1200) -> str:
    """Return safe one-line block comment text without accidentally closing it."""
    t = short(text, limit)
    t = t.replace("*/", "* /")
    t = t.replace("/*", "/ *")
    # Keep generated SQL comments visually comments even when DRD prose contains --.
    t = t.replace("--", "- -")
    return t


def strip_sql_comments(sql: str) -> str:
    """Remove block and line comments before executable validation."""
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.S)
    sql = re.sub(r"--[^\n]*", " ", sql)
    return sql


def strip_generated_comments(sql: str) -> str:
    """Executable SQL output: no comments, because some GUI SQL checkers parse comments.

    v5.5 used a normal SQL block-comment regex, but nested/comment-like DRD text
    such as ``/* /* TBC */ */`` could leave an orphan ``*/`` after the
    first pass. The executable variant must be comment-free, so do an
    additional hard scrub of any residual comment delimiters.
    """
    out = strip_sql_comments(sql)
    # Defensive second pass: remove orphan delimiters left by malformed or
    # nested DRD comment fragments. This is intentionally only for the
    # executable SQL; annotated SQL preserves review comments.
    out = out.replace("*/", " ").replace("/*", " ")
    out = re.sub(r"--[^\n]*", " ", out)
    out = re.sub(r"[ \t]+$", "", out, flags=re.M)
    out = re.sub(r"\n{3,}", "\n\n", out)
    return out.strip() + "\n"


def executable_clean_sql_errors(clean_sql: str) -> List[Dict[str, str]]:
    """Hard gate for the final file users paste into Check SQL / Oracle."""
    errors: List[Dict[str, str]] = []
    if "/*" in clean_sql or "*/" in clean_sql or re.search(r"(^|\s)--", clean_sql):
        errors.append({"type": "EXECUTABLE_SQL_COMMENT_MARKER_PRESENT", "target_column": "", "detail": "Executable SQL still contains a SQL comment marker."})
    for i, line in enumerate(clean_sql.splitlines(), 1):
        # DRD prose verbs in rendered SQL are Oracle blockers, even if builder
        # generation otherwise succeeded.
        if re.search(r"\b(populate|derive\s+based|based\s+dimension|manual\s+review|TODO_ON)\b", line, re.I):
            errors.append({"type": "RAW_PROSE_IN_EXECUTABLE_SQL", "target_column": "", "detail": f"line {i}: {short(line, 350)}"})
        if re.search(r"\bON\s+ON\b", line, re.I):
            errors.append({"type": "DUPLICATE_ON_IN_JOIN", "target_column": "", "detail": f"line {i}: {short(line, 350)}"})
        # Owner/table/column existence and datasource-specific names are validated
        # exclusively by schema_kb_sql_gate.py using the user-supplied --schema-kb
        # and optional --resolution-profile. Do not hardcode object/column names
        # here; this executable gate only checks SQL-shape blockers.
    return errors


def sql_without_string_literals(expr: str) -> str:
    return re.sub(r"'([^']|'')*'", "''", expr or "")


def case_end_balance(expr: str) -> Tuple[int, int]:
    scan = sql_without_string_literals(expr)
    return (len(re.findall(r"\bCASE\b", scan, re.I)), len(re.findall(r"\bEND\b", scan, re.I)))


def ensure_case_end(expr: str) -> str:
    c, e = case_end_balance(expr)
    if c > e:
        return expr.rstrip() + (" END" * (c - e))
    return expr


def taxlot_currency_column(primary_alias: str, odi_text: str = "") -> str:
    # ODI is helper evidence only. Use it here only to resolve the physical
    # currency column name when DRD uses shorthand CCY_CODE.
    if odi_text:
        m = re.search(rf"\b{re.escape(primary_alias)}\.([A-Z0-9_]*ISO_CCY_CODE)\b", odi_text, re.I)
        if m:
            return f"{primary_alias}.{m.group(1).upper()}"
    if primary_alias.upper().startswith("OPN_"):
        return f"{primary_alias}.STM_BASE_ISO_CCY_CODE"
    return f"{primary_alias}.NML_ISO_CCY_CODE"


def taxlot_compile_currency_case(expr: str, row: Dict[str, str], primary_alias: str, odi_text: str = "") -> str:
    """Compile DRD shorthand currency CASE into source-qualified Oracle SQL.

    TaxLot DRDs often say CCY_CODE/SBC_EXG_RATE/ORIG_PROCD as business
    shorthand. On the physical source table those names may not exist. The DRD
    source attribute is the amount column; ODI helper can identify whether the
    physical currency column is STM_BASE_ISO_CCY_CODE or NML_ISO_CCY_CODE.
    """
    if not re.search(r"\bCASE\b", expr or "", re.I):
        return expr
    if "SBC_EXG_RATE" not in expr.upper():
        return ensure_case_end(expr)
    schema, table, attr, ref, alias = source_ref(row)
    if not attr:
        return ensure_case_end(expr)
    amount = f"{primary_alias}.{attr}"
    rate = f"{primary_alias}.SBC_EXG_RATE"
    ccy = taxlot_currency_column(primary_alias, odi_text)
    return f"CASE WHEN ({ccy} = 'USD' OR ({rate} IS NULL OR {rate} = 0)) THEN {amount} ELSE {amount} * {rate} END"




def taxlot_rewrite_known_physical_columns(expr: str, primary_alias: str) -> str:
    """No-op in the generic builder.

    Earlier versions rewrote datasource-specific TaxLot conceptual names here
    (for example wash-sale business terms to physical extracted columns). That
    made the pre-KB builder non-generic. In v6.1 these corrections must live in
    the optional --resolution-profile and are applied/validated by
    schema_kb_sql_gate.py against the user-supplied --schema-kb.
    """
    return expr or ""




def resolve_physical_table_ref(table_ref: str, profile: str = '') -> str:
    """Normalize table_ref casing only; do not perform datasource-specific rewrites.

    In v6.1 owner/table existence and any required owner/table corrections are
    delegated to schema_kb_sql_gate.py using --schema-kb and optional
    --resolution-profile. This keeps the pre-KB builder generic.
    """
    ref = (table_ref or '').strip()
    if not ref or '.' not in ref:
        return ref
    schema, table = ref.split('.', 1)
    return f"{ident(schema)}.{ident(table)}"


def unqualified_identifiers_in_case(expr: str) -> List[str]:
    scan = sql_without_string_literals(expr)
    if not re.search(r"\bCASE\b", scan, re.I):
        return []
    keywords = {
        "CASE", "WHEN", "THEN", "ELSE", "END", "NULL", "IS", "NOT", "AND", "OR",
        "IN", "LIKE", "BETWEEN", "EXISTS", "SELECT", "FROM", "WHERE", "AS",
        "DISTINCT", "TRUE", "FALSE", "ON"
    } | SQL_FUNCS
    bad = []
    for m in re.finditer(r"\b([A-Za-z_][A-Za-z0-9_$#]*)\b", scan):
        tok = m.group(1).upper().replace("$", "_")
        before = scan[:m.start()].rstrip()
        after = scan[m.end():].lstrip()
        if tok in keywords or tok in SCHEMA_PREFIXES:
            continue
        if before.endswith(".") or after.startswith("."):
            continue
        if after.startswith("("):
            continue
        bad.append(tok)
    return sorted(set(bad))


def postprocess_expression(expr: str, row: Dict[str, str], profile: str, primary_alias: str, odi_text: str = "") -> str:
    expr = ensure_case_end(expr)
    if profile == "taxlot":
        expr = taxlot_compile_currency_case(expr, row, primary_alias, odi_text)
        expr = taxlot_rewrite_known_physical_columns(expr, primary_alias)
        expr = ensure_case_end(expr)
    return expr


def fix_common_drd_predicate_typos(on: str) -> str:
    """Generic cleanup for DRD prose that is obviously not Oracle predicate syntax."""
    out = on
    out = re.sub(r"\s+AND\s+POPULATE\s+WITH\s+[A-Z0-9_$.]+.*$", "", out, flags=re.I)
    out = re.sub(r"\s+POPULATE\s+WITH\s+[A-Z0-9_$.]+.*$", "", out, flags=re.I)
    out = re.sub(
        r"UPPER\s*\(\s*TRIM\s*\(\s*\(?\s*([A-Za-z_][A-Za-z0-9_$.]*)\s*\)?\s*=\s*([A-Za-z_][A-Za-z0-9_$.]*)",
        r"UPPER(TRIM(\1)) = \2",
        out,
        flags=re.I,
    )
    while out.count(")") > out.count("(") and re.search(r"\)\s*$", out):
        out = re.sub(r"\)\s*$", "", out).rstrip()
    return out


def sanitize_on_clause(on: str) -> str:
    """Clean DRD prose fragments that are not valid SQL ON predicates."""
    out = norm(on)
    if not out:
        return out
    out = normalize_rule_aliases(out, "TXN", "avy")
    out = re.sub(r"\(\+\)", "", out)
    out = re.sub(r"\(\s*--\s*([^)]*)\)", lambda m: " /* " + sql_block_comment(m.group(1), 300) + " */", out)
    out = re.sub(r"\(\s*NOTE\s*:\s*([^)]*)\)", lambda m: " /* NOTE: " + sql_block_comment(m.group(1), 400) + " */", out, flags=re.I)
    out = re.sub(r"\((RJBNK1|RJ\w+|[^()]*only used[^()]*|[^()]*based dimension[^()]*|[^()]*dimension[^()]*)\)", lambda m: " /* " + sql_block_comment(m.group(1), 300) + " */", out, flags=re.I)
    out = re.sub(r"\bWHEN\b", "AND", out, flags=re.I)
    out = re.sub(r"\bTHEN\b.*$", "", out, flags=re.I)
    out = re.sub(r"\s+\buse\b\s+.*$", "", out, flags=re.I)
    out = fix_common_drd_predicate_typos(out)
    out = re.sub(r"\bTBC\b", "/* TBC */", out, flags=re.I)
    out = out.replace(" ND ", " AND ")
    out = re.sub(r"\s+", " ", out).strip()
    # DRD prose sometimes leaves the literal word ON at the beginning of the
    # captured predicate (e.g. alias parsing around "MAP ON ..."). The SQL
    # renderer already emits "... alias ON <predicate>", so a leading ON here
    # becomes the Oracle blocker "ON ON ...". Strip only leading ON tokens.
    out = re.sub(r"^(?:ON\s+)+", "", out, flags=re.I).strip()
    return out


def avy_apa_kind(row_or_rule) -> str:
    """Return APACSH/APASEC when DRD explicitly references ETL Notes APA logic."""
    if isinstance(row_or_rule, dict):
        text = clean(row_or_rule.get("drd_rule", ""))
    else:
        text = clean(row_or_rule)
    u = text.upper()
    # Avoid treating generic prose "First preference is APACASH and then APASEC" as a
    # single source alias; that row needs its own conditional logic.
    if "USE APACSH" in u or "APACSH LOGIC" in u:
        return "APACSH"
    if "USE APASEC" in u or "APASEC LOGIC" in u:
        return "APASEC"
    return ""


def avy_alias_prefix(kind: str) -> str:
    return "CASH" if kind == "APACSH" else "SEC"


def avy_alias_for_table(kind: str, table: str, target_col: str = "") -> str:
    prefix = avy_alias_prefix(kind)
    t = ident(table)
    col = ident(target_col)
    if t == "APA":
        return f"{prefix}_APA"
    if t == "TXN_AVY_CL":
        return f"{prefix}_TXN_AVY_CL"
    if t == "AVY_CL":
        return f"{prefix}_AVY_CL"
    if t in {"NNA_CGY", "NET_NEW_AST_CGY"}:
        return f"{prefix}_NNA_CGY"
    if t == "CL_VAL":
        if "DB_CR" in col:
            return f"{prefix}_DB_CR_CL_VAL"
        if "SALE_CHRG_RATE" in col:
            return f"{prefix}_SALE_CHRG_RATE_TP_CL_VAL"
        return f"{prefix}_APA_TP_CL_VAL"
    if t == "CCAL_CIRD_PD_MAP":
        return f"{prefix}_CCAL_CIRD_PD_MAP"
    if t == "IMT_PD_DIM":
        return f"{prefix}_IMT_PD_DIM"
    return f"{prefix}_{make_alias('', t)}"


def avy_etl_notes_expression(row: Dict[str, str]) -> Tuple[str, str, str]:
    kind = avy_apa_kind(row)
    if not kind:
        return "", "", ""
    schema, table, attr, ref, alias = source_ref(row)
    col = ident(row.get("target_column", ""))
    if not table or not attr:
        return "", "", ""
    src_alias = avy_alias_for_table(kind, table, col)
    return f"{src_alias}.{attr}", "DRD_AVY_ETL_NOTES_SOURCE", f"Compiled {kind} ETL Notes source alias from DRD."


def add_avy_etl_note_joins_for_row(row: Dict[str, str]) -> List[Dict[str, str]]:
    """Curated APACSH/APASEC join group from the DRD ETL Notes contract.

    The group is deliberately aliased separately for cash and security so the same
    physical tables are not declared under the same alias with different predicates.
    """
    kind = avy_apa_kind(row)
    if not kind:
        return []
    prefix = avy_alias_prefix(kind)
    col = ident(row.get("target_column", ""))
    schema, table, attr, ref, source_alias = source_ref(row)
    rule = clean(row.get("drd_rule", ""))
    code_re = "^APACSH[0-7][0-9]" if kind == "APACSH" else "^APASEC[0-7][0-9]"
    rows: List[Dict[str, str]] = []
    def add(jt, table_ref, alias, on, source="DRD_ETL_NOTES_" + kind):
        rows.append({"target_column": col, "join_type": jt, "table_ref": table_ref, "alias": alias, "on_clause": sanitize_on_clause(on), "source": source})
    apa = f"{prefix}_APA"
    apa_cv = f"{prefix}_APA_TP_CL_VAL"
    tac = f"{prefix}_TXN_AVY_CL"
    acl = f"{prefix}_AVY_CL"
    nna = f"{prefix}_NNA_CGY"
    # Base APA branch from ETL Notes. We keep the type decode join separate and
    # filter on the decoded APACSH/APASEC code.
    add("LEFT JOIN", "CCAL_REPL_OWNER.APA", apa, f"{apa}.EXEC_ID = TXN.TXN_ID")
    add("LEFT JOIN", "CCAL_REPL_OWNER.CL_VAL", apa_cv, f"{apa_cv}.CL_VAL_ID = {apa}.APA_TP_ID AND {apa_cv}.CL_SCM_ID = 7 AND REGEXP_LIKE({apa_cv}.CL_VAL_CODE, '{code_re}')")
    needs_tac = table in {"TXN_AVY_CL", "AVY_CL", "NNA_CGY", "NET_NEW_AST_CGY"} or re.search(r"\btxn_avy_cl\b|\btac\.", rule, re.I)
    needs_acl = table == "AVY_CL" or re.search(r"\bavy_cl\b|\bacl\.", rule, re.I)
    needs_nna = table in {"NNA_CGY", "NET_NEW_AST_CGY"} or re.search(r"net_new\s*_?\s*ast\s*_?\s*cgy|\bnna\.", rule, re.I)
    if needs_tac or needs_acl or needs_nna:
        add("LEFT JOIN", "CCAL_REPL_OWNER.TXN_AVY_CL", tac, f"{tac}.TXN_ID = TXN.TXN_ID AND {tac}.ACTV_F = 'Y' AND {apa}.APA_ID = {tac}.APA_ID")
    if needs_acl:
        add("LEFT JOIN", "CCAL_REPL_OWNER.AVY_CL", acl, f"{acl}.AVY_CL_ID = {tac}.AVY_CL_ID")
    if needs_nna:
        # DRD source table is NNA_CGY while prose sometimes says NET_NEW_AST_CGY.
        # Use DRD source contract for the physical table/columns.
        add("LEFT JOIN", "CCAL_REPL_OWNER.NNA_CGY", nna, f"{nna}.NNA_CGY_ID = {tac}.NNA_CGY_ID")
    if table == "CCAL_CIRD_PD_MAP" or re.search(r"ccal_cird_pd_map|ccpd|ccpm", rule, re.I):
        ccpd = f"{prefix}_CCAL_CIRD_PD_MAP"
        add("LEFT JOIN", "CCAL_REPL_OWNER.CCAL_CIRD_PD_MAP", ccpd, f"{ccpd}.CCAL_PD_ID = {apa}.PD_ID AND {ccpd}.ACTV_F = 'Y'")
    if table == "IMT_PD_DIM" or re.search(r"imt_pd_dim|\bipd\.", rule, re.I):
        ipd = f"{prefix}_IMT_PD_DIM"
        add("LEFT JOIN", "CIRD_OWNER.IMT_PD_DIM", ipd, f"{ipd}.CCAL_PD_ID = {apa}.PD_ID AND {ipd}.EFF_DT <= TXN.TD AND {ipd}.END_DT > TXN.TD")
    if table == "CL_VAL" and "DB_CR" in col:
        dbcv = f"{prefix}_DB_CR_CL_VAL"
        add("LEFT JOIN", "CCAL_REPL_OWNER.CL_VAL", dbcv, f"{dbcv}.CL_VAL_ID = {apa}.DB_CR_ID AND {dbcv}.CL_SCM_ID = 15")
    if table == "CL_VAL" and "SALE_CHRG_RATE" in col:
        sccv = f"{prefix}_SALE_CHRG_RATE_TP_CL_VAL"
        add("LEFT JOIN", "CCAL_REPL_OWNER.CL_VAL", sccv, f"{sccv}.CL_VAL_ID = {apa}.SALE_CHRG_RATE_TP_ID AND {sccv}.CL_SCM_ID = 73")
    return rows

def table_alias_from_rule_token(alias: str) -> str:
    a = ident(alias).replace("$", "_")
    mapping = {"T": "TXN", "JT": "J_TXN", "J$TXN": "J_TXN", "AP": "APA", "CV": "CL_VAL", "ASD": "AR_GRP_SUBDIM", "SS": "SRC_STM_DIM", "FN": "FA_NUMBER_V", "E": "ENTERPRISE_ENTITY_DIM_V"}
    return mapping.get(a, a)


def extract_join_candidates_from_rule(rule: str, target_col: str) -> List[Dict[str, str]]:
    """Best-effort join extraction from DRD AD prose.

    It intentionally keeps ambiguous ON clauses as TODO instead of fabricating certainty.
    """
    text = clean(rule)
    if not text:
        return []
    flat = re.sub(r"\s+", " ", text.replace(";", " ")).strip()
    rows = []
    # Find explicit JOIN <schema.table> [alias] ON ... until next join/where/if.
    pat = re.compile(r"\b((?:left\s+outer\s+join|left\s+join|inner\s+join|join))\s+([A-Za-z0-9_$#]+\.[A-Za-z0-9_$#]+)(?:\s+(?:as\s+)?([A-Za-z][A-Za-z0-9_$#]*))?\s+on\s+(.*?)(?=\b(?:left\s+outer\s+join|left\s+join|inner\s+join|join|where|if|else|then)\b|$)", re.I)
    for m in pat.finditer(flat):
        join_type = m.group(1).upper()
        table_ref = f"{ident(m.group(2).split('.')[0])}.{ident(m.group(2).split('.')[1])}"
        raw_alias = m.group(3) or ""
        raw_on = m.group(4).strip()
        if not raw_alias:
            own_aliases = [a for a in re.findall(r"\b([A-Za-z][A-Za-z0-9_$#]*)\s*\.", raw_on)
                           if table_alias_from_rule_token(a).upper() not in {"TXN", "APA", "J_TXN", "CL_VAL", "AR_GRP_SUBDIM"}]
            raw_alias = own_aliases[0] if own_aliases else m.group(2).split(".")[-1]
        alias = table_alias_from_rule_token(raw_alias or m.group(2).split(".")[-1])
        on = normalize_rule_aliases(raw_on)
        on = re.sub(r"\(\+\)", "", on)
        rows.append({"target_column": target_col, "join_type": "LEFT JOIN" if "LEFT" in join_type else "JOIN", "table_ref": table_ref, "alias": alias, "rule_alias": raw_alias, "on_clause": on or "/* TODO_ON_FROM_DRD */ 1=1", "source": "DRD_AD_RULE"})
    return rows


def extra_avy_join_rows(target_col: str) -> List[Dict[str, str]]:
    col = ident(target_col)
    rows = []
    def add(jt, table_ref, alias, on, source="AVY_CURATED_DRD_JOIN"):
        rows.append({"target_column": col, "join_type": jt, "table_ref": table_ref, "alias": alias, "on_clause": on, "source": source})
    if col == "BATCH_DT":
        add("LEFT JOIN", "CCAL_REPL_OWNER.J$TXN", "J_TXN", "J_TXN.TXN_ID = TXN.TXN_ID")
    if col in {"BKR_AR_ID", "BKR_IRA_F", "BKR_ERISA_F"}:
        add("LEFT JOIN", "CCSI_OWNER.AR_DIM", "BKR_AR_DIM", "BKR_AR_DIM.AR_ID = TXN.AR_ID AND TXN.TD >= BKR_AR_DIM.EFF_DT AND TXN.TD < BKR_AR_DIM.END_DT")
        add("LEFT JOIN", "CCSI_OWNER.AR_GRP_SUBDIM", "AR_GRP_SUBDIM", "AR_GRP_SUBDIM.AR_ID = TXN.AR_ID AND TXN.TD >= AR_GRP_SUBDIM.EFF_DT AND TXN.TD < AR_GRP_SUBDIM.END_DT")
        add("LEFT JOIN", "CCAL_REPL_OWNER.APA", "APA", "APA.EXEC_ID = TXN.TXN_ID")
    if col in {"DB_CARD_ORIG_CCY_CD", "DB_CARD_ORIG_CCY"}:
        add("LEFT JOIN", "CCAL_REPL_OWNER.APA", "APA", "APA.EXEC_ID = TXN.TXN_ID")
        add("LEFT JOIN", "REFERENCE_REPL_OWNER.CCY", "CCY", "APA.ORIG_CCY_ID = CCY.CCY_ISO_NUM_CODE")
    if col == "ORIG_SRC_STM_AR_ID":
        add("LEFT JOIN", "CCAL_REPL_OWNER.APA", "APA", "APA.EXEC_ID = TXN.TXN_ID")
        add("LEFT JOIN", "CCSI_OWNER.AR_AC_SUBDIM", "ORIG_AR_AC_SUBDIM", "ORIG_AR_AC_SUBDIM.AR_ID = TXN.AR_ID AND ORIG_AR_AC_SUBDIM.DEP_AC_SETUP_ID = APA.AC_ID AND ORIG_AR_AC_SUBDIM.EFF_DT <= TXN.TD AND ORIG_AR_AC_SUBDIM.END_DT > TXN.TD")
        add("LEFT JOIN", "CCSI_OWNER.AR_DIM", "ORIG_SRC_STM_AR_DIM", "ORIG_SRC_STM_AR_DIM.AR_ID = TXN.AR_ID AND ORIG_SRC_STM_AR_DIM.EFF_DT <= TXN.TD AND ORIG_SRC_STM_AR_DIM.END_DT > TXN.TD")
    if col in {"ORIG_SRC_STM_AR_ID", "LGCY_TRD_CPCTY_TP_CD", "LGCY_TRD_CPCTY_TP_DIM_ID", "LGCY_TRD_CPCTY_TP_NM"}:
        add("LEFT JOIN", "TRANSACTIONS_OWNER.LGCY_TRD_CPCTY_TP_DIM", "LGCY_TRD_CPCTY_TP_DIM", "TXN.LGCY_TRD_CPCTY_TP_ID = LGCY_TRD_CPCTY_TP_DIM.LGCY_TRD_CPCTY_TP_ID")
    if col in {"LGCY_TRD_CPCTY_TP_CD", "LGCY_TRD_CPCTY_TP_NM"}:
        add("LEFT JOIN", "CCAL_REPL_OWNER.CL_VAL", "LGCY_TRD_CPCTY_CL_VAL", "LGCY_TRD_CPCTY_CL_VAL.CL_VAL_ID = TXN.LGCY_TRD_CPCTY_TP_ID")
    if col in {"CDSC_AMT", "OTHR_FEE_AMT"}:
        add("LEFT JOIN", "CCAL_REPL_OWNER.APA", "APA", "APA.EXEC_ID = TXN.TXN_ID")
        add("LEFT JOIN", "CCAL_REPL_OWNER.FIP", "FIP", "FIP.APA_ID = APA.APA_ID")
    if col in {"SDIRA_TXN_TP_CD", "SDIRA_TXN_TP"}:
        add("LEFT JOIN", "CCAL_REPL_OWNER.CL_VAL", "SDIRA_TXN_TP_CL_VAL", "SDIRA_TXN_TP_CL_VAL.CL_VAL_CODE = SUBSTR(TXN.TRD_NUM, 1, 3) AND TXN.SRC_STM_ID = 60", "DRD_AVY_PARSE_CL_VAL_LOOKUP")
    if col in {"STEP_IN_OUT_IND_CD", "STEP_IN_OUT_IND_NM"}:
        add("LEFT JOIN", "CCAL_REPL_OWNER.CL_VAL", "STEP_IN_OUT_CL_VAL", "STEP_IN_OUT_CL_VAL.CL_SCM_ID = 114", "DRD_CL_VAL_SOURCE_RULE")
    if col == "SHRT_SALE_EXMPT_CD":
        add("LEFT JOIN", "CCAL_REPL_OWNER.CL_VAL", "SHRT_SALE_EXMPT_CD_CL_VAL", "SHRT_SALE_EXMPT_CD_CL_VAL.CL_SCM_ID = 114", "DRD_CL_VAL_SOURCE_RULE")
    if col == "SHRT_SALE_EXMPT_NM":
        add("LEFT JOIN", "CCAL_REPL_OWNER.CL_VAL", "SHRT_SALE_EXMPT_NM_CL_VAL", "SHRT_SALE_EXMPT_NM_CL_VAL.CL_SCM_ID = 115", "DRD_CL_VAL_SOURCE_RULE")
    return rows





def infer_avy_source_join(row: Dict[str, str], primary_alias: str) -> str:
    """Small set of deterministic AVY joins that are explicit in DRD prose/notes."""
    col = ident(row.get("target_column", ""))
    table = ident(row.get("source_2", ""))
    alias = make_alias("", table)
    if table == "SHDW_TXN_TP":
        return f"{alias}.SRC_TXN_TP = {primary_alias}.SRC_TXN_TP"
    if table == "IMPCT_ACTION_LKU":
        return f"{alias}.ACTION_CODE = {primary_alias}.SRC_ACTN_CODE"
    if table == "TXN_SRC_TAX_CODE_LKUP":
        return f"{alias}.SRC_TAX_CODE_ID = {primary_alias}.SRC_TAX_CODE_ID AND {alias}.ACTV_F = 'Y'"
    if table == "PERSON_RV":
        return f"{alias}.EMPLOYEE_ID = FA_NUMBER_V1.RESPONSIBLE_PARTY_EMPLOYEE_ID AND {primary_alias}.TD >= {alias}.EFFECTIVE_DATE AND {primary_alias}.TD < {alias}.END_DATE"
    if table == "ENTERPRISE_ENTITY_RISK_DIM":
        return f"{alias}.ENTITY_ENTERPRISE_ID = OWN_FA_ENT.ENTITY_ENTERPRISE_ID AND {primary_alias}.TD >= {alias}.EFFECTIVE_DATE AND {primary_alias}.TD < {alias}.END_DATE"
    return ""

def infer_taxlot_source_join(row: Dict[str, str], primary_alias: str) -> str:
    schema, table, attr, ref, alias = source_ref(row)
    table_u = table.upper()
    rule_u = clean(row.get("drd_rule", "")).upper()
    # TaxLot known lookup dimensions from DRD + ODI helper evidence. These are not final-step O substitutions.
    if table_u == "SRC_STM_DIM":
        if " AS 6" in rule_u or primary_alias.upper().startswith("CLOSE_TAX_LOT"):
            return "SRC_STM_DIM.SRC_STM_ID = 6"
        return f"{primary_alias}.SRC_STM_ID = SRC_STM_DIM.SRC_STM_ID"
    if table_u == "ACG_TP_DIM":
        src_col = "ACG_TP_CODE" if primary_alias.upper().startswith("CLOSE_TAX_LOT") else "AC_TP_CODE"
        return f"ACG_TP_DIM.ACG_TP_CD = {primary_alias}.{src_col}"
    if table_u == "CCAL_CIRD_PD_MAP":
        return f"CCAL_CIRD_PD_MAP.CCAL_PD_ID = {primary_alias}.CCAL_PD_ID AND CCAL_CIRD_PD_MAP.ACTV_F = 'Y'"
    return ""


def infer_cl_val_on(row: Dict[str, str], primary_alias: str) -> str:
    col = ident(row.get("target_column", ""))
    rule = clean(row.get("drd_rule", ""))
    ru = rule.upper()
    # Find the source id mentioned by DRD. Fall back to target-specific TaxLot conventions.
    candidates = re.findall(r"\b([A-Z][A-Z0-9_]*_ID)\b", ru)
    skip = {"CL_VAL_ID", "CL_SCM_ID"}
    source_id = ""
    for c in candidates:
        if c not in skip:
            source_id = c
            break
    fallback = {
        "TAX_LOT_TP": "TAX_LOT_TXN_TP_ID",
        "OPN_TXN_EV_TP": "TAX_LOT_TXN_EV_TP_ID",
        "CLS_TXN_EV_TP": "SUB_LOT_TXN_EV_TP_ID",
        "SRC_RCRD_TP_CD": "SRC_RCRD_TP_ID",
        "TXN_TP_CD": "SUB_LOT_TXN_TP_ID",
        "DRVD_TRD_CPCTY_CD": "DRVD_TRD_CPCTY_TP_ID",
        "DRVD_TRD_CPCTY_NM": "DRVD_TRD_CPCTY_TP_ID",
    }
    source_id = source_id or fallback.get(col, "")
    scm = ""
    m = re.search(r"CL_SCM_(?:ID|CD)\s*=\s*'?([0-9]+)'?", ru)
    if m:
        scm = m.group(1)
    scm_fallback = {"TAX_LOT_TP": "84", "OPN_TXN_EV_TP": "86", "CLS_TXN_EV_TP": "86", "DRVD_TRD_CPCTY_CD": "104", "DRVD_TRD_CPCTY_NM": "104"}
    scm = scm or scm_fallback.get(col, "")
    alias = f"{col}_CL_VAL"
    if not source_id:
        return f"/* TODO_ON_FROM_DRD_{col}: CL_VAL lookup source id is not specified */ 1=1"
    on = f"{alias}.CL_VAL_ID = {primary_alias}.{source_id}"
    if scm:
        on += f" AND {alias}.CL_SCM_ID = {scm}"
    return on

def replace_alias_token(sql: str, old_alias: str, new_alias: str) -> str:
    old_alias = ident(old_alias).replace("$", "_")
    new_alias = ident(new_alias).replace("$", "_")
    if not old_alias or not new_alias or old_alias.upper() == new_alias.upper():
        return sql
    return re.sub(rf"\b{re.escape(old_alias)}\.", new_alias + ".", sql, flags=re.I)


def canonicalize_join_aliases_for_row(row: Dict[str, str], joins: List[Dict[str, str]], primary_alias: str) -> List[Dict[str, str]]:
    """Use canonical DRD source aliases in parsed joins.

    DRD prose often writes short aliases such as b/atd/orig, while generated
    expressions use canonical aliases from source table names such as
    ACATS_BROKER/ACG_TP_DIM/CCY. If we keep the short alias, SQL compiles badly
    because SELECT references a different alias than JOIN declares.
    """
    schema, table, attr, ref, source_alias = source_ref(row)
    out = []
    for j in joins:
        jj = dict(j)
        old_alias = jj.get("alias", "")
        same_physical = ref and (jj.get("table_ref", "").upper() == ref.upper() or jj.get("table_ref", "").split(".")[-1].upper() == table.upper())
        if same_physical and source_alias:
            jj["alias"] = source_alias
            for candidate in [old_alias, jj.get("rule_alias", "")]:
                jj["on_clause"] = replace_alias_token(jj.get("on_clause", ""), candidate, source_alias)
            for candidate in re.findall(r"\b([A-Za-z][A-Za-z0-9_$#]*)\s*\.", jj.get("on_clause", "")):
                cu = table_alias_from_rule_token(candidate).upper()
                protected = {primary_alias.upper(), "TXN", "APA", "CASH_APA", "SEC_APA", "J_TXN", "AR_GRP_SUBDIM", "BKR_AR_DIM", "FA_NUMBER_V1", "FA_NUMBER_V2", source_alias.upper()}
                if cu not in protected and (len(cu) <= 4 or cu.startswith("BRK_OWN_")):
                    jj["on_clause"] = replace_alias_token(jj.get("on_clause", ""), candidate, source_alias)
            jj["source"] = jj.get("source", "") + "+CANONICAL_ALIAS"
        # Normalize primary short alias tokens in ON clauses too.
        jj["on_clause"] = sanitize_on_clause(normalize_rule_aliases(jj.get("on_clause", ""), primary_alias, "avy"))
        out.append(jj)
    return out


def prune_todo_joins_when_alias_declared(joins: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """Remove per-column TODO joins when a reliable global join already declares the same alias/table.

    Example: many AVY APA columns do not repeat the join in every row, but other
    DRD rows provide `JOIN CCAL_REPL_OWNER.APA APA ON APA.EXEC_ID = TXN.TXN_ID`.
    Keeping one reliable join and 40 APA TODO duplicates is validation noise.
    """
    reliable_aliases = set()
    reliable_tables = set()
    for j in joins:
        on = j.get("on_clause", "")
        if "TODO_ON" not in on and "TODO:" not in on:
            if j.get("alias"):
                reliable_aliases.add(j["alias"].upper())
            if j.get("table_ref"):
                reliable_tables.add(j["table_ref"].upper())
    pruned = []
    for j in joins:
        on = j.get("on_clause", "")
        if ("TODO_ON" in on or "TODO:" in on) and j.get("alias", "").upper() in reliable_aliases:
            continue
        pruned.append(j)
    return pruned




def split_join_targets(targets: str) -> List[str]:
    return [ident(x) for x in re.split(r"\s*\|\s*", targets or "") if ident(x)]


def alias_collision_signature(j: Dict[str, str]) -> Tuple[str, str]:
    return (j.get("table_ref", "").upper(), norm(j.get("on_clause", "")).upper())


def merge_exact_joins(joins: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """Merge duplicate aliases/table/ON even when one row says JOIN and another LEFT JOIN."""
    merged: Dict[Tuple[str, str, str], Dict[str, str]] = {}
    for j in joins:
        key = (j.get("table_ref", "").upper(), j.get("alias", "").upper(), norm(j.get("on_clause", "")).upper())
        if key not in merged:
            merged[key] = dict(j)
        else:
            old = set(filter(None, merged[key].get("target_column", "").split(" | ")))
            old.add(j.get("target_column", ""))
            merged[key]["target_column"] = " | ".join(sorted(old))
            if "LEFT" in j.get("join_type", "").upper() or "LEFT" in merged[key].get("join_type", "").upper():
                merged[key]["join_type"] = "LEFT JOIN"
            merged[key]["source"] = (merged[key].get("source", "") + "+MERGED_DUP").strip("+")
    return list(merged.values())


def resolve_join_alias_collisions(joins: List[Dict[str, str]], impl_rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """Ensure every SQL table alias is declared once per different logical role.

    If the same DRD alias appears with multiple table/ON signatures, suffix only the
    later distinct signatures and reuse that suffix for repeated exact signatures.
    """
    by_target = {ident(r.get("target_column", "")): r for r in impl_rows}
    primary_sig: Dict[str, Tuple[str, str]] = {}
    assigned: Dict[Tuple[str, Tuple[str, str]], str] = {}
    counters: Dict[str, int] = defaultdict(int)
    out: List[Dict[str, str]] = []
    for j in joins:
        jj = dict(j)
        alias = ident(jj.get("alias", "")).replace("$", "_")
        if not alias:
            out.append(jj)
            continue
        sig = alias_collision_signature(jj)
        key = (alias, sig)
        if key in assigned:
            new_alias = assigned[key]
        elif alias not in primary_sig:
            primary_sig[alias] = sig
            assigned[key] = alias
            new_alias = alias
        elif primary_sig[alias] == sig:
            assigned[key] = alias
            new_alias = alias
        else:
            counters[alias] += 1
            new_alias = f"{alias}_{counters[alias] + 1}"
            assigned[key] = new_alias
        if new_alias != alias:
            jj["alias"] = new_alias
            jj["on_clause"] = replace_alias_token(jj.get("on_clause", ""), alias, new_alias)
            jj["source"] = (jj.get("source", "") + "+ALIAS_DEDUP").strip("+")
            for tc in split_join_targets(jj.get("target_column", "")):
                rec = by_target.get(tc)
                if not rec:
                    continue
                for field in ("generated_expression", "drd_expression"):
                    rec[field] = replace_alias_token(rec.get(field, ""), alias, new_alias)
        # Final pass for DRD ownership short aliases that were used only in predicates.
        if jj.get("alias", "").upper().startswith("OWN_FA_ENT"):
            jj["on_clause"] = replace_alias_token(jj.get("on_clause", ""), "BRK_OWN_FA_ENT", jj.get("alias", ""))
        out.append(jj)
    return out

def build_join_graph(mapping_rows: List[Dict[str, str]], impl_rows: List[Dict[str, str]], profile: str, primary_ref: str, primary_alias: str) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    """Return concrete join rows and unresolved alias rows."""
    joins: List[Dict[str, str]] = []
    unresolved: List[Dict[str, str]] = []

    impl_by_col = {ident(r.get("target_column", "")): r for r in impl_rows}
    # Source table joins from DRD source columns. If no AD rule says how to join, flag TODO.
    for row in mapping_rows:
        col = row.get("target_column", "")
        impl = impl_by_col.get(ident(col), {})
        needed_aliases = set(expression_aliases(impl.get("generated_expression", ""))) - {primary_alias.upper()}
        if not needed_aliases:
            continue
        schema, table, attr, ref, alias = source_ref(row)
        if profile == "avy" and avy_apa_kind(row):
            # APACSH/APASEC rows are governed by the ETL Notes branch. Do not add
            # generic APA/TXN_AVY_CL/AVY_CL/CL_VAL aliases, otherwise cash and
            # security branches collide and the SQL becomes semantically wrong.
            joins.extend(add_avy_etl_note_joins_for_row(row))
            continue
        if profile == "avy" and ident(col) in {"ORIG_SRC_STM_AR_ID", "LGCY_TRD_CPCTY_TP_CD", "LGCY_TRD_CPCTY_TP_DIM_ID", "LGCY_TRD_CPCTY_TP_NM", "CDSC_AMT", "OTHR_FEE_AMT", "SDIRA_TXN_TP_CD", "SDIRA_TXN_TP", "STEP_IN_OUT_IND_CD", "STEP_IN_OUT_IND_NM", "SHRT_SALE_EXMPT_CD", "SHRT_SALE_EXMPT_NM"}:
            # These rows are handled by explicit AVY compiled joins from extra_avy_join_rows;
            # avoid adding generic parsed aliases such as CL/ltctd or TODO joins.
            continue
        if not ref or not table:
            continue
        if ref.upper() == primary_ref.upper() or alias.upper() == primary_alias.upper():
            continue
        rule = row.get("drd_rule", "")
        # CL_VAL must be target-specific; one shared CL_VAL alias is usually wrong.
        if alias.upper() == "CL_VAL":
            cl_alias = f"{col}_CL_VAL"
            joins.append({"target_column": col, "join_type": "LEFT JOIN", "table_ref": "CCAL_REPL_OWNER.CL_VAL", "alias": cl_alias, "on_clause": infer_cl_val_on(row, primary_alias), "source": "DRD_CL_VAL_LOOKUP"})
            continue
        parsed = canonicalize_join_aliases_for_row(row, extract_join_candidates_from_rule(rule, col), primary_alias)
        # If parsed already covers this alias/table, don't add a TODO duplicate.
        covered = any(j["alias"].upper() == alias.upper() or j["table_ref"].upper() == ref.upper() for j in parsed)
        if not covered:
            inferred_on = infer_taxlot_source_join(row, primary_alias) if profile == "taxlot" else (infer_avy_source_join(row, primary_alias) if profile == "avy" else "")
            joins.append({
                "target_column": col, "join_type": "LEFT JOIN", "table_ref": ref,
                "alias": alias, "on_clause": inferred_on or f"/* TODO_ON_FROM_DRD_ROW_{row.get('excel_row','')}: source table used for {col}; AD rule did not provide a complete ON clause */ 1=1",
                "source": ("DRD_TAXLOT_INFERRED_JOIN" if profile == "taxlot" and inferred_on else ("DRD_AVY_INFERRED_JOIN" if profile == "avy" and inferred_on else "DRD_SOURCE_TABLE_TODO_ON"))
            })
        joins.extend(parsed)
        if profile == "avy":
            joins.extend(extra_avy_join_rows(col))

    # Add curated AVY joins for overrides even if source table columns did not trigger them.
    if profile == "avy":
        for row in mapping_rows:
            joins.extend(extra_avy_join_rows(row.get("target_column", "")))

    # Resolve DRD/source object names to executable physical objects before merging/gating.
    for j in joins:
        j["table_ref"] = resolve_physical_table_ref(j.get("table_ref", ""), profile)
    # Never emit a JOIN back to the driving source alias/table; DRD prose often repeats
    # `ccal_repl_owner.txn t` as context, not as an actual second table instance.
    joins = [j for j in joins if not (j.get("table_ref", "").upper() == primary_ref.upper() or j.get("alias", "").upper() == primary_alias.upper())]
    # Remove noisy per-column TODO joins when a reliable global join already exists.
    joins = prune_todo_joins_when_alias_declared(joins)
    for j in joins:
        j["on_clause"] = sanitize_on_clause(j.get("on_clause", ""))
    joins = merge_exact_joins(joins)
    joins = resolve_join_alias_collisions(joins, impl_rows)

    # Check aliases used in generated expressions.
    declared = {primary_alias.upper()}
    for j in joins:
        if j.get("alias"):
            declared.add(j["alias"].upper())
    for r in impl_rows:
        for a in expression_aliases(r.get("generated_expression", "")):
            if a.upper() not in declared:
                unresolved.append({
                    "target_column": r.get("target_column", ""), "alias": a,
                    "detail": f"Alias {a} is used in generated_expression but no reliable DRD join was declared.",
                    "generated_expression": r.get("generated_expression", ""),
                })
    # Deduplicate exact joins and merge target column lists.
    return merge_exact_joins(joins), unresolved


def choose_primary(mapping_rows: List[Dict[str, str]], profile: str, override: str = "") -> Tuple[str, str]:
    if override:
        parts = override.split()
        ref = parts[0]
        alias = ident(parts[1]) if len(parts) > 1 else make_alias("", ref.split(".")[-1])
        return ref, alias
    if profile == "avy":
        return "CCAL_REPL_OWNER.TXN", "TXN"
    refs = []
    for r in mapping_rows:
        schema, table, attr, ref, alias = source_ref(r)
        if ref:
            refs.append(ref)
    if refs:
        ref = Counter(refs).most_common(1)[0][0]
        return ref, make_alias("", ref.split(".")[-1])
    return "/* TODO_SOURCE_TABLE */", "SRC"


def build_sql(target: str, target_schema: str, mapping_rows: List[Dict[str, str]], odi_by_col: Dict[str, Dict[str, str]], cls_by_col: Dict[str, Dict[str, str]], profile: str, primary_source: str, quote: bool = False) -> Tuple[str, List[Dict[str, str]], List[Dict[str, str]], List[Dict[str, str]]]:
    if target_schema and "." not in target:
        full_target = f"{ident(target_schema)}.{ident(target)}"
    else:
        full_target = ".".join(ident(p) for p in target.split(".") if ident(p)) or "TODO_TARGET"
    primary_ref, primary_alias = choose_primary(mapping_rows, profile, primary_source)
    cols_sql = ",\n".join(f"    {qident(r['target_column'], quote)}" for r in mapping_rows)
    select_lines, impl_rows = [], []
    for i, r in enumerate(mapping_rows, 1):
        col = r["target_column"]
        c = cls_by_col.get(col, {"class": "NO_ODI_HELPER", "reason": ""})
        odi = odi_by_col.get(col, {})
        expr, status, notes = drd_expr(r, profile)
        expr = normalize_rule_aliases(expr, primary_alias, profile)
        expr = postprocess_expression(expr, r, profile, primary_alias, odi.get("xml_logic_full", "") or odi.get("expression", ""))
        schema, table, attr, ref, src_alias = source_ref(r)
        if src_alias.upper() == "CL_VAL" and "CL_VAL." in expr.upper():
            expr = re.sub(r"\bCL_VAL\.", f"{col}_CL_VAL.", expr, flags=re.I)
        comma = "," if i < len(mapping_rows) else ""
        select_lines.append(f"    /* DRD row {sql_block_comment(str(r.get('excel_row')), 40)}; impl={sql_block_comment(status, 100)}; compare={sql_block_comment(c.get('class',''), 100)}; reason={sql_block_comment(c.get('reason',''), 200)} */")
        if r.get("drd_rule"):
            select_lines.append(f"    /* DRD rule: {sql_block_comment(r.get('drd_rule'), 1000)} */")
        # Keep ODI evidence out of generated SQL text to avoid confusing helper/stage
        # references with executable source dependencies. ODI evidence remains in
        # implementation_map.csv.
        select_lines.append(f"    {expr} AS {qident(col, quote)}{comma}")
        rec = {
            "ordinal": str(i), "target_column": col, "excel_row": r.get("excel_row", ""),
            "generated_expression": expr, "implementation_status": status, "implementation_source": "DRD",
            "comparison_class": c.get("class", ""), "comparison_reason": c.get("reason", ""),
            "drd_expression": expr, "drd_status": status, "drd_rule": r.get("drd_rule", ""),
            "source_schema": r.get("source_1", ""), "source_table": r.get("source_2", ""), "source_attribute": r.get("source_3", ""),
            "odi_expression": odi.get("xml_logic_full", "") or odi.get("expression", ""),
            "odi_raw_expression": odi.get("expression", ""), "notes": notes,
        }
        impl_rows.append(rec)
    joins, unresolved_aliases = build_join_graph(mapping_rows, impl_rows, profile, primary_ref, primary_alias)
    join_lines = []
    for j in joins:
        join_lines.append(f"    {j['join_type']} {j['table_ref']} {j['alias']} ON {j['on_clause']} /* {sql_block_comment(j.get('source',''), 160)}; cols={sql_block_comment(j.get('target_column',''), 500)} */")
    join_sql = "\n".join(join_lines) if join_lines else "    /* No DRD joins detected. */"
    sql = f"""/*
Generated by universal_insert_builder.py v{__VERSION__}

Rules:
1. DRD is the source of truth for target contract and generated expressions.
2. ODI XML is helper evidence only; generated SQL does not select from ODI staging output.
3. No synthetic final-source alias is generated.
4. TODO_ON joins are explicit validation blockers, not silent assumptions.
*/

INSERT INTO {full_target} (
{cols_sql}
)
SELECT
{chr(10).join(select_lines)}
FROM {primary_ref} {primary_alias}
{join_sql};
"""
    return sql, impl_rows, joins, unresolved_aliases


def referenced_aliases_from_sql(sql: str) -> set:
    scan = strip_sql_comments(sql)
    for schema in list(SCHEMA_PREFIXES) + ["CCSI_REPL_OWNER"]:
        scan = re.sub(rf"\b{re.escape(schema)}\s*\.\s*[A-Z0-9_$#]+", " ", scan, flags=re.I)
    refs = set()
    for a in re.findall(r"\b([A-Za-z_][A-Za-z0-9_$#]*)\s*\.", scan):
        au = a.upper().replace("$", "_")
        if au not in SCHEMA_PREFIXES and au not in SQL_FUNCS:
            refs.add(au)
    return refs


def executable_sql_errors(sql: str) -> List[Dict[str, str]]:
    errors: List[Dict[str, str]] = []
    scan = strip_sql_comments(sql)
    for i, line in enumerate(scan.splitlines(), 1):
        if not re.search(r"\b(?:LEFT\s+JOIN|JOIN)\b", line, re.I):
            continue
        if re.search(r"\b(populate|note|exception|skip|load|record written|manual|TODO_ON)\b", line, re.I):
            errors.append({"type": "RAW_PROSE_IN_JOIN", "target_column": "", "detail": f"line {i}: {short(line, 350)}"})
        bal = line.count("(") - line.count(")")
        if bal != 0:
            errors.append({"type": "UNBALANCED_PARENS_IN_JOIN", "target_column": "", "detail": f"line {i}: balance={bal}; {short(line, 350)}"})
    alias_counts = Counter()
    for m in re.finditer(r"\bFROM\s+[A-Z0-9_$#]+\.[A-Z0-9_$#]+\s+([A-Z][A-Z0-9_$#]*)", scan, flags=re.I):
        a = m.group(1).upper().replace("$", "_")
        if a not in {"WHERE", "ON", "JOIN", "LEFT", "INNER"}:
            alias_counts[a] += 1
    for m in re.finditer(r"\b(?:LEFT\s+JOIN|JOIN)\s+[A-Z0-9_$#]+\.[A-Z0-9_$#]+\s+([A-Z][A-Z0-9_$#]*)\s+ON\b", scan, flags=re.I):
        alias_counts[m.group(1).upper().replace("$", "_")] += 1
    for alias, n in alias_counts.items():
        if n > 1:
            errors.append({"type": "DUPLICATE_SQL_ALIAS", "target_column": "", "detail": f"{alias} declared {n} times"})
    declared = set(alias_counts)
    for a in sorted(referenced_aliases_from_sql(sql) - declared):
        errors.append({"type": "UNDECLARED_SQL_ALIAS_AFTER_RENDER", "target_column": "", "detail": a})
    return errors


def validate_sql(sql: str, impl_rows: List[Dict[str, str]], joins: List[Dict[str, str]], unresolved_aliases: List[Dict[str, str]]) -> List[Dict[str, str]]:
    errors: List[Dict[str, str]] = []
    scan_sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.S)
    scan_sql = re.sub(r"--[^\n]*", " ", scan_sql)
    up = scan_sql.upper()
    for bad in ["ODI_FINAL_SOURCE", "FROM ODI_FINAL_SOURCE", " AVY_FACT_STEP5_STG_RT", "SSDS_AVY_FACT_STEP5_STG"]:
        if bad in up:
            errors.append({"type": "FORBIDDEN_ODI_FINAL_HELPER_IN_GENERATED_SQL", "target_column": "", "detail": bad})
    alias_sigs = defaultdict(set)
    for j in joins:
        alias_sigs[j.get("alias", "").upper()].add(alias_collision_signature(j))
        if "TODO_ON" in j.get("on_clause", "") or "TODO:" in j.get("on_clause", ""):
            errors.append({"type": "JOIN_REQUIRES_REVIEW", "target_column": j.get("target_column", ""), "detail": f"{j.get('table_ref')} {j.get('alias')} ON {j.get('on_clause')}"})
    for alias, sigs in alias_sigs.items():
        if alias and len(sigs) > 1:
            errors.append({"type": "JOIN_ALIAS_COLLISION", "target_column": "", "detail": f"Alias {alias} is used for {len(sigs)} different joins."})
    for a in unresolved_aliases:
        errors.append({"type": "UNDECLARED_GENERATED_ALIAS", "target_column": a.get("target_column", ""), "detail": a.get("detail", "")})
    for r in impl_rows:
        expr = r.get("generated_expression", "")
        if re.match(r"(?i)^\s*(use|fetch|lookup|look\s+up|if\s+there)\b", expr):
            errors.append({"type": "RAW_PROSE_IN_GENERATED_EXPRESSION", "target_column": r.get("target_column", ""), "detail": short(expr, 500)})
        c, e = case_end_balance(expr)
        if c != e:
            errors.append({"type": "CASE_END_IMBALANCE", "target_column": r.get("target_column", ""), "detail": f"CASE={c}; END={e}; {short(expr, 500)}"})
        # Oracle accepts unqualified columns only when they physically exist in
        # the joined row source. For TaxLot currency CASE shorthand this was the
        # source of live ORA-00904; reject unqualified tokens in CASE expressions
        # so the generator must source-qualify them.
        if r.get("implementation_status", "").startswith("DRD"):
            bad = unqualified_identifiers_in_case(expr)
            if bad:
                errors.append({"type": "UNQUALIFIED_IDENTIFIER_IN_CASE", "target_column": r.get("target_column", ""), "detail": ", ".join(bad) + "; " + short(expr, 500)})
    errors.extend(executable_sql_errors(sql))
    return errors


def tri_compare(impl_rows: List[Dict[str, str]]):
    rows = []
    drd_odi_review = set()
    for r in impl_rows:
        col = r["target_column"]
        cls = r.get("comparison_class", "")
        if cls in {"REVIEW_REQUIRED", "MISSING_IN_ODI"}:
            drd_odi_review.add(col)
        rows.append({
            "target_column": col,
            "drd_vs_odi_class": cls,
            "generated_vs_drd": "DRD_SOURCE",
            "generated_vs_odi": "EXPECTED_REVIEW_MISMATCH" if cls in {"REVIEW_REQUIRED", "MISSING_IN_ODI"} else "ODI_HELPER_NOT_USED_FOR_SQL",
            "same_mismatch_as_drd_odi": "Y",
        })
    return rows, {"drd_odi_review_fields": sorted(drd_odi_review), "generated_odi_review_fields": sorted(drd_odi_review), "same_review_field_set": True}



def extract_rendered_join_inventory(sql: str):
    """Return the actual JOIN inventory from the final rendered executable SQL.

    This is intentionally derived after schema-KB patching so summary/join_inventory.csv
    cannot go stale when a resolution profile injects or rewrites joins.
    """
    clean = strip_generated_comments(sql)
    rows = []
    # Capture FROM as context plus every JOIN. Join ON text stops at the next JOIN/WHERE/GROUP/ORDER/semicolon.
    pat = re.compile(
        r"\b(LEFT\s+JOIN|RIGHT\s+JOIN|INNER\s+JOIN|FULL\s+JOIN|JOIN)\s+"
        r"([A-Z0-9_$#]+\.[A-Z0-9_$#]+)\s+([A-Z][A-Z0-9_$#]*)\s+ON\s+"
        r"(.*?)(?=\n\s*(?:LEFT\s+JOIN|RIGHT\s+JOIN|INNER\s+JOIN|FULL\s+JOIN|JOIN|WHERE|GROUP\s+BY|ORDER\s+BY|HAVING|;|$))",
        re.I | re.S,
    )
    for m in pat.finditer(clean):
        rows.append({
            "target_column": "",
            "join_type": cmp.normalize_space(m.group(1).upper()),
            "table_ref": cmp.normalize_space(m.group(2).upper()),
            "alias": cmp.normalize_space(m.group(3).upper()),
            "on_clause": cmp.normalize_space(m.group(4)),
            "source": "RENDERED_EXECUTABLE_SQL",
        })
    return rows

def run(args):
    out = Path(args.out).expanduser().resolve()
    out.mkdir(parents=True, exist_ok=True)
    xlsx = Path(args.xlsx).expanduser().resolve()
    xml = Path(args.xml).expanduser().resolve() if args.xml else None
    if not xlsx.exists():
        raise FileNotFoundError(xlsx)
    if xml and not xml.exists():
        raise FileNotFoundError(xml)
    xml_targets, odi_by_col, final_lineage, sql_blocks = load_odi(xml)
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    # AVY workbooks have a stable Table-View layout. Use it directly when profile=avy
    # to avoid slow/fragile workbook-wide autodetection over many auxiliary sheets.
    mapping_sheet_override = args.mapping_sheet or ""
    header_row_override = args.header_row
    target_col_override = args.target_col or ""
    source_cols_override = args.source_cols or ""
    rule_col_override = args.rule_col or ""
    if args.profile == "avy":
        if not mapping_sheet_override:
            if "Table-View" in wb.sheetnames:
                mapping_sheet_override = "Table-View"
            elif "Table-View (2)" in wb.sheetnames:
                mapping_sheet_override = "Table-View (2)"
        header_row_override = header_row_override or 12
        target_col_override = target_col_override or "B"
        source_cols_override = source_cols_override or "Y,Z,AA"
        rule_col_override = rule_col_override or "AD"

    detection = cmp.auto_detect_mapping(
        wb, xml_targets=xml_targets, target_table_override=args.target_table or "",
        mapping_sheet_override=mapping_sheet_override, header_row_override=header_row_override,
        target_col_override=target_col_override, source_cols_override=source_cols_override, rule_col_override=rule_col_override,
    )
    profile = infer_profile(detection, args.profile)
    mapping_rows, etl_notes = cmp.extract_mapping_from_xlsx(xlsx, detection, notes_sheet=args.notes_sheet)
    cls_by_col, column_diff = comparison_classes(mapping_rows, odi_by_col, profile)
    target = args.target_table or detection.target_table_from_sheet or "TODO_TARGET"
    sql, impl_rows, joins, unresolved_aliases = build_sql(target, args.target_schema, mapping_rows, odi_by_col, cls_by_col, profile, args.primary_source, args.quote_identifiers)
    clean_sql = strip_generated_comments(sql)
    kb_audit = []
    kb_errors = []
    kb_meta = {}
    if getattr(args, "schema_kb", ""):
        if schema_kb_gate is None:
            raise RuntimeError("schema_kb_sql_gate.py is required for --schema-kb")
        objects, table_owners, owners, kb_data = schema_kb_gate.load_kb(args.schema_kb)
        resolution_profile = schema_kb_gate.load_resolution_profile(getattr(args, "resolution_profile", ""))
        clean_sql, kb_audit = schema_kb_gate.apply_kb_patch(clean_sql, objects, table_owners, drop_bad_target_cols=True, profile=resolution_profile)
        kb_errors = schema_kb_gate.validate_sql(clean_sql, objects, table_owners, profile=resolution_profile)
        ds = kb_data.get("pdm", {}).get("datasource", {})
        kb_meta = {"schema_kb_datasource_id": ds.get("id"), "schema_kb_name": ds.get("name"), "schema_kb_host": ds.get("host"), "schema_kb_database": ds.get("database")}
    errors = validate_sql(sql, impl_rows, joins, unresolved_aliases)
    errors.extend(executable_clean_sql_errors(clean_sql))
    for etype, item, detail in kb_errors:
        errors.append({"type": "SCHEMA_KB_" + etype, "target_column": item, "detail": detail})
    tri_rows, tri_summary = tri_compare(impl_rows)

    (out / "generated_insert_select_candidate.sql").write_text(clean_sql, encoding="utf-8")
    (out / "generated_insert_select_candidate_annotated.sql").write_text(sql, encoding="utf-8")
    (out / "generated_insert_select_candidate_drd_blueprint.sql").write_text(sql, encoding="utf-8")
    write_csv(out / "implementation_map.csv", impl_rows, ["ordinal", "target_column", "excel_row", "generated_expression", "implementation_status", "implementation_source", "comparison_class", "comparison_reason", "drd_expression", "drd_status", "drd_rule", "source_schema", "source_table", "source_attribute", "odi_expression", "odi_raw_expression", "notes"])
    rendered_joins = extract_rendered_join_inventory(clean_sql)
    write_csv(out / "join_inventory_pre_kb.csv", joins, ["target_column", "join_type", "table_ref", "alias", "on_clause", "source"])
    write_csv(out / "join_inventory.csv", rendered_joins, ["target_column", "join_type", "table_ref", "alias", "on_clause", "source"])
    write_csv(out / "validation_errors.csv", errors, ["type", "target_column", "detail"])
    write_csv(out / "schema_kb_resolution_audit.csv", [{"action": a, "item": b, "detail": c} for a,b,c in kb_audit], ["action", "item", "detail"])
    write_csv(out / "schema_kb_validation_errors.csv", [{"type": a, "item": b, "detail": c} for a,b,c in kb_errors], ["type", "item", "detail"])
    write_csv(out / "tri_compare_report.csv", tri_rows, ["target_column", "drd_vs_odi_class", "generated_vs_drd", "generated_vs_odi", "same_mismatch_as_drd_odi"])
    if column_diff:
        write_csv(out / "drd_vs_odi_column_diff.csv", column_diff, list(column_diff[0].keys()))
    mapping_fields = ["excel_row", "target_column", "source_1", "source_2", "source_3", "drd_rule"]
    write_csv(out / "mapping_extract.csv", mapping_rows, mapping_fields)
    det = detection.as_human()
    det.update({"profile": profile, "target_table_used": target, "target_schema": args.target_schema, "mapping_rows": str(len(mapping_rows)), "sql_source_model": "DRD_SOURCE_NO_ODI_FINAL_CTE"})
    (out / "detected_layout.json").write_text(json.dumps(det, indent=2), encoding="utf-8")
    summary = {
        "version": __VERSION__, "profile": profile, "target_table": target, "mapping_rows": len(mapping_rows),
        "implementation_status_counts": dict(Counter(r["implementation_status"] for r in impl_rows)),
        "comparison_class_counts": dict(Counter(r.get("comparison_class", "") for r in impl_rows)),
        "join_rows": len(rendered_joins), "builder_join_rows_pre_kb": len(joins), "validation_error_counts": dict(Counter(e["type"] for e in errors)),
        "schema_kb_audit_counts": dict(Counter(a[0] for a in kb_audit)),
        "schema_kb_validation_error_counts": dict(Counter(e[0] for e in kb_errors)),
        **kb_meta,
        "forbidden_odi_final_helper_present": any(e["type"] == "FORBIDDEN_ODI_FINAL_HELPER_IN_GENERATED_SQL" for e in errors),
        **tri_summary,
        "note": "v6.1: generated SQL is DRD-driven. Pre-KB builder does not hardcode datasource-specific owner/table/column resolver rules. Schema KB path is supplied by --schema-kb; optional datasource-specific corrections are supplied by --resolution-profile. join_inventory.csv and join_rows are derived from the final executable SQL after KB/profile patching."
    }
    (out / "final_consistency_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    if not args.quiet:
        print(json.dumps(summary, indent=2))
    return out


def build_arg_parser():
    p = argparse.ArgumentParser(description="Universal DRD-driven INSERT builder. ODI XML is comparison evidence only.")
    p.add_argument("--xlsx", required=True)
    p.add_argument("--xml", default="")
    p.add_argument("--out", default="insert_builder_output")
    p.add_argument("--target-schema", default="")
    p.add_argument("--target-table", default="")
    p.add_argument("--primary-source", default="", help="Override source as 'SCHEMA.TABLE ALIAS' or 'SCHEMA.TABLE'.")
    p.add_argument("--mapping-sheet", default="")
    p.add_argument("--notes-sheet", default="ETL Notes")
    p.add_argument("--target-col", default="")
    p.add_argument("--source-cols", default="")
    p.add_argument("--rule-col", default="")
    p.add_argument("--header-row", type=int, default=None)
    p.add_argument("--profile", default="auto", choices=["auto", "generic", "avy", "taxlot"])
    p.add_argument("--quote-identifiers", action="store_true")
    p.add_argument("--schema-kb", default="", help="Oracle schema KB JSON. When provided, owner/table/column refs are validated/resolved against KB as source of truth.")
    p.add_argument("--resolution-profile", default="", help="Optional JSON with datasource/project-specific table/column/expression overrides. Keeps KB engine generic.")
    p.add_argument("--quiet", action="store_true")
    return p


def main(argv=None):
    args = build_arg_parser().parse_args(argv)
    try:
        run(args)
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
