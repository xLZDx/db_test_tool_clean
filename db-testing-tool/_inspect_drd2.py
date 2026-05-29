"""Inspect Table-View (2) sheet in detail to find mapping rows."""
import openpyxl
import json

wb = openpyxl.load_workbook("DRD_Activity_Fact.xlsx", read_only=True, data_only=True)
ws = wb["Table-View (2)"]

# Print first 15 rows to find header row
print("=== First 15 rows of Table-View (2) ===")
for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
    print(f"Row {i}: {row[:8]}")

print("\n=== Looking for header row (Physical Name, etc.) ===")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    cells = [str(c) if c is not None else '' for c in row]
    row_str = ' | '.join(cells[:6])
    if any(kw in row_str.upper() for kw in ['PHYSICAL', 'TARGET', 'SOURCE', 'COLUMN_NAME', 'COLUMN NAME']):
        print(f"Row {i}: {row[:8]}")
    if i > 20:
        break

wb.close()
