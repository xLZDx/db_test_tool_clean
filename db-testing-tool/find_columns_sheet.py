#!/usr/bin/env python3
"""Find the sheet with the most columns in DRD Excel."""
import openpyxl
import json

excel_file = "c:/Users/ikorostelev/Downloads/DRD_Activity_Fact.xlsx"
wb = openpyxl.load_workbook(excel_file)

print("Analyzing all sheets for column count:\n")
sheet_info = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Count non-empty cells in first row
    col_count = 0
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            col_count += 1
    row_count = ws.max_row
    sheet_info.append((sheet_name, col_count, row_count))
    print(f"{sheet_name:30s} - Columns: {col_count:3d}, Rows: {row_count:4d}")

# Find sheet with most columns
best_sheet = max(sheet_info, key=lambda x: x[1])
print(f"\n✓ Using sheet with most columns: {best_sheet[0]} ({best_sheet[1]} columns)")

# Extract columns
ws = wb[best_sheet[0]]
columns = []
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    if cell.value:
        col_name = str(cell.value).strip()
        columns.append(col_name)

print(f"\nTotal columns found: {len(columns)}")
print("\nFirst 20 columns:")
for i, col in enumerate(columns[:20], 1):
    print(f"  {i:3d}. {col}")

if len(columns) > 20:
    print(f"\n... ({len(columns)-20} more) ...\n")
    print("Last 10 columns:")
    for i, col in enumerate(columns[-10:], len(columns)-9):
        print(f"  {i:3d}. {col}")

# Save to JSON and SQL
with open('drd_columns.json', 'w') as f:
    json.dump(columns, f, indent=2)

# Create SQL column list for INSERT statement
sql_cols = ",\n    ".join(columns)
print(f"\n\nSQL column list (for INSERT):\n({sql_cols})")

print(f"\n✓ Saved {len(columns)} columns to drd_columns.json")
