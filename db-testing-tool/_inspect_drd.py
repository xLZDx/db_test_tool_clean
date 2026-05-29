"""Inspect DRD_Activity_Fact.xlsx sheets and columns directly."""
import openpyxl
import json

wb = openpyxl.load_workbook("DRD_Activity_Fact.xlsx", read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    print(f"\n=== {sn} === ({ws.max_row} rows, {ws.max_column} cols)")
    if rows:
        print("Row1:", rows[0][:10])
    if len(rows) > 1:
        print("Row2:", rows[1][:10])
wb.close()
