"""P7 -- Multi-sheet DRD Excel parser.

Reads ALL sheets from a DRD workbook in non-read-only mode so that
openpyxl can expose hyperlink targets.  Returns a structured result that:

  - classifies every sheet by its role (mapping, etl_notes, grain, model,
    consumer_view, attribute, unknown)
  - extracts transformation rules from ETL-Notes style sheets
  - extracts grain / key columns from Grain-Columns style sheets
  - extracts JOIN / filter snippets from Model / Consumer-View style sheets
  - records hyperlinks whose target cannot be resolved locally as
    ``deferred_refs`` -- anything in deferred_refs degrades the verdict
    from FULL_DRD to PARTIAL_DRD

Design invariants (operator-locked):
  - deferred_refs is NEVER silently empty when hyperlinks exist
  - verdict in {'FULL_DRD', 'PARTIAL_DRD'}
  - PARTIAL_DRD is visible to the caller -- never masked as FULL
  - Cross-sheet extracted_rules feed into ODI model construction (caller
    is responsible for merging them into ColumnMapping annotations)
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from enum import Enum
from typing import Any, Dict, List, Optional


# ---------------------------------------------------------------------------
# Sheet role classification
# ---------------------------------------------------------------------------

class SheetRole(str, Enum):
    MAPPING      = "mapping"        # primary DRD column mapping (Table-View*)
    ETL_NOTES    = "etl_notes"      # transformation rules, business logic
    GRAIN        = "grain"          # grain / primary-key columns
    MODEL        = "model"          # entity-relationship / data model
    CONSUMER_VIEW = "consumer_view" # downstream consumer projections
    ATTRIBUTE    = "attribute"      # attribute / column dictionary
    TABLE        = "table"          # table-level DDL / metadata
    UNKNOWN      = "unknown"


_ROLE_PATTERNS: list[tuple[str, SheetRole]] = [
    (r"etl.?note",        SheetRole.ETL_NOTES),
    (r"grain",            SheetRole.GRAIN),
    (r"model",            SheetRole.MODEL),
    (r"consumer.?view",   SheetRole.CONSUMER_VIEW),
    (r"additional.?view", SheetRole.CONSUMER_VIEW),
    (r"rj[mb]?.*view",    SheetRole.CONSUMER_VIEW),
    (r"attribute",        SheetRole.ATTRIBUTE),
    (r"^table$",          SheetRole.TABLE),
    (r"table.?view",      SheetRole.MAPPING),
    (r"^view$",           SheetRole.MAPPING),
    (r"table_?view_?sorted", SheetRole.MAPPING),
    (r"mapping",          SheetRole.MAPPING),
    (r"avy_fact",         SheetRole.MAPPING),
]


def _classify_sheet(name: str) -> SheetRole:
    n = name.strip().lower()
    for pattern, role in _ROLE_PATTERNS:
        if re.search(pattern, n):
            return role
    return SheetRole.UNKNOWN


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------

@dataclass
class DeferredRef:
    """A reference that could not be resolved locally."""
    source: str        # e.g. sheet name or cell address
    ref_type: str      # "hyperlink", "pbi", "external_formula", "cross_sheet"
    raw_value: str     # the unresolved value / URL
    reason: str        # why it's deferred


@dataclass
class ExtractionRule:
    """A transformation / join / filter rule extracted from a non-mapping sheet."""
    sheet: str
    role: SheetRole
    rule_type: str     # "transformation", "join_condition", "filter", "grain_col", "note"
    target_col: Optional[str]
    description: str
    raw_row: List[Any] = field(default_factory=list)


@dataclass
class SheetSummary:
    name: str
    role: SheetRole
    row_count: int
    non_empty_rows: int
    hyperlink_count: int
    sample_rows: List[List[Any]] = field(default_factory=list)  # up to 10 rows


@dataclass
class DrdMultiSheetResult:
    """Full result of parsing ALL sheets in a DRD workbook."""
    sheets: List[SheetSummary] = field(default_factory=list)
    deferred_refs: List[DeferredRef] = field(default_factory=list)
    extracted_rules: List[ExtractionRule] = field(default_factory=list)
    verdict: str = "FULL_DRD"   # "FULL_DRD" | "PARTIAL_DRD"
    mapping_sheet: Optional[str] = None   # best-scored mapping sheet name
    grain_columns: List[str] = field(default_factory=list)
    total_hyperlinks: int = 0

    def to_dict(self) -> dict:
        return {
            "verdict": self.verdict,
            "mapping_sheet": self.mapping_sheet,
            "grain_columns": self.grain_columns,
            "total_hyperlinks": self.total_hyperlinks,
            "sheets": [
                {
                    "name": s.name,
                    "role": s.role.value,
                    "row_count": s.row_count,
                    "non_empty_rows": s.non_empty_rows,
                    "hyperlink_count": s.hyperlink_count,
                    "sample_rows": [
                        [str(c) if c is not None else "" for c in row]
                        for row in s.sample_rows
                    ],
                }
                for s in self.sheets
            ],
            "deferred_refs": [
                {
                    "source": d.source,
                    "ref_type": d.ref_type,
                    "raw_value": d.raw_value,
                    "reason": d.reason,
                }
                for d in self.deferred_refs
            ],
            "extracted_rules": [
                {
                    "sheet": r.sheet,
                    "role": r.role.value,
                    "rule_type": r.rule_type,
                    "target_col": r.target_col,
                    "description": r.description,
                }
                for r in self.extracted_rules
            ],
        }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _looks_like_pbi(text: str) -> bool:
    return bool(re.search(r"\b(pbi|tfs|bug|work.?item|#\d{4,})\b", text, re.I))


def _looks_like_sql(text: str) -> bool:
    return bool(re.search(
        r"\b(select|from|where|join|left|right|inner|outer|case|when|then|nvl|coalesce|decode|null|not null)\b",
        text, re.I,
    ))


# ---------------------------------------------------------------------------
# Per-sheet extractors
# ---------------------------------------------------------------------------

def _extract_etl_notes(ws_rows: list[list[Any]], sheet_name: str) -> list[ExtractionRule]:
    """Extract transformation rules from an ETL Notes style sheet."""
    rules: list[ExtractionRule] = []
    for row in ws_rows:
        non_empty = [_cell_text(c) for c in row if _cell_text(c)]
        if not non_empty:
            continue
        desc = " | ".join(non_empty[:6])
        if len(desc) < 5:
            continue
        # Try to identify a target column reference in first two cells
        target = non_empty[0] if len(non_empty) > 0 else None
        if target and re.match(r"^[A-Z][A-Z0-9_]{1,59}$", target):
            rule_type = "transformation"
        elif _looks_like_sql(desc):
            rule_type = "sql_snippet"
            target = None
        else:
            rule_type = "note"
            target = None
        rules.append(ExtractionRule(
            sheet=sheet_name,
            role=SheetRole.ETL_NOTES,
            rule_type=rule_type,
            target_col=target,
            description=desc,
            raw_row=row,
        ))
    return rules


def _extract_grain(ws_rows: list[list[Any]], sheet_name: str) -> tuple[list[str], list[ExtractionRule]]:
    """Extract grain / key column names from a Grain Columns sheet."""
    grain_cols: list[str] = []
    rules: list[ExtractionRule] = []
    for row in ws_rows:
        for cell in row:
            t = _cell_text(cell)
            if t and re.match(r"^[A-Z][A-Z0-9_]{1,59}$", t) and t not in grain_cols:
                grain_cols.append(t)
                rules.append(ExtractionRule(
                    sheet=sheet_name,
                    role=SheetRole.GRAIN,
                    rule_type="grain_col",
                    target_col=t,
                    description=f"Grain / key column: {t}",
                    raw_row=row,
                ))
    return grain_cols, rules


def _extract_model(ws_rows: list[list[Any]], sheet_name: str) -> list[ExtractionRule]:
    """Extract join conditions and filter criteria from a Model sheet."""
    rules: list[ExtractionRule] = []
    for row in ws_rows:
        non_empty = [_cell_text(c) for c in row if _cell_text(c)]
        if not non_empty:
            continue
        desc = " | ".join(non_empty[:6])
        if _looks_like_sql(desc):
            rules.append(ExtractionRule(
                sheet=sheet_name,
                role=SheetRole.MODEL,
                rule_type="join_condition" if re.search(r"\bjoin\b", desc, re.I) else "filter",
                target_col=None,
                description=desc,
                raw_row=row,
            ))
        elif len(desc) > 10:
            rules.append(ExtractionRule(
                sheet=sheet_name,
                role=SheetRole.MODEL,
                rule_type="note",
                target_col=None,
                description=desc,
                raw_row=row,
            ))
    return rules


def _extract_consumer_view(ws_rows: list[list[Any]], sheet_name: str) -> list[ExtractionRule]:
    rules: list[ExtractionRule] = []
    for row in ws_rows:
        non_empty = [_cell_text(c) for c in row if _cell_text(c)]
        if not non_empty:
            continue
        desc = " | ".join(non_empty[:6])
        if len(desc) > 5:
            rules.append(ExtractionRule(
                sheet=sheet_name,
                role=SheetRole.CONSUMER_VIEW,
                rule_type="consumer_projection",
                target_col=non_empty[0] if re.match(r"^[A-Z][A-Z0-9_]{1,59}$", non_empty[0]) else None,
                description=desc,
                raw_row=row,
            ))
    return rules


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_all_sheets(xlsx_bytes: bytes) -> DrdMultiSheetResult:
    """Parse ALL sheets in a DRD workbook.

    Uses non-read-only mode so openpyxl exposes .hyperlink on cells.
    Never raises -- returns a result with PARTIAL_DRD verdict on any issue.
    """
    try:
        import openpyxl
    except ImportError:
        return DrdMultiSheetResult(
            verdict="PARTIAL_DRD",
            deferred_refs=[DeferredRef(
                source="import",
                ref_type="external_formula",
                raw_value="openpyxl",
                reason="openpyxl not installed",
            )],
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=True)
    except Exception as exc:
        return DrdMultiSheetResult(
            verdict="PARTIAL_DRD",
            deferred_refs=[DeferredRef(
                source="workbook",
                ref_type="external_formula",
                raw_value=str(exc),
                reason="Failed to open workbook",
            )],
        )

    result = DrdMultiSheetResult()

    # Mapping sheet: pick the best-scored Table-View* sheet (reuse logic inline)
    mapping_candidates = [sn for sn in wb.sheetnames if re.search(r"table.?view", sn, re.I)]
    if mapping_candidates:
        # prefer "(2)" variant if present
        result.mapping_sheet = next(
            (s for s in mapping_candidates if "(2)" in s or s.endswith("2")),
            mapping_candidates[0],
        )
    else:
        result.mapping_sheet = wb.sheetnames[0] if wb.sheetnames else None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        role = _classify_sheet(sheet_name)

        all_rows: list[list[Any]] = []
        hyperlink_count = 0

        for row in ws.iter_rows():
            row_vals: list[Any] = []
            for cell in row:
                val = cell.value
                row_vals.append(val)

                # Check for hyperlinks (openpyxl exposes .hyperlink in non-read-only mode)
                hl = getattr(cell, "hyperlink", None)
                if hl is not None:
                    target = getattr(hl, "target", None) or str(hl)
                    hyperlink_count += 1
                    result.total_hyperlinks += 1
                    # Determine if it can be resolved locally
                    if target and (
                        target.startswith("http") or
                        target.startswith("\\\\") or
                        re.search(r"(sharepoint|tfs|devops|visualstudio)", target, re.I)
                    ):
                        result.deferred_refs.append(DeferredRef(
                            source=f"{sheet_name}!{cell.coordinate}",
                            ref_type="hyperlink",
                            raw_value=target[:200],
                            reason="External URL/TFS link cannot be followed without network+auth",
                        ))
                    elif _looks_like_pbi(str(val or "") + " " + str(target or "")):
                        result.deferred_refs.append(DeferredRef(
                            source=f"{sheet_name}!{cell.coordinate}",
                            ref_type="pbi",
                            raw_value=str(target or val or "")[:200],
                            reason="PBI/TFS work item link not resolved (needs AD auth)",
                        ))

            all_rows.append(row_vals)

        non_empty = sum(1 for r in all_rows if any(c is not None for c in r))
        sample = [r for r in all_rows if any(c is not None for c in r)][:10]

        result.sheets.append(SheetSummary(
            name=sheet_name,
            role=role,
            row_count=len(all_rows),
            non_empty_rows=non_empty,
            hyperlink_count=hyperlink_count,
            sample_rows=sample,
        ))

        # Per-role extraction
        non_empty_rows = [r for r in all_rows if any(c is not None for c in r)]

        if role == SheetRole.ETL_NOTES:
            result.extracted_rules.extend(_extract_etl_notes(non_empty_rows, sheet_name))
        elif role == SheetRole.GRAIN:
            grain_cols, grain_rules = _extract_grain(non_empty_rows, sheet_name)
            result.grain_columns.extend(g for g in grain_cols if g not in result.grain_columns)
            result.extracted_rules.extend(grain_rules)
        elif role == SheetRole.MODEL:
            result.extracted_rules.extend(_extract_model(non_empty_rows, sheet_name))
        elif role in (SheetRole.CONSUMER_VIEW, SheetRole.ATTRIBUTE):
            result.extracted_rules.extend(_extract_consumer_view(non_empty_rows, sheet_name))

    # Verdict
    result.verdict = "PARTIAL_DRD" if result.deferred_refs else "FULL_DRD"
    return result
