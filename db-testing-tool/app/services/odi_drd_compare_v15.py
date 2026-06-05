# """VENDORED (R2, 2026-06-05) from
# D:\test 2\odi_drd_compare_tool_universal_v15_final -- the v15 universal
# DRD-vs-ODI comparator. ASCII-sanitized (3 em-dashes -> --). Pipeline functions
# are reused verbatim; two thin wrappers (compare_summary / compare_to_dir) added
# at the end force profile='generic' so no AVY/TaxLot curated heuristics fire.
# NOT yet wired into any endpoint (that is R3). See
# core/V15_COMPARE_TOOL_EVAL_AND_PLAN_2026-06-05.md.
# """
#!/usr/bin/env python3
"""
compare_drd_odi_universal.py

Universal DRD Excel vs ODI Scenario XML comparison tool.

What is new in the universal version:
- Auto-detects the best DRD mapping sheet.
- Auto-detects header row.
- Auto-detects target column, source schema/table/attribute columns, rule column.
- Supports AVY_FACT style and TaxLot closed/open style without hardcoding.
- Still allows manual overrides when auto-detection is wrong.
- Normalizes ODI getObjectName(...) macros before SQL parsing.
- Works with both INSERT...SELECT and MERGE-style ODI SQL where possible.

This is a static-analysis tool. It surfaces differences and review candidates.
It does not prove semantic equivalence automatically.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

__VERSION__ = "15.0-final"

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Missing dependency: openpyxl. Install with: pip install -r requirements.txt") from exc


# ======================================================================================
# Helpers
# ======================================================================================

def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value)
    text = html.unescape(text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", clean_text(text)).strip()


def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    total = 0
    for ch in letter:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid Excel column letter: {letter}")
        total = total * 26 + (ord(ch) - ord("A") + 1)
    return total


def index_to_col_letter(index: int) -> str:
    out = ""
    while index:
        index, rem = divmod(index - 1, 26)
        out = chr(65 + rem) + out
    return out


def normalize_identifier(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ""
    value = value.strip().strip('"').strip("'").strip()
    # Remove Excel formula errors / booleans / accidental descriptions.
    if value.upper() in {"#N/A", "TRUE", "FALSE", "YES", "NO", "NULL", "N/A"}:
        return ""
    value = re.sub(r"\[[^\]]+\]", "", value)
    # If it looks like a formula, skip.
    if value.startswith("="):
        return ""
    # If it contains whitespace and is not a normal identifier, keep first token only.
    if re.search(r"\s", value):
        value = value.split()[0]
    # If schema.table.column, compare by column.
    value = value.split(".")[-1]
    value = value.strip('"').strip("'").strip()
    value = re.sub(r"[^A-Za-z0-9_#$]", "", value)
    return value.upper()


def is_probable_column_name(value: str) -> bool:
    v = normalize_identifier(value)
    if not v:
        return False
    if len(v) < 2:
        return False
    if v in {
        "TARGET", "COLUMN", "COLUMNS", "ATTRIBUTE", "ATTRIBUTES", "FIELD", "FIELDS",
        "SOURCE", "TRANSFORMATION", "NOTES", "VIEW", "TABLE", "NUMBER", "VARCHAR2",
        "DATE", "TIMESTAMP", "NUMERIC", "CHARACTER",
    }:
        return False
    # Oracle physical columns normally start with a letter, but DRD documents can contain
    # numeric-prefixed business column names such as 1099/401K-style fields. Keep those too.
    if re.fullmatch(r"[A-Z][A-Z0-9_#$]*", v):
        return True
    if re.fullmatch(r"[0-9][A-Z0-9_#$]*", v) and "_" in v:
        return True
    return False


def write_csv(path: Path, rows: Sequence[Dict[str, str]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(fieldnames), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def md_escape(text: str) -> str:
    return clean_text(text).replace("|", "\\|").replace("\n", "<br>")


def short(text: str, limit: int = 500) -> str:
    text = normalize_space(text)
    if len(text) <= limit:
        return text
    return text[:limit - 3] + "..."


def unique_keep_order(items: Iterable[str]) -> List[str]:
    seen = set()
    out = []
    for item in items:
        item = clean_text(item)
        if item and item not in seen:
            seen.add(item)
            out.append(item)
    return out


# ======================================================================================
# ODI XML extraction
# ======================================================================================

def parse_odi_objects(xml_path: Path) -> List[Dict[str, str]]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    objects: List[Dict[str, str]] = []

    for obj in root.findall(".//Object"):
        rec: Dict[str, str] = {"class": obj.attrib.get("class", "")}
        for field in obj.findall("Field"):
            name = field.attrib.get("name", "")
            value = field.text or ""
            rec[name] = clean_text(value)
        objects.append(rec)

    return objects


def extract_target_resources_from_xml(objects: List[Dict[str, str]]) -> List[str]:
    resources = []
    # SnpScenStep fields.
    for o in objects:
        for key in ("ResName", "TableName"):
            v = normalize_identifier(o.get(key, ""))
            if v:
                resources.append(v)
        gi = o.get("GenInfo", "")
        if gi:
            # target name/resource inside map-generation-info.
            for m in re.finditer(r'<target[^>]+(?:name|resource)="([^"]+)"', gi, flags=re.I):
                v = normalize_identifier(m.group(1))
                if v:
                    resources.append(v)
    # Scenario name sometimes contains target.
    for o in objects:
        if o.get("class", "").endswith("SnpScen"):
            v = normalize_identifier(o.get("ScenName", ""))
            if v:
                resources.append(v)
    return unique_keep_order(resources)


def extract_odi_summary(objects: List[Dict[str, str]]) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], List[Dict[str, str]]]:
    scenarios = [o for o in objects if o.get("class", "").endswith("SnpScen")]
    steps = [o for o in objects if o.get("class", "").endswith("SnpScenStep")]
    tasks = [o for o in objects if o.get("class", "").endswith("SnpScenTask")]

    step_rows: List[Dict[str, str]] = []
    for s in sorted(steps, key=lambda x: int(x.get("Nno") or 0)):
        nno = s.get("Nno", "")
        related_tasks = [t for t in tasks if t.get("Nno") == nno]
        sql_count = 0
        for t in related_tasks:
            if clean_text(t.get("DefTxt", "")) or clean_text(t.get("ColTxt", "")):
                sql_count += 1
        step_rows.append({
            "step_no": nno,
            "step_name": s.get("StepName", ""),
            "step_type": s.get("StepType", ""),
            "ok_next_step": s.get("OkNextStep", ""),
            "ko_next_step": s.get("KoNextStep", ""),
            "task_count": str(len(related_tasks)),
            "sql_task_count": str(sql_count),
            "variable_name": s.get("VarName", ""),
            "variable_value": s.get("VarLongValue", ""),
            "resource_name": s.get("ResName", ""),
            "table_name": s.get("TableName", ""),
        })

    sql_blocks: List[Dict[str, str]] = []
    for t in sorted(tasks, key=lambda x: (int(x.get("Nno") or 0), int(x.get("ScenTaskNo") or 0))):
        for field_name in ("DefTxt", "ColTxt"):
            sql = clean_text(t.get(field_name, ""))
            if not sql:
                continue
            if not looks_like_sql(sql):
                continue
            sql_blocks.append({
                "step_no": t.get("Nno", ""),
                "task_no": t.get("ScenTaskNo", ""),
                "task_type": t.get("TaskType", ""),
                "task_name_1": t.get("TaskName1", ""),
                "task_name_2": t.get("TaskName2", ""),
                "task_name_3": t.get("TaskName3", ""),
                "field": field_name,
                "sql": sql,
            })

    scenario_rows: List[Dict[str, str]] = []
    for s in scenarios:
        scenario_rows.append({
            "scenario_name": s.get("ScenName", ""),
            "scenario_version": s.get("ScenVersion", ""),
            "scenario_no": s.get("ScenNo", ""),
            "first_date": s.get("FirstDate", ""),
            "last_date": s.get("LastDate", ""),
            "first_user": s.get("FirstUser", ""),
            "last_user": s.get("LastUser", ""),
        })

    return scenario_rows, step_rows, sql_blocks


def looks_like_sql(text: str) -> bool:
    t = normalize_space(text).lower()
    return bool(re.search(r"\b(select|insert|merge|update|delete|create|truncate|drop|alter|begin|commit)\b", t))


# ======================================================================================
# SQL parsing
# ======================================================================================

def normalize_odi_sql(sql: str) -> str:
    """
    Replace ODI macros with simple identifiers where possible.
    Example:
      <?= odiRef.getObjectName("L", "TABLE", "SCHEMA", "D") ?> -> SCHEMA.TABLE
    """
    s = sql

    def repl_getobj(m: re.Match) -> str:
        table = m.group(1)
        schema = m.group(2)
        if schema and schema.upper() not in {"D", "L"}:
            return f"{schema}.{table}"
        return table

    s = re.sub(
        r'<\?=\s*odiRef\.getObjectName\(\s*"[^"]*"\s*,\s*"([^"]+)"\s*,\s*"([^"]+)"\s*,\s*"[^"]*"\s*\)\s*\?>',
        repl_getobj,
        s,
        flags=re.I | re.S,
    )
    # Session variables become literals/identifiers.
    s = re.sub(r'<\?=.*?\?>', "ODI_RUNTIME_VALUE", s, flags=re.S)
    return s


def strip_sql_comments(sql: str) -> str:
    sql = normalize_odi_sql(sql)
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.S)
    sql = re.sub(r"--[^\n]*", " ", sql)
    return sql


def split_top_level_commas(text: str) -> List[str]:
    parts = []
    current = []
    depth = 0
    in_single = False
    in_double = False
    i = 0

    while i < len(text):
        ch = text[i]
        prev = text[i - 1] if i > 0 else ""

        if ch == "'" and not in_double and prev != "\\":
            # Handle escaped Oracle single quotes by checking doubled quote.
            if i + 1 < len(text) and text[i + 1] == "'":
                current.append(ch)
                current.append(text[i + 1])
                i += 2
                continue
            in_single = not in_single
        elif ch == '"' and not in_single and prev != "\\":
            in_double = not in_double
        elif not in_single and not in_double:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth = max(0, depth - 1)
            elif ch == "," and depth == 0:
                parts.append("".join(current).strip())
                current = []
                i += 1
                continue

        current.append(ch)
        i += 1

    if current:
        parts.append("".join(current).strip())
    return [p for p in parts if p]


def find_matching_paren(text: str, open_pos: int) -> int:
    depth = 0
    in_single = False
    in_double = False

    i = open_pos
    while i < len(text):
        ch = text[i]
        prev = text[i - 1] if i > 0 else ""
        if ch == "'" and not in_double and prev != "\\":
            if i + 1 < len(text) and text[i + 1] == "'":
                i += 2
                continue
            in_single = not in_single
        elif ch == '"' and not in_single and prev != "\\":
            in_double = not in_double
        elif not in_single and not in_double:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    return i
        i += 1

    return -1


def extract_insert_columns(sql: str) -> List[str]:
    s = strip_sql_comments(sql)
    matches = list(re.finditer(r"\binsert\b", s, flags=re.I))
    if not matches:
        return []

    # Last INSERT often corresponds to the actual final MERGE INSERT or Insert Rows.
    for m in reversed(matches):
        open_pos = s.find("(", m.end())
        if open_pos < 0:
            continue
        end = find_matching_paren(s, open_pos)
        if end < 0:
            continue
        cols = [normalize_identifier(c) for c in split_top_level_commas(s[open_pos + 1:end])]
        cols = [c for c in cols if c]
        # A real target column list usually has several identifiers.
        if len(cols) >= 2:
            return cols
    return []


def extract_update_set_pairs(sql: str) -> List[Tuple[str, str]]:
    s = strip_sql_comments(sql)
    m = re.search(r"\bupdate\s+set\b", s, flags=re.I)
    if not m:
        m = re.search(r"\bset\b", s, flags=re.I)
        if not m:
            return []
    start = m.end()
    # End at WHEN NOT MATCHED / WHERE / DELETE etc.
    end_candidates = []
    for pat in [r"\bwhen\s+not\s+matched\b", r"\bwhere\b", r"\bdelete\b"]:
        mm = re.search(pat, s[start:], flags=re.I)
        if mm:
            end_candidates.append(start + mm.start())
    end = min(end_candidates) if end_candidates else len(s)
    assignments = split_top_level_commas(s[start:end])
    pairs = []
    for a in assignments:
        if "=" in a:
            left, right = a.split("=", 1)
            left_id = normalize_identifier(left)
            if left_id:
                pairs.append((left_id, normalize_space(right)))
    return pairs


def extract_merge_values(sql: str) -> List[str]:
    s = strip_sql_comments(sql)
    m = re.search(r"\bvalues\s*\(", s, flags=re.I | re.S)
    if not m:
        return []
    open_pos = s.find("(", m.start())
    end = find_matching_paren(s, open_pos)
    if end < 0:
        return []
    return split_top_level_commas(s[open_pos + 1:end])


def is_word_boundary(text: str, pos: int) -> bool:
    if pos < 0 or pos >= len(text):
        return True
    return not (text[pos].isalnum() or text[pos] == "_")


def extract_select_list(sql: str) -> List[str]:
    s = strip_sql_comments(sql)
    lower = s.lower()
    select_match = re.search(r"\bselect\b", lower)
    if not select_match:
        return []

    start = select_match.end()
    depth = 0
    in_single = False
    in_double = False

    i = start
    while i < len(s):
        ch = s[i]
        prev = s[i - 1] if i > 0 else ""
        if ch == "'" and not in_double and prev != "\\":
            if i + 1 < len(s) and s[i + 1] == "'":
                i += 2
                continue
            in_single = not in_single
        elif ch == '"' and not in_single and prev != "\\":
            in_double = not in_double
        elif not in_single and not in_double:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth = max(0, depth - 1)
            elif depth == 0 and lower[i:i+4] == "from" and is_word_boundary(lower, i-1) and is_word_boundary(lower, i+4):
                return split_top_level_commas(s[start:i])
        i += 1

    return []


def extract_select_list_at(sql: str, select_start: int) -> Tuple[List[str], int]:
    """
    Extract SELECT list for a SELECT keyword at a specific position.
    Returns (expressions, from_position). from_position is -1 when not found.
    """
    s = strip_sql_comments(sql)
    lower = s.lower()
    if lower[select_start:select_start + 6] != "select":
        return [], -1

    start = select_start + 6
    depth = 0
    in_single = False
    in_double = False

    i = start
    while i < len(s):
        ch = s[i]
        prev = s[i - 1] if i > 0 else ""
        if ch == "'" and not in_double and prev != "\\":
            if i + 1 < len(s) and s[i + 1] == "'":
                i += 2
                continue
            in_single = not in_single
        elif ch == '"' and not in_single and prev != "\\":
            in_double = not in_double
        elif not in_single and not in_double:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth = max(0, depth - 1)
            elif depth == 0 and lower[i:i+4] == "from" and is_word_boundary(lower, i-1) and is_word_boundary(lower, i+4):
                return split_top_level_commas(s[start:i]), i
        i += 1

    return [], -1


def extract_alias(expr: str) -> str:
    expr = clean_text(expr)
    if not expr:
        return ""

    # AS alias
    m = re.search(r"\bas\s+([A-Za-z_][A-Za-z0-9_#$]*)\s*$", expr, flags=re.I)
    if m:
        return normalize_identifier(m.group(1))

    # Simple table.column
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_#$]*\.[A-Za-z_][A-Za-z0-9_#$]*", expr.strip()):
        return normalize_identifier(expr)

    # Trailing alias after expression, but avoid keywords.
    m = re.search(r"\s+([A-Za-z_][A-Za-z0-9_#$]*)\s*$", expr)
    if m:
        alias = m.group(1)
        if alias.upper() not in {"END", "NULL", "DUAL"}:
            return normalize_identifier(alias)

    return normalize_identifier(expr)


def extract_tables(sql: str) -> List[str]:
    s = strip_sql_comments(sql)
    tables = []
    patterns = [
        r"\bfrom\s+([A-Za-z0-9_#$\.]+)",
        r"\bjoin\s+([A-Za-z0-9_#$\.]+)",
        r"\bmerge\s+into\s+([A-Za-z0-9_#$\.]+)",
        r"\binsert\s+into\s+([A-Za-z0-9_#$\.]+)",
        r"\bupdate\s+([A-Za-z0-9_#$\.]+)",
    ]
    for p in patterns:
        for m in re.finditer(p, s, flags=re.I):
            tables.append(m.group(1).upper())
    return unique_keep_order(tables)




def extract_sql_context_for_expression(sql: str, expr: str, limit: int = 5000) -> str:
    """
    Build a fuller human-review context for a target expression:
    - expression used for the target column
    - WITH clause if present
    - FROM / JOIN / WHERE / USING context from the ODI SQL block

    This intentionally keeps more context than xml_expressions because review of
    JOIN/CASE/lookup differences requires source path, not only final alias.
    """
    s = normalize_odi_sql(sql)
    s_clean = strip_sql_comments(s)
    expr_clean = normalize_space(expr)

    parts = []
    if expr_clean:
        parts.append("Target expression:\n" + expr_clean)

    # WITH clause, if present, up to first top-level SELECT.
    m_with = re.search(r"\bwith\b", s_clean, flags=re.I)
    m_select = re.search(r"\bselect\b", s_clean, flags=re.I)
    if m_with and m_select and m_with.start() < m_select.start():
        with_text = s_clean[m_with.start():m_select.start()]
        parts.append("WITH / CTE context:\n" + normalize_space(with_text))

    # Try to include FROM/JOIN/WHERE context from the SQL containing the expression.
    lower = s_clean.lower()
    from_match = re.search(r"\bfrom\b", lower, flags=re.I)
    if from_match:
        context = s_clean[from_match.start():]
        # Stop before merge update/insert control if context is huge.
        stop_patterns = [
            r"\bwhen\s+matched\b",
            r"\bwhen\s+not\s+matched\b",
            r"\bcommit\b",
            r"\bend;\b",
        ]
        stops = []
        for pat in stop_patterns:
            mm = re.search(pat, context, flags=re.I)
            if mm and mm.start() > 0:
                stops.append(mm.start())
        if stops:
            context = context[:min(stops)]
        parts.append("FROM / JOIN / WHERE context:\n" + normalize_space(context))

    out = "\n\n".join(parts)
    if len(out) > limit:
        out = out[:limit - 30] + "\n... [truncated]"
    return out


def extract_final_using_context(sql: str, limit: int = 6000) -> str:
    """
    For MERGE statements, extract USING (...) or final SELECT context.
    Useful when final VALUES are S.COL aliases and actual logic lives in USING SELECT.
    """
    s = normalize_odi_sql(sql)
    s_clean = strip_sql_comments(s)

    # MERGE USING ( ... ) S
    m = re.search(r"\busing\s*\(", s_clean, flags=re.I)
    if m:
        open_pos = s_clean.find("(", m.start())
        end = find_matching_paren(s_clean, open_pos)
        if end > open_pos:
            using_text = s_clean[open_pos + 1:end]
            if len(using_text) > limit:
                using_text = using_text[:limit - 30] + "\n... [truncated]"
            return normalize_space(using_text)

    # Fallback: main SELECT context.
    select_idx = re.search(r"\bselect\b", s_clean, flags=re.I)
    if select_idx:
        ctx = s_clean[select_idx.start():]
        if len(ctx) > limit:
            ctx = ctx[:limit - 30] + "\n... [truncated]"
        return normalize_space(ctx)
    return ""





def expression_table_aliases(expr: str) -> List[str]:
    """Return table aliases referenced in a SQL expression, e.g. ACG_TP_DIM.COL -> ACG_TP_DIM."""
    expr = normalize_odi_sql(expr)
    aliases = re.findall(r"\b([A-Za-z_][A-Za-z0-9_#$]*)\s*\.", expr)
    # Remove SQL/schema-ish noise.
    skip = {"ODIREF", "SYS", "TO_DATE", "TO_CHAR"}
    return unique_keep_order(a.upper() for a in aliases if a.upper() not in skip)


def compact_relevant_sql_context(expr: str, context: str, max_len: int = 2500) -> str:
    """
    Keep only source/JOIN/lookup context relevant to one attribute expression.

    Rules:
    - Always show the exact attribute expression.
    - Show the driving source.
    - Show a JOIN only if the *joined alias* is directly referenced by the expression.
      Do NOT include every join just because the base source alias appears in ON clauses.
    - Show a filter only if it references an alias used by the expression.
    """
    expr_norm = normalize_space(expr)
    context_norm = normalize_space(context)
    aliases = expression_table_aliases(expr_norm)

    parts = ["ODI attribute expression:\n" + expr_norm]

    if not context_norm:
        return "\n\n".join(parts)

    join_pat = r"\b(?:left\s+outer\s+join|left\s+join|inner\s+join|right\s+join|full\s+join|join)\b"
    m_join = re.search(join_pat, context_norm, flags=re.I)
    m_where = re.search(r"\bwhere\b", context_norm, flags=re.I)

    cut_points = [m.start() for m in [m_join, m_where] if m]
    base_part = context_norm[:min(cut_points)] if cut_points else context_norm
    base_part = normalize_space(base_part)

    # Infer base/driving alias from "FROM schema.table alias".
    base_alias = ""
    if base_part:
        tokens = re.findall(r"[A-Za-z_][A-Za-z0-9_#$]*", base_part)
        if tokens:
            base_alias = tokens[-1].upper()
        parts.append("Driving source:\n" + short(base_part, 500))

    non_base_aliases = [a for a in aliases if a.upper() != base_alias]

    def joined_alias(seg: str) -> str:
        m = re.search(r"\bjoin\b\s+(.+?)\s+\bon\b", seg, flags=re.I)
        if not m:
            return ""
        between = m.group(1)
        # Remove nested SELECT text as much as possible; alias is usually the last identifier before ON.
        tokens = re.findall(r"[A-Za-z_][A-Za-z0-9_#$]*", between)
        if not tokens:
            return ""
        return tokens[-1].upper()

    # Split JOIN segments and keep only the joined alias used by this expression.
    join_segments = []
    matches = list(re.finditer(join_pat, context_norm, flags=re.I))
    for i, m in enumerate(matches):
        start = m.start()
        end_candidates = []
        if i + 1 < len(matches):
            end_candidates.append(matches[i + 1].start())
        where_after = re.search(r"\bwhere\b", context_norm[start:], flags=re.I)
        if where_after:
            end_candidates.append(start + where_after.start())
        end = min(end_candidates) if end_candidates else len(context_norm)
        seg = context_norm[start:end].strip()
        if seg:
            join_segments.append(seg)

    relevant_joins = []
    for seg in join_segments:
        ja = joined_alias(seg)
        if ja and ja in non_base_aliases:
            relevant_joins.append(seg)

    if relevant_joins:
        parts.append("Relevant JOIN / lookup context:\n" + "\n".join(short(j, 1200) for j in relevant_joins[:4]))

    # Do not add a generic WHERE block here. In ODI generated SQL the first WHERE may
    # belong to an unrelated inline view and can pollute every attribute. If a lookup
    # has a filter such as CL_SCM_ID, it is preserved inside the relevant JOIN segment.

    out = "\n\n".join(parts)
    if len(out) > max_len:
        out = out[:max_len - 30] + "\n... [truncated]"
    return out

def build_select_alias_logic_map(sql: str) -> Dict[str, str]:
    """
    Map SELECT aliases to compact, attribute-specific logic.

    For MERGE final values like S.ACG_TP_NM, actual logic lives in USING SELECT.
    This function scans nested SELECT blocks and stores, per alias:
    - the exact expression that creates the alias
    - only the relevant join/filter context for aliases used by that expression
    """
    s = normalize_odi_sql(sql)
    s_clean = strip_sql_comments(s)
    alias_map: Dict[str, str] = {}

    select_positions = [m.start() for m in re.finditer(r"\bselect\b", s_clean, flags=re.I)]
    for pos in select_positions:
        select_exprs, from_pos = extract_select_list_at(s_clean, pos)
        if not select_exprs:
            continue

        context = ""
        if from_pos >= 0:
            context = s_clean[from_pos:]
            stops = []
            for pat in [r"\bwhen\s+matched\b", r"\bwhen\s+not\s+matched\b"]:
                mm = re.search(pat, context, flags=re.I)
                if mm and mm.start() > 0:
                    stops.append(mm.start())
            if stops:
                context = context[:min(stops)]
            context = normalize_space(context)

        for expr in select_exprs:
            alias = extract_alias(expr)
            if not alias:
                continue

            expr_norm = normalize_space(expr)
            simple_id = normalize_identifier(expr_norm)
            aliases = expression_table_aliases(expr_norm)
            is_pass_through = (simple_id == alias and not aliases and " " not in expr_norm)

            # Prefer real source expressions over outer pass-through aliases.
            # Example: prefer "ACG_TP_DIM.ACG_TP_NM ACG_TP_NM" over "ACG_TP_NM".
            richness = 0
            if not is_pass_through:
                richness += 100000
            if aliases:
                richness += 10000
            if re.search(r"\bcase\b|\bwhen\b|\bnvl\b|\bcoalesce\b|\bdecode\b|\bsubstr\b|\binstr\b", expr_norm, re.I):
                richness += 5000
            richness += len(expr_norm)

            logic = compact_relevant_sql_context(expr_norm, context, max_len=3000)

            prev = alias_map.get(alias)
            prev_score = -1
            if prev:
                # Roughly infer previous score from whether it had real alias/join context.
                prev_score = len(prev)
                if "Relevant JOIN / lookup context" in prev:
                    prev_score += 10000
                if "ODI attribute expression:\n" in prev and "." in prev.split("ODI attribute expression:\n", 1)[1].split("\n", 1)[0]:
                    prev_score += 100000

            if not prev or richness > prev_score:
                alias_map[alias] = logic

    return alias_map

def build_odi_lineage(sql_blocks: List[Dict[str, str]]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []

    for block in sql_blocks:
        raw_sql = block["sql"]
        sql = normalize_odi_sql(raw_sql)
        select_alias_logic_map = None

        def get_select_alias_logic_map() -> Dict[str, str]:
            nonlocal select_alias_logic_map
            if select_alias_logic_map is None:
                select_alias_logic_map = build_select_alias_logic_map(sql)
            return select_alias_logic_map

        tables = extract_tables(sql)
        insert_cols = extract_insert_columns(sql)
        merge_values = extract_merge_values(sql)
        select_exprs = extract_select_list(sql)
        update_pairs = extract_update_set_pairs(sql)

        base = {k: block.get(k, "") for k in ["step_no", "task_no", "task_name_1", "task_name_2", "task_name_3", "field"]}

        # MERGE update SET pairs.
        for col, expr in update_pairs:
            rows.append({
                **base,
                "target_column": col,
                "expression": expr,
                "alias": col,
                "source_tables": " | ".join(tables),
                "lineage_type": "merge_update_set",
                "has_case": "Y" if re.search(r"\bcase\b", expr, flags=re.I) else "N",
                "has_join": "Y" if re.search(r"\bjoin\b|\(\+\)", sql, flags=re.I) else "N",
                "sql_excerpt": short(sql, 1200),
                "xml_logic_full": get_select_alias_logic_map().get(normalize_identifier(expr), extract_sql_context_for_expression(sql, expr, 5000)),
            })

        # MERGE INSERT columns + VALUES.
        if insert_cols and merge_values and len(insert_cols) == len(merge_values):
            for col, expr in zip(insert_cols, merge_values):
                rows.append({
                    **base,
                    "target_column": col,
                    "expression": normalize_space(expr),
                    "alias": col,
                    "source_tables": " | ".join(tables),
                    "lineage_type": "merge_insert_values",
                    "has_case": "Y" if re.search(r"\bcase\b", expr, flags=re.I) else "N",
                    "has_join": "Y" if re.search(r"\bjoin\b|\(\+\)", sql, flags=re.I) else "N",
                    "sql_excerpt": short(sql, 1200),
                    "xml_logic_full": get_select_alias_logic_map().get(normalize_identifier(expr), extract_sql_context_for_expression(sql, expr, 5000)),
                })
            continue

        # INSERT INTO (cols) SELECT exprs.
        # Here target columns and SELECT expressions are already position-aligned,
        # so use the expression directly. Do NOT resolve through alias_map; that is
        # only needed for MERGE final values like S.COL.
        if insert_cols and select_exprs and len(insert_cols) == len(select_exprs):
            for col, expr in zip(insert_cols, select_exprs):
                rows.append({
                    **base,
                    "target_column": col,
                    "expression": normalize_space(expr),
                    "alias": extract_alias(expr) or col,
                    "source_tables": " | ".join(tables),
                    "lineage_type": "insert_select",
                    "has_case": "Y" if re.search(r"\bcase\b", expr, flags=re.I) else "N",
                    "has_join": "Y" if re.search(r"\bjoin\b|\(\+\)", sql, flags=re.I) else "N",
                    "sql_excerpt": short(sql, 1200),
                    "xml_logic_full": extract_sql_context_for_expression(sql, expr, 5000),
                })
            continue

        # SELECT aliases for stage creation/enrichment.
        if select_exprs:
            for expr in select_exprs:
                alias = extract_alias(expr)
                if not alias:
                    continue
                rows.append({
                    **base,
                    "target_column": alias,
                    "expression": normalize_space(expr),
                    "alias": alias,
                    "source_tables": " | ".join(tables),
                    "lineage_type": "select_alias",
                    "has_case": "Y" if re.search(r"\bcase\b", expr, flags=re.I) else "N",
                    "has_join": "Y" if re.search(r"\bjoin\b|\(\+\)", sql, flags=re.I) else "N",
                    "sql_excerpt": short(sql, 1200),
                    "xml_logic_full": get_select_alias_logic_map().get(alias, extract_sql_context_for_expression(sql, expr, 5000)),
                })

    return rows


# ======================================================================================
# DRD auto-detection and extraction
# ======================================================================================

@dataclass
class MappingDetection:
    mapping_sheet: str
    header_row: int
    target_col: int
    source_schema_col: Optional[int]
    source_table_col: Optional[int]
    source_attribute_col: Optional[int]
    rule_col: Optional[int]
    notes_col: Optional[int]
    confidence: float
    target_table_from_sheet: str
    target_resources_from_xml: List[str]

    def as_human(self) -> Dict[str, str]:
        return {
            "mapping_sheet": self.mapping_sheet,
            "header_row": str(self.header_row),
            "target_col": index_to_col_letter(self.target_col),
            "source_schema_col": index_to_col_letter(self.source_schema_col) if self.source_schema_col else "",
            "source_table_col": index_to_col_letter(self.source_table_col) if self.source_table_col else "",
            "source_attribute_col": index_to_col_letter(self.source_attribute_col) if self.source_attribute_col else "",
            "rule_col": index_to_col_letter(self.rule_col) if self.rule_col else "",
            "notes_col": index_to_col_letter(self.notes_col) if self.notes_col else "",
            "confidence": f"{self.confidence:.2f}",
            "target_table_from_sheet": self.target_table_from_sheet,
            "target_resources_from_xml": " | ".join(self.target_resources_from_xml),
        }


def get_cell(ws, row: int, col: int) -> str:
    return clean_text(ws.cell(row, col).value)


def detect_table_name_in_sheet(ws) -> str:
    for r in range(1, min(ws.max_row, 15) + 1):
        row_text = " ".join(get_cell(ws, r, c) for c in range(1, min(ws.max_column, 10) + 1))
        if "table name" in row_text.lower():
            # Prefer a physical uppercase table name from row cells.
            candidates = []
            for c in range(1, min(ws.max_column, 10) + 1):
                v = normalize_identifier(get_cell(ws, r, c))
                if v and v not in {"TABLENAME", "FROMDATEAM"} and "_" in v:
                    candidates.append(v)
            if candidates:
                return candidates[-1]
    return ""


def header_score_row(values: List[str]) -> float:
    text = " ".join(v.lower() for v in values if v)
    score = 0.0
    if "physical name" in text and "attribute" in text:
        score += 4
    if "name of column in table" in text:
        score += 4
    if "source schema" in text:
        score += 2
    if "source table" in text:
        score += 2
    if "source attribute" in text:
        score += 2
    if "transformation" in text:
        score += 3
    if "business rules" in text or "join conditions" in text:
        score += 2
    if "logical name of attribute" in text:
        score += 1
    return score


def find_header_row(ws) -> Tuple[int, float]:
    best_row = 1
    best_score = -1.0
    for r in range(1, min(ws.max_row, 40) + 1):
        values = [get_cell(ws, r, c) for c in range(1, ws.max_column + 1)]
        score = header_score_row(values)
        if score > best_score:
            best_row = r
            best_score = score
    return best_row, best_score


def header_contains(header: str, *needles: str) -> bool:
    h = normalize_space(header).lower()
    return all(n.lower() in h for n in needles)


def detect_columns(ws, header_row: int) -> Tuple[int, Optional[int], Optional[int], Optional[int], Optional[int], Optional[int], float]:
    headers = {c: get_cell(ws, header_row, c) for c in range(1, ws.max_column + 1)}
    target_col = None
    source_schema_col = None
    source_table_col = None
    source_attribute_col = None
    rule_col = None
    notes_col = None
    score = 0.0

    # Target column detection.
    for c, h in headers.items():
        hl = normalize_space(h).lower()
        if "physical name of attribute" in hl and "table" in hl:
            target_col = c
            score += 4
            break
    if not target_col:
        for c, h in headers.items():
            hl = normalize_space(h).lower()
            if "name of column in table" in hl or ("physical name" in hl and "attribute" in hl):
                target_col = c
                score += 3
                break

    # Source columns.
    for c, h in headers.items():
        hl = normalize_space(h).lower()
        if "source schema" in hl:
            source_schema_col = c
            score += 2
        elif "source table" in hl:
            source_table_col = c
            score += 2
        elif "source attribute" in hl:
            source_attribute_col = c
            score += 2

    # Transformation / rule column.
    for c, h in headers.items():
        hl = normalize_space(h).lower()
        if "transformation" in hl and ("business" in hl or "rule" in hl or "join" in hl):
            rule_col = c
            score += 3
            break
    if not rule_col:
        for c, h in headers.items():
            hl = normalize_space(h).lower()
            if "transformation" in hl:
                rule_col = c
                score += 2
                break

    # Notes column.
    for c, h in headers.items():
        hl = normalize_space(h).lower()
        if "notes" in hl or "comments" in hl:
            notes_col = c
            break

    return target_col or 0, source_schema_col, source_table_col, source_attribute_col, rule_col, notes_col, score


def sheet_score(wb, sheet_name: str, xml_targets: List[str], target_table_override: str = "") -> Tuple[float, int, str]:
    ws = wb[sheet_name]
    header_row, header_score = find_header_row(ws)
    sheet_target = detect_table_name_in_sheet(ws)
    score = header_score

    lname = sheet_name.lower()
    if "table-view" in lname or "table view" in lname or "table_view" in lname:
        score += 2
    if sheet_name.lower() == "table-view":
        score += 0.5

    compare_targets = [normalize_identifier(target_table_override)] if target_table_override else xml_targets
    if sheet_target and compare_targets:
        for t in compare_targets:
            if sheet_target == t:
                score += 8
            elif sheet_target in t or t in sheet_target:
                score += 3

    return score, header_row, sheet_target


def auto_detect_mapping(wb, xml_targets: List[str], target_table_override: str = "", mapping_sheet_override: str = "", header_row_override: Optional[int] = None,
                        target_col_override: str = "", source_cols_override: str = "", rule_col_override: str = "") -> MappingDetection:
    # Sheet selection.
    if mapping_sheet_override:
        if mapping_sheet_override not in wb.sheetnames:
            # case-insensitive fallback
            matches = [s for s in wb.sheetnames if s.lower() == mapping_sheet_override.lower()]
            if not matches:
                raise ValueError(f"Mapping sheet '{mapping_sheet_override}' not found. Available: {wb.sheetnames}")
            sheet_name = matches[0]
        else:
            sheet_name = mapping_sheet_override
        ws = wb[sheet_name]
        header_row, hscore = find_header_row(ws)
        sheet_target = detect_table_name_in_sheet(ws)
        base_score = hscore
    else:
        candidates = []
        for s in wb.sheetnames:
            score, header_row, sheet_target = sheet_score(wb, s, xml_targets, target_table_override)
            if score >= 5:
                candidates.append((score, s, header_row, sheet_target))
        if not candidates:
            raise ValueError("Could not auto-detect mapping sheet. Use --mapping-sheet.")
        candidates.sort(reverse=True)
        base_score, sheet_name, header_row, sheet_target = candidates[0]
        ws = wb[sheet_name]

    if header_row_override:
        header_row = header_row_override

    target_col, src_schema, src_table, src_attr, rule_col, notes_col, col_score = detect_columns(ws, header_row)

    # Manual overrides.
    if target_col_override:
        target_col = col_letter_to_index(target_col_override)
    if source_cols_override:
        source_indexes = [col_letter_to_index(c.strip()) for c in source_cols_override.split(",") if c.strip()]
        src_schema = source_indexes[0] if len(source_indexes) > 0 else None
        src_table = source_indexes[1] if len(source_indexes) > 1 else None
        src_attr = source_indexes[2] if len(source_indexes) > 2 else None
    if rule_col_override:
        rule_col = col_letter_to_index(rule_col_override)

    if not target_col:
        raise ValueError("Could not auto-detect target column. Use --target-col.")
    if not rule_col:
        # Keep running, but report as missing.
        rule_col = None

    confidence = min(1.0, (base_score + col_score) / 25.0)
    return MappingDetection(
        mapping_sheet=sheet_name,
        header_row=header_row,
        target_col=target_col,
        source_schema_col=src_schema,
        source_table_col=src_table,
        source_attribute_col=src_attr,
        rule_col=rule_col,
        notes_col=notes_col,
        confidence=confidence,
        target_table_from_sheet=sheet_target,
        target_resources_from_xml=xml_targets,
    )


def extract_mapping_from_xlsx(xlsx_path: Path, detection: MappingDetection, notes_sheet: str = "ETL Notes") -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    """Fast extraction using iter_rows, not ws.cell loops."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[detection.mapping_sheet]

    mapping_rows: List[Dict[str, str]] = []
    src_cols = [detection.source_schema_col, detection.source_table_col, detection.source_attribute_col]
    src_cols = [c for c in src_cols if c]

    needed_cols = [detection.target_col] + src_cols
    if detection.rule_col:
        needed_cols.append(detection.rule_col)
    if detection.notes_col:
        needed_cols.append(detection.notes_col)
    max_needed = max(needed_cols) if needed_cols else ws.max_column

    # Iterate mapping rows with a manual Excel row counter.
    for row_num, row in enumerate(ws.iter_rows(min_row=detection.header_row + 1, max_col=max_needed, values_only=True), start=detection.header_row + 1):
        def val(col: Optional[int]) -> str:
            if not col:
                return ""
            idx = col - 1
            if idx < 0 or idx >= len(row):
                return ""
            return clean_text(row[idx])

        raw_target = val(detection.target_col)
        target_column = normalize_identifier(raw_target)
        if not target_column:
            continue
        if not is_probable_column_name(raw_target):
            continue

        source_values = [val(c) for c in src_cols]
        rule = val(detection.rule_col) if detection.rule_col else ""
        notes = val(detection.notes_col) if detection.notes_col else ""

        mapping_rows.append({
            "excel_row": str(row_num),
            "target_column_raw": raw_target,
            "target_column": target_column,
            "source_1": source_values[0] if len(source_values) > 0 else "",
            "source_2": source_values[1] if len(source_values) > 1 else "",
            "source_3": source_values[2] if len(source_values) > 2 else "",
            "source_tables_combined": " | ".join([s for s in source_values if s]),
            "drd_rule": rule,
            "drd_notes": notes,
        })

    notes_rows: List[Dict[str, str]] = []
    sheet_match = None
    if notes_sheet in wb.sheetnames:
        sheet_match = notes_sheet
    else:
        for sh in wb.sheetnames:
            if sh.lower().replace(" ", "") == notes_sheet.lower().replace(" ", ""):
                sheet_match = sh
                break
    if sheet_match:
        nws = wb[sheet_match]
        for rnum, row in enumerate(nws.iter_rows(values_only=True), start=1):
            values = [clean_text(v) for v in row]
            if any(values):
                out = {"excel_row": str(rnum), "row_text": " | ".join([v for v in values if v])}
                for i, v in enumerate(values, start=1):
                    if v:
                        out[f"col_{index_to_col_letter(i)}"] = v
                notes_rows.append(out)

    return mapping_rows, notes_rows


# ======================================================================================
# Compare and report
# ======================================================================================



def select_final_target_lineage(lineage_rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Pick only the lineage rows that represent the final target load.

    Why this matters:
    ODI XML contains many intermediate aliases from staging inserts, exception tables,
    variables, journal procedures, and session logging. Those are useful in
    odi_column_lineage_best_effort.csv, but they should not be treated as target columns
    in mapping_vs_xml_column_diff.csv.

    Selection rule:
    1. Prefer the largest MERGE INSERT VALUES group.
    2. If no MERGE INSERT exists, use the largest INSERT SELECT group.
    3. Deduplicate by target_column while preserving the selected group.
    """
    groups: Dict[Tuple[str, str, str], List[Dict[str, str]]] = defaultdict(list)
    for r in lineage_rows:
        lt = r.get("lineage_type", "")
        if lt not in {"merge_insert_values", "insert_select"}:
            continue
        key = (r.get("step_no", ""), r.get("task_no", ""), lt)
        groups[key].append(r)

    if not groups:
        return lineage_rows

    def step_int(key):
        try:
            return int(key[0])
        except Exception:
            return -1

    # Prefer final MERGE target insert.
    merge_groups = {k: v for k, v in groups.items() if k[2] == "merge_insert_values"}
    if merge_groups:
        best_key = max(merge_groups.keys(), key=lambda k: (len({r.get("target_column","") for r in merge_groups[k]}), step_int(k), int(k[1] or 0)))
        selected = merge_groups[best_key]
    else:
        insert_groups = {k: v for k, v in groups.items() if k[2] == "insert_select"}
        best_key = max(insert_groups.keys(), key=lambda k: (len({r.get("target_column","") for r in insert_groups[k]}), step_int(k), int(k[1] or 0)))
        selected = insert_groups[best_key]

    # Deduplicate columns, preserving the first selected expression.
    seen = set()
    out = []
    for r in selected:
        col = r.get("target_column", "")
        if not col or col in seen:
            continue
        seen.add(col)
        out.append(r)
    return out

def compare_columns(mapping_rows: List[Dict[str, str]], lineage_rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    mapping_by_col = {r["target_column"]: r for r in mapping_rows if r.get("target_column")}
    xml_by_col: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for r in lineage_rows:
        col = normalize_identifier(r.get("target_column", ""))
        if col:
            xml_by_col[col].append(r)

    all_cols = sorted(set(mapping_by_col) | set(xml_by_col))
    rows: List[Dict[str, str]] = []
    for col in all_cols:
        m = mapping_by_col.get(col)
        xs = xml_by_col.get(col, [])
        if m and xs:
            status = "IN_BOTH"
        elif m and not xs:
            status = "MAPPING_ONLY"
        else:
            status = "XML_ONLY"

        rows.append({
            "target_column": col,
            "status": status,
            "mapping_excel_row": m.get("excel_row", "") if m else "",
            "mapping_source_tables": m.get("source_tables_combined", "") if m else "",
            "drd_rule": m.get("drd_rule", "") if m else "",
            "drd_notes": m.get("drd_notes", "") if m else "",
            "xml_occurrences": str(len(xs)),
            "xml_steps": ", ".join(sorted({x.get("step_no", "") for x in xs if x.get("step_no", "")})),
            "xml_lineage_types": ", ".join(sorted({x.get("lineage_type", "") for x in xs if x.get("lineage_type", "")})),
            "xml_expressions": " || ".join(short(x.get("expression", ""), 400) for x in xs[:8]),
            "xml_source_tables": " || ".join(short(x.get("source_tables", ""), 300) for x in xs[:5]),
            "xml_logic_full": "\n\n---\n\n".join(short(x.get("xml_logic_full", ""), 5000) for x in xs[:3] if x.get("xml_logic_full", "")),
        })

    return rows


LOGIC_KEYWORDS = re.compile(
    r"\b(case|when|then|else|end|join|where|lookup|coalesce|nvl|decode|cl_val|src_stm_id|src_stm|regexp|like|in\s*\(|exists|substr|instr|trim|round|to_date|to_char)\b|\(\+\)",
    flags=re.I,
)


def extract_condition_tokens(text: str) -> List[str]:
    """Fast conservative condition-token extractor. Bounded regexes avoid catastrophic backtracking."""
    text = normalize_space(text).upper()
    if len(text) > 8000:
        text = text[:8000]
    patterns = [
        r"\bSRC_STM_ID\s*(?:=|<>|!=)\s*'?\w{1,40}'?",
        r"\bSRC_STM_ID\s+IN\s*\([^\)]{0,200}\)",
        r"\bCL_SCM_ID\s*=\s*\d+",
        r"\b[A-Z0-9_\.]{1,80}\s+LIKE\s+'[^']{0,120}'",
        r"\b[A-Z0-9_\.]{1,80}\s*(?:=|<>|!=)\s*'[^']{0,120}'",
        r"\b[A-Z0-9_\.]{1,80}\s+IS\s+NULL",
        r"\b[A-Z0-9_\.]{1,80}\s+IS\s+NOT\s+NULL",
    ]
    tokens: List[str] = []
    for pat in patterns:
        for m in re.finditer(pat, text, flags=re.I):
            tokens.append(normalize_space(m.group(0).upper()))
    return unique_keep_order(tokens)


def extract_table_tokens(text: str) -> List[str]:
    text = normalize_space(text).upper()
    tokens = re.findall(r"\b[A-Z][A-Z0-9_#$]*\.[A-Z][A-Z0-9_#$]*\b", text)
    tokens += re.findall(
        r"\b(CL_VAL|TXN|APA|APA_CASH|APA_SECURITY|ACATS_BROKER|AR_DIM|BKR_AR_DIM|AR_GRP_SUBDIM|CCY|J\$[A-Z0-9_#$]+|ACG_TP_DIM|CCAL_CIRD_PD_MAP|SRC_STM_DIM|CLOSE_TAX_LOT_NONBKR_RJTRUST_TGT|OPN_TAX_LOTS_NONBKR_TGT)\b",
        text,
    )
    return unique_keep_order(tokens)


def build_logic_diff_candidates(mapping_rows: List[Dict[str, str]], lineage_rows: List[Dict[str, str]], sql_blocks: List[Dict[str, str]]) -> List[Dict[str, str]]:
    xml_by_col: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for r in lineage_rows:
        col = normalize_identifier(r.get("target_column", ""))
        if col:
            xml_by_col[col].append(r)

    all_sql_text_by_step = {
        (b.get("step_no", ""), b.get("task_no", "")): short(normalize_odi_sql(b.get("sql", "")), 5000)
        for b in sql_blocks
    }

    rows: List[Dict[str, str]] = []
    for m in mapping_rows:
        col = m["target_column"]
        drd = m.get("drd_rule", "")
        if not LOGIC_KEYWORDS.search(drd):
            continue

        xs = xml_by_col.get(col, [])
        xml_expr = " || ".join(x.get("expression", "") for x in xs[:10])
        xml_sql = " ".join(all_sql_text_by_step.get((x.get("step_no", ""), x.get("task_no", "")), "") for x in xs[:3])
        xml_combined = short(normalize_space(xml_expr + " " + xml_sql), 8000)

        drd_conditions = extract_condition_tokens(drd)
        xml_conditions = extract_condition_tokens(xml_combined)
        drd_tables = extract_table_tokens(m.get("source_tables_combined", "") + " " + drd)
        xml_tables = extract_table_tokens(xml_combined + " " + " ".join(x.get("source_tables", "") for x in xs[:5]))

        missing_conditions = [c for c in drd_conditions if c not in xml_conditions]
        missing_tables = []
        for t in drd_tables:
            t_suffix = t.split(".")[-1]
            if not any(t == xt or t_suffix == xt.split(".")[-1] for xt in xml_tables):
                missing_tables.append(t)

        has_drd_case = "Y" if re.search(r"\bcase\b|\bwhen\b", drd, flags=re.I) else "N"
        has_xml_case = "Y" if re.search(r"\bcase\b|\bwhen\b", xml_combined, flags=re.I) else "N"
        has_drd_join = "Y" if re.search(r"\bjoin\b|\bwhere\b|\blookup\b", drd, flags=re.I) else "N"
        has_xml_join = "Y" if re.search(r"\bjoin\b|\(\+\)", xml_combined, flags=re.I) else "N"

        if not xs:
            severity = "HIGH"
            conclusion = "DRD has join/case/lookup logic, but no ODI XML lineage was found for this target column."
        elif missing_conditions or missing_tables or (has_drd_case == "Y" and has_xml_case == "N"):
            severity = "MEDIUM"
            conclusion = "Potential logic drift: some DRD conditions/tables/case patterns are not visible in the XML expression."
        else:
            severity = "LOW"
            conclusion = "No obvious static mismatch found; review still recommended for semantic equivalence."

        rows.append({
            "target_column": col,
            "severity": severity,
            "conclusion": conclusion,
            "mapping_excel_row": m.get("excel_row", ""),
            "drd_source_tables": m.get("source_tables_combined", ""),
            "drd_rule": drd,
            "drd_notes": m.get("drd_notes", ""),
            "xml_steps": ", ".join(sorted({x.get("step_no", "") for x in xs if x.get("step_no", "")})),
            "xml_expression_excerpt": short(xml_expr, 1200),
            "drd_conditions": " | ".join(drd_conditions),
            "xml_conditions": " | ".join(xml_conditions),
            "missing_drd_conditions_in_xml": " | ".join(missing_conditions),
            "drd_tables": " | ".join(drd_tables),
            "xml_tables": " | ".join(xml_tables),
            "missing_drd_tables_in_xml": " | ".join(missing_tables),
            "drd_has_case": has_drd_case,
            "xml_has_case": has_xml_case,
            "drd_has_join_or_where_or_lookup": has_drd_join,
            "xml_has_join": has_xml_join,
        })

    severity_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    rows.sort(key=lambda r: (severity_order.get(r["severity"], 9), r["target_column"]))
    return rows


def write_sql_blocks(path: Path, sql_blocks: List[Dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for b in sql_blocks:
            f.write("\n")
            f.write("-- " + "=" * 100 + "\n")
            f.write(
                f"-- Step {b.get('step_no')} | Task {b.get('task_no')} | "
                f"{b.get('task_name_1')} / {b.get('task_name_2')} / {b.get('task_name_3')} | {b.get('field')}\n"
            )
            f.write("-- " + "=" * 100 + "\n")
            f.write(normalize_odi_sql(b.get("sql", "")).rstrip())
            f.write("\n;\n")




def classify_difference(row: Dict[str, str]) -> Tuple[str, str, str]:
    """
    Convert raw diff row into business-readable difference type, conclusion, and action.
    This is heuristic but useful for review triage.
    """
    status = row.get("status", "")
    col = row.get("target_column", "")
    drd_rule = normalize_space(row.get("drd_rule", ""))
    xml_expr = normalize_space(row.get("xml_expressions", ""))
    mapping_sources = normalize_space(row.get("mapping_source_tables", ""))
    xml_sources = normalize_space(row.get("xml_source_tables", ""))

    drd_has_lookup = bool(re.search(r"\b(CL_VAL|lookup|join|where|src_stm_id|case|when|nvl|coalesce|decode|substr|instr|trim|regexp|like)\b", drd_rule, re.I))
    xml_has_case = bool(re.search(r"\bcase\b|\bwhen\b", xml_expr, re.I))
    drd_has_case = bool(re.search(r"\bcase\b|\bwhen\b", drd_rule, re.I))
    xml_missing = status == "MAPPING_ONLY"
    drd_missing = status == "XML_ONLY"

    if xml_missing:
        if drd_has_lookup:
            return (
                "Missing implementation",
                "**Real gap.** DRD contains transformation / lookup logic, but ODI XML lineage does not populate this target column.",
                "Add the target column and implement the DRD transformation / lookup logic in ODI, or document why this column is intentionally excluded."
            )
        return (
            "Missing target column",
            "**Structural gap.** Column exists in DRD but was not found in ODI XML lineage.",
            "Add the column to ODI target mapping or confirm it is intentionally out of scope."
        )

    if drd_missing:
        return (
            "XML-only column / alias",
            "**ODI-only lineage.** XML contains a target/alias not present in DRD mapping extraction.",
            "Confirm whether this is a technical/control column, staging alias, or missing DRD documentation."
        )

    # Both present.
    if drd_has_case and not xml_has_case:
        return (
            "Missing CASE logic",
            "**Potential logic drift.** DRD describes conditional logic, but the ODI expression does not visibly contain equivalent CASE/WHEN logic.",
            "Verify generated ODI SQL. If condition is not implemented elsewhere, add the CASE logic or update the DRD."
        )

    # Source-system condition in DRD but not visible in XML expression.
    drd_conditions = set(extract_condition_tokens(drd_rule))
    xml_conditions = set(extract_condition_tokens(xml_expr + " " + xml_sources))
    missing_conditions = sorted([c for c in drd_conditions if c not in xml_conditions])
    if missing_conditions:
        return (
            "Condition mismatch",
            "**Potential logic drift.** Some DRD conditions are not visible in the ODI expression / source context.",
            "Review the ODI SQL block and verify whether the missing condition is applied in a join, filter, or prior stage."
        )

    # CL_VAL/schema lookup mismatch.
    if re.search(r"\bCL_VAL\b|CL_SCM_ID", drd_rule, re.I) and not re.search(r"\bCL_VAL\b|CL_SCM_ID", xml_expr + " " + xml_sources, re.I):
        return (
            "Lookup mismatch",
            "**Potential lookup gap.** DRD expects a CL_VAL-style lookup, but ODI lineage does not visibly show it for this column.",
            "Validate lookup alias and CL_SCM_ID in ODI. Add the lookup or update DRD if the value is derived earlier."
        )

    # Expression differs materially.
    if drd_rule and xml_expr and normalize_space(drd_rule).upper() != normalize_space(xml_expr).upper():
        return (
            "Expression / implementation difference",
            "**Review required.** DRD and ODI both populate the column, but ODI expression differs from the textual DRD rule.",
            "Compare the detailed ODI SQL block with the DRD rule and classify as acceptable rewrite or defect."
        )

    return (
        "No obvious difference",
        "**No obvious static difference.** Column exists in both DRD and ODI lineage.",
        "No action unless business review identifies semantic drift."
    )




def build_avy_review_rules_diff(column_diff_rows: List[Dict[str, str]], logic_rows: List[Dict[str, str]], detection: MappingDetection) -> List[Dict[str, str]]:
    """
    Curated AVY_FACT review-level grouping.
    This produces the compact business-review table requested in the original AVY analysis,
    instead of hundreds of static parser candidates.
    """
    target = (detection.target_table_from_sheet or "").upper()
    resources = " ".join(detection.target_resources_from_xml).upper()
    if "AVY_FACT" not in target and "AVY_FACT" not in resources:
        return []

    by_col = {r.get("target_column", ""): r for r in column_diff_rows}
    logic_by_col = {r.get("target_column", ""): r for r in logic_rows}

    def q(cols: List[str]) -> str:
        return ", ".join(f"`{c}`" for c in cols)

    def mapping_logic(cols: List[str], fallback: str) -> str:
        parts = []
        for c in cols:
            r = by_col.get(c, {})
            rule = normalize_space(r.get("drd_rule", ""))
            src = normalize_space(r.get("mapping_source_tables", ""))
            if rule:
                parts.append(f"{c}: {rule}")
            elif src:
                parts.append(f"{c}: source {src}")
        return " | ".join(parts) if parts else fallback

    def odi_logic(cols: List[str], fallback: str) -> str:
        parts = []
        for c in cols:
            r = by_col.get(c, {})
            expr = normalize_space(r.get("xml_expressions", ""))
            steps = normalize_space(r.get("xml_steps", ""))
            status = r.get("status", "")
            if status == "MAPPING_ONLY":
                parts.append(f"{c}: not present in ODI XML lineage / final target mapping.")
            elif expr:
                parts.append(f"{c}: {expr}" + (f" [steps {steps}]" if steps else ""))
        return " | ".join(parts) if parts else fallback

    rows = []

    def add(area, dtype, mlogic, xlogic, conclusion, action):
        rows.append({
            "#": str(len(rows) + 1),
            "Area / Columns": area,
            "Difference Type": dtype,
            "Mapping Logic": mlogic,
            "ODI XML Logic": xlogic,
            "Conclusion": conclusion,
            "Recommended Action": action,
        })

    # 1-3 missing implementation groups.
    add(
        q(["STEP_IN_OUT_IND_CD", "STEP_IN_OUT_IND_NM"]),
        "Missing implementation",
        mapping_logic(["STEP_IN_OUT_IND_CD", "STEP_IN_OUT_IND_NM"], "Mapping expects lookup from CCAL_REPL_OWNER.CL_VAL with CL_SCM_ID = 114."),
        odi_logic(["STEP_IN_OUT_IND_CD", "STEP_IN_OUT_IND_NM"], "Columns are not present in the final XML insert/update target list."),
        "**Real gap.** Mapping logic is not implemented in the ODI final load.",
        "Add these columns to the final step and implement the required CL_VAL lookup.",
    )
    add(
        q(["SHRT_SALE_EXMPT_CD"]),
        "Missing implementation",
        mapping_logic(["SHRT_SALE_EXMPT_CD"], "Mapping expects CL_VAL.CL_VAL_CODE, likely using short-sale exemption lookup logic."),
        odi_logic(["SHRT_SALE_EXMPT_CD"], "Column is not present in the final XML insert/update target list."),
        "**Real gap.** Mapping rule exists but XML does not populate the target column.",
        "Add column and lookup logic to ODI final load.",
    )
    add(
        q(["SHRT_SALE_EXMPT_NM"]),
        "Missing implementation",
        mapping_logic(["SHRT_SALE_EXMPT_NM"], "Mapping expects CL_VAL.CL_VAL_NM, with CL_SCM_ID = 115."),
        odi_logic(["SHRT_SALE_EXMPT_NM"], "Column is not present in the final XML insert/update target list."),
        "**Real gap.** Name value is not loaded.",
        "Add column and CL_VAL lookup with the correct schema ID.",
    )

    # 4 SDIRA
    sdira_cols = ["SDIRA_TXN_TP_CD", "SDIRA_TXN_TP", "SDIRA_TXN_YR"]
    add(
        q(sdira_cols),
        "Possible wrong lookup / missing parse",
        mapping_logic(sdira_cols, "Mapping says logic applies only for SRC_STM_ID = 60 and requires parsing TXN.TRD_NUM, for example splitting transaction type and year."),
        odi_logic(sdira_cols, "XML lineage shows SDIRA_TXN_YR <- TXN.TRD_NUM, while type code/name come from CL_VAL; visible CL_VAL join appears not tied to parsed TRD_NUM."),
        "**High-risk mismatch.** XML may not implement the intended SDIRA parsing rule.",
        "Validate actual XML alias for CL_VAL. If no dedicated SDIRA lookup exists, add parsing logic for TRD_NUM and correct CL_VAL join.",
    )

    # 5-6 debit card
    add(
        q(["DB_CARD_TXN_DT"]),
        "Missing source-system condition",
        mapping_logic(["DB_CARD_TXN_DT"], "Mapping says debit card fields should apply only when SRC_STM_ID = 60."),
        odi_logic(["DB_CARD_TXN_DT"], "XML lineage shows DB_CARD_TXN_DT from TXN.ORIG_TD / staged value, with no clear CASE WHEN SRC_STM_ID = 60."),
        "**Potential overpopulation.** Field may be filled for non-debit-card source systems.",
        "Add or verify CASE WHEN TXN.SRC_STM_ID = 60 THEN TXN.ORIG_TD END.",
    )
    debit_ccy_cols = ["DB_CARD_ORIG_CCY_CD", "DB_CARD_ORIG_CCY_NM"]
    add(
        q(debit_ccy_cols),
        "Missing source-system condition",
        mapping_logic(debit_ccy_cols, "Mapping limits debit card currency logic to SRC_STM_ID = 60."),
        odi_logic(debit_ccy_cols, "XML does not clearly show the same source-system guard."),
        "**Potential logic drift.** Currency may be populated outside the intended debit card scope.",
        "Confirm final XML expression. Add explicit SRC_STM_ID = 60 guard if missing.",
    )

    # 7 MM_ALT_ID
    add(
        q(["MM_ALT_ID"]),
        "WHERE-vs-CASE structural difference",
        mapping_logic(["MM_ALT_ID"], "Mapping describes the rule like a filter: SRC_STM_ID = 60, ORIG_SRC_STM_CODE LIKE 'MM%', SRC_CRT_USRNM = 'BPMWRAPB'."),
        odi_logic(["MM_ALT_ID"], "XML uses CASE WHEN ... THEN TXN.ORIG_SRC_STM_CODE END."),
        "**Likely acceptable structural difference.** XML preserves all fact rows and only nulls the column when the condition fails.",
        "No change unless the business requirement truly intended to filter out rows, which would be unusual for a fact load.",
    )

    # 8 ACAT group.
    acat_cols = [c for c in ["ACAT_CNTRA_FIRM_CLRG_ID_TP_CD", "ACAT_CNTRA_FIRM_CLRG_NUM", "ACAT_CNTRA_FIRM_NM", "ACAT_CNTRA_FIRM_SHRT_NM"] if c in by_col]
    add(
        "`ACAT_CNTRA_FIRM_*`",
        "JOIN filter moved into CASE",
        mapping_logic(acat_cols, "Mapping puts SRC_STM_ID IN (53,54) inside the join condition to ACATS_BROKER."),
        odi_logic(acat_cols, "XML joins ACATS_BROKER more broadly and applies CASE WHEN SRC_STM_ID IN (53,54) in the selected columns."),
        "**Mostly structural difference.** Output should be equivalent if BROKER_ID is unique and there are no duplicate join matches.",
        "Validate uniqueness of ACATS_BROKER.BROKER_ID. If not guaranteed, move source-system filter back into join.",
    )

    # 9 BKR_AR_ID.
    add(
        q(["BKR_AR_ID"]),
        "Priority logic expanded",
        mapping_logic(["BKR_AR_ID"], "Mapping says: if broker account source is BKRBO1, use TXN.AR_ID; else if AP.BKR_AR_ID exists, use it; else use linked broker account."),
        odi_logic(["BKR_AR_ID"], "XML uses APA_CASH.BKR_AR_ID first, then APA_SECURITY.BKR_AR_ID, then fallback."),
        "**Business-significant difference.** XML introduces explicit cash-before-security priority.",
        "Confirm with business whether cash should always override security. If yes, update mapping documentation. If no, adjust XML.",
    )

    # 10 APA enrichment operational detail.
    add(
        "APA enrichment logic",
        "XML has more detailed implementation than mapping",
        "Mapping refers generally to APA / AP enrichment.",
        "XML splits APA into APA_CASH and APA_SECURITY, applies regex filters, exception staging, and COALESCE logic.",
        "**XML is more operationally specific.** Mapping is less detailed than implementation.",
        "Update mapping document to describe APA split, precedence, regex validation, and exception handling.",
    )

    # 11 LGCY
    add(
        q(["LGCY_TRD_CPCTY_TP_DIM_ID"]),
        "XML-only exception",
        mapping_logic(["LGCY_TRD_CPCTY_TP_DIM_ID"], "Mapping describes dimension lookup logic."),
        odi_logic(["LGCY_TRD_CPCTY_TP_DIM_ID"], "XML adds special handling: when SRC_STM_ID = 3, dimension ID is forced to 0."),
        "**Mapping omission.** XML contains a hardcoded source-system exception not documented in mapping.",
        "Add this exception to mapping or remove it from XML if no longer valid.",
    )

    # 12 BATCH_DT
    add(
        q(["BATCH_DT"]),
        "Different journal source",
        mapping_logic(["BATCH_DT"], "Mapping references J$TXN."),
        odi_logic(["BATCH_DT"], "XML uses J$AVY_FACT.BATCH_DT / AVY_FACT journal staging."),
        "**Structural lineage difference.** This may be correct if the scenario uses AVY_FACT journal staging, but it differs from mapping.",
        "Confirm intended journal source for batch date. Update mapping or XML accordingly.",
    )

    # 13 Final target
    add(
        "Final target",
        "Target ambiguity",
        "User requested TRANSACTIONS_OWNER.AVY_FACT / DRD sheet target is AVY_FACT.",
        "XML scenario name/header/resources indicate AVY_FACT_SIDE or runtime SSDS_AVY_FACT_TABLE_NAME side-table behavior.",
        "**Environment/target risk.** The XML may not represent the production AVY_FACT load directly.",
        "Check runtime variable SSDS.SSDS_AVY_FACT_TABLE_NAME and confirm whether it resolves to AVY_FACT or AVY_FACT_SIDE.",
    )

    # 14 Column count
    mapping_count = len([r for r in column_diff_rows if r.get("status") in {"IN_BOTH", "MAPPING_ONLY"}])
    final_missing = [c for c in ["STEP_IN_OUT_IND_CD", "STEP_IN_OUT_IND_NM", "SHRT_SALE_EXMPT_CD", "SHRT_SALE_EXMPT_NM"] if by_col.get(c, {}).get("status") == "MAPPING_ONLY"]
    add(
        "Final insert/update column count",
        "Structural mismatch",
        f"Mapping contains {mapping_count} target columns.",
        "XML final merge loads fewer target columns; mapped missing columns include " + ", ".join(final_missing) + ".",
        "**Confirmed structural gap.** Four mapped columns are missing from XML final load.",
        "Add the four missing columns or document why they are intentionally excluded.",
    )

    return rows



def normalize_business_logic_text(text: str) -> str:
    """
    Normalize DRD and ODI logic text for equivalence checks.
    This intentionally maps common DRD business field names to ODI physical fields
    for TaxLot-style mappings.
    """
    s = normalize_space(text).upper()

    replacements = {
        "TAX_LOT_OPN_MSTR.": "",
        "TAXLOT_DTL_OPN.": "",
        "OPN_TAX_LOTS_NONBKR_TGT.": "",
        "CLOSE_TAX_LOT_NONBKR_RJTRUST_TGT.": "",
        "CLS_TAX_LOTS_NONBKR_TGT.": "",
        "ACG_TP_DIM.": "",
        "SRC_STM_DIM.": "",
        "CL_VAL1.": "",
        "CL_VAL2.": "",
        "CL_VAL3.": "",
        "CL_VAL4.": "",
        "CL_VAL.": "",

        # Source/field synonyms seen in TaxLot DRD vs ODI generated SQL.
        "STM_BASE_ISO_CCY_CODE": "CCY_CODE",
        "AC_TP_CODE": "ACG_TP_CODE",
        "CL_VAL_CODE": "CL_VAL_CD",
        "WASH_SALE_F": "WASH_SALE_IND",
        "SBC_WASH_SALE_AMT": "DISALWD_WASHSALE_AMT",
        "SBC_ASOF_ADJ_COST_AMT": "ADJ_COST",
        "NML_ASOF_CRN_COST_AMT": "ORIG_COST",
        "SBC_OPN_COST_AMT": "ORIG_COST",
        "SBC_ASOF_FAIR_MKT_VAL_AMT": "FMV",
        "SBC_OPN_PRC_FCTR": "OPN_FCTR",
        "SBC_ORIG_PROCD_AMT": "ORIG_PROCD",
        "SBC_ORIG_PRC": "ORIG_PRC",
        "SBC_OPN_PER_UC_AMT": "UC",
        "NULL_BSS_F": "NULL_BSS_F",
        "ZERO_BSS_IND_F": "ZERO_BSS_IND",
    }

    for src, dst in replacements.items():
        s = s.replace(src, dst)

    # Remove report labels and context headings.
    s = re.sub(r"ODI ATTRIBUTE EXPRESSION:|DRIVING SOURCE:|RELEVANT JOIN / LOOKUP CONTEXT:", " ", s)
    s = re.sub(r"\bAS\b", " ", s)
    s = re.sub(r"[^A-Z0-9_'><=]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def sql_string_constants(text: str) -> set:
    return set(re.findall(r"'([^']+)'", text.upper()))


def likely_equivalent_logic_row(row: Dict[str, str]) -> Tuple[bool, str]:
    """
    Decide whether a reported candidate is actually an equivalent implementation.

    Returns:
      (is_equivalent, reason)
    """
    area = row.get("Area / Columns", "")
    diff_type = row.get("Difference Type", "")
    mapping = row.get("Mapping Logic", "")
    odi = row.get("ODI XML Logic", "")
    combined = (mapping + " " + odi).upper()

    # Never auto-hide true structural gaps.
    if any(k in diff_type.upper() for k in ["MISSING", "STRUCTURAL", "TARGET AMBIGUITY"]):
        return False, ""

    # Do not auto-hide transformation source drift.
    # Example: same CASE constants, but DRD derives from TAX_LOT_OPN_MSTR while ODI derives from OPN_TAX_LOTS_NONBKR_TGT.
    # That must remain in mismatches/review-required, even if branch outcomes are equivalent.
    mapping_upper = mapping.upper()
    odi_upper = odi.upper()
    source_drift_pairs = [
        ("TAX_LOT_OPN_MSTR", "OPN_TAX_LOTS_NONBKR_TGT"),
        ("TAX_LOT_OPN_MSTR", "CLOSE_TAX_LOT_NONBKR_RJTRUST_TGT"),
    ]
    for drd_src, odi_src in source_drift_pairs:
        if drd_src in mapping_upper and odi_src in odi_upper and drd_src not in odi_upper:
            return False, "Transformation source drift: DRD and ODI use different physical source tables."


    m_norm = normalize_business_logic_text(mapping)
    o_norm = normalize_business_logic_text(odi)
    o_upper = odi.upper()

    # Equivalent lookup/dimension patterns.
    lookup_requirements = {
        "ACG_TP_NM": ["ACG_TP_DIM", "ACG_TP_CD", "ACG_TP_NM"],
        "CIRD_PD_ID": ["CCAL_CIRD_PD_MAP", "CIRD_PD_ID"],
        "OPN_TXN_EV_TP": ["CL_VAL", "TAX_LOT_TXN_EV_TP_ID", "CL_VAL_NM", "86"],
        "SRC_RCRD_TP_CD": ["CL_VAL", "SRC_RCRD_TP_ID", "CL_VAL"],
        "SRC_STM_CD": ["SRC_STM_DIM", "SRC_STM_ID", "SRC_STM_CD"],
        "SRC_STM_NM": ["SRC_STM_DIM", "SRC_STM_ID", "SRC_STM_NM"],
        "TAX_LOT_TP": ["CL_VAL", "TAX_LOT_TXN_TP_ID", "CL_VAL_NM", "84"],
    }

    if area in lookup_requirements:
        reqs = lookup_requirements[area]
        if all(req in o_upper for req in reqs):
            return True, "Equivalent lookup/dimension implementation."

    # Textual ZERO cost basis flag pattern.
    if (
        "ZERO_COST_BSS_F" in area.upper()
        and "ZERO_BSS_IND" in m_norm
        and "ZERO_BSS_IND" in o_norm
        and "01" in m_norm
        and "01" in o_norm
        and "Y" in m_norm
        and "Y" in o_norm
        and "N" in m_norm
        and "N" in o_norm
    ):
        return True, "Equivalent ZERO cost basis flag rule."

    # Currency conversion pattern:
    # DRD and ODI both use USD-or-null/zero exchange-rate condition and multiply by exchange rate otherwise.
    if (
        "CCY_CODE" in m_norm and "EXG_RATE" in m_norm and
        "CCY_CODE" in o_norm and "EXG_RATE" in o_norm and
        "*" in mapping and "*" in odi and
        "USD" in m_norm and "USD" in o_norm
    ):
        return True, "Equivalent currency/exchange-rate CASE structure."

    # Boolean / enumerated CASE pattern: constants match and CASE exists in both.
    m_consts = sql_string_constants(mapping)
    o_consts = sql_string_constants(odi)
    if m_consts and "CASE" in m_norm and "CASE" in o_norm:
        # ODI may add ELSE NULL; that is equivalent to SQL CASE with no ELSE.
        if m_consts.issubset(o_consts):
            return True, "Equivalent CASE constants and branches after source-name normalization."

    # Direct source-field pass-through or simple condition equivalence after normalization.
    important = [
        tok for tok in re.findall(r"\b[A-Z][A-Z0-9_]{2,}\b", m_norm)
        if tok not in {
            "CASE", "WHEN", "THEN", "ELSE", "END", "NULL", "AND", "OR", "IS",
            "USE", "GET", "PICK", "LOOKUP", "TABLE", "WHERE", "CODE", "NAME"
        }
    ]
    if important:
        missing = [tok for tok in important if tok not in o_norm]
        if not missing:
            return True, "Equivalent expression after DRD/ODI source-name normalization."

    return False, ""


def split_mismatch_and_equivalent_rows(rows: List[Dict[str, str]]) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    mismatches: List[Dict[str, str]] = []
    equivalent: List[Dict[str, str]] = []

    for row in rows:
        is_eq, reason = likely_equivalent_logic_row(row)
        if is_eq:
            eq = dict(row)
            eq["Conclusion"] = "MATCH_EQUIVALENT. " + reason
            eq["Recommended Action"] = "No mismatch action required. Keep as matched/equivalent unless business SMEs disagree with source-field synonym mapping."
            equivalent.append(eq)
        else:
            m = dict(row)
            mu = m.get("Mapping Logic", "").upper()
            ou = m.get("ODI XML Logic", "").upper()
            if "TAX_LOT_OPN_MSTR" in mu and "OPN_TAX_LOTS_NONBKR_TGT" in ou and "TAX_LOT_OPN_MSTR" not in ou:
                m["Difference Type"] = "Transformation source drift"
                m["Conclusion"] = "**Source drift.** DRD and ODI use equivalent-looking transformation branches, but derive the attribute from different physical source tables."
                m["Recommended Action"] = "Confirm whether TAX_LOT_OPN_MSTR and OPN_TAX_LOTS_NONBKR_TGT are semantically equivalent for this attribute. If not, adjust ODI source or update DRD."
            mismatches.append(m)

    # Re-number both outputs.
    for i, row in enumerate(mismatches, start=1):
        row["#"] = str(i)
    for i, row in enumerate(equivalent, start=1):
        row["#"] = str(i)

    return mismatches, equivalent


def write_rules_table_markdown(path: Path, rows: List[Dict[str, str]], title: str, description: str) -> None:
    fields = ["#", "Area / Columns", "Difference Type", "Mapping Logic", "ODI XML Logic", "Conclusion", "Recommended Action"]
    with path.open("w", encoding="utf-8") as f:
        f.write(f"# {title}\n\n")
        if description:
            f.write(description + "\n\n")
        if not rows:
            f.write("No rows.\n")
            return
        f.write("| " + " | ".join(fields) + " |\n")
        f.write("|" + "|".join(["---"] * len(fields)) + "|\n")
        for r in rows:
            f.write("| " + " | ".join(md_escape(short(r.get(field, ""), 900)) for field in fields) + " |\n")

def build_full_drd_vs_odi_xml_rules_diff(column_diff_rows: List[Dict[str, str]], logic_rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Build a business-readable differences-only table:
    #, Area / Columns, Difference Type, Mapping Logic, ODI XML Logic,
    Conclusion, Recommended Action.

    Only rows with actual / potential differences are included.
    """
    logic_by_col: Dict[str, Dict[str, str]] = {}
    for r in logic_rows:
        col = r.get("target_column", "")
        if not col:
            continue
        # Prefer higher severity row.
        prev = logic_by_col.get(col)
        sev_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
        if prev is None or sev_order.get(r.get("severity", "LOW"), 9) < sev_order.get(prev.get("severity", "LOW"), 9):
            logic_by_col[col] = r

    out: List[Dict[str, str]] = []
    for row in column_diff_rows:
        col = row.get("target_column", "")
        status = row.get("status", "")

        # Ignore obvious SQL literals / parser artifacts, not real business columns.
        if not col or re.fullmatch(r"[0-9]+", col):
            continue
        if col in {"DUAL", "ODI_RUNTIME_VALUE", "NULL"}:
            continue

        logic = logic_by_col.get(col)

        difference_type, conclusion, action = classify_difference(row)

        # Keep only differences, not "No obvious difference".
        include = False
        if status in {"MAPPING_ONLY", "XML_ONLY"}:
            include = True
        elif logic and logic.get("severity") in {"HIGH", "MEDIUM"}:
            include = True
            # Use more specific logic candidate conclusion where useful.
            if logic.get("conclusion"):
                conclusion = logic["conclusion"]
            if logic.get("missing_drd_conditions_in_xml"):
                difference_type = "Condition mismatch"
            elif logic.get("missing_drd_tables_in_xml"):
                difference_type = "Join / lookup source mismatch"
            elif logic.get("drd_has_case") == "Y" and logic.get("xml_has_case") == "N":
                difference_type = "Missing CASE logic"
            else:
                # Preserve classify_difference if already meaningful.
                if difference_type == "Expression / implementation difference":
                    difference_type = "Transformation logic difference"
        elif difference_type not in {"No obvious difference"} and status == "IN_BOTH":
            # Include expression differences only when DRD has non-trivial rule/logic.
            drd_rule = row.get("drd_rule", "")
            if LOGIC_KEYWORDS.search(drd_rule or ""):
                include = True

        if not include:
            continue

        mapping_logic = row.get("drd_rule", "") or row.get("mapping_source_tables", "")
        odi_logic = row.get("xml_logic_full", "") or row.get("xml_expressions", "") or row.get("xml_source_tables", "")

        # Enrich with condition details if available.
        if logic:
            missing_conditions = logic.get("missing_drd_conditions_in_xml", "")
            missing_tables = logic.get("missing_drd_tables_in_xml", "")
            if missing_conditions:
                conclusion += f" Missing DRD conditions not visible in XML: {missing_conditions}."
            if missing_tables:
                conclusion += f" DRD source/lookup tables not visible in XML: {missing_tables}."
            if logic.get("xml_expression_excerpt") and not odi_logic:
                odi_logic = logic.get("xml_expression_excerpt", "")

        out.append({
            "#": str(len(out) + 1),
            "Area / Columns": col,
            "Difference Type": difference_type,
            "Mapping Logic": mapping_logic,
            "ODI XML Logic": odi_logic,
            "Conclusion": conclusion,
            "Recommended Action": action,
        })

    return out


def write_full_rules_diff_markdown(path: Path, rows: List[Dict[str, str]]) -> None:
    fields = ["#", "Area / Columns", "Difference Type", "Mapping Logic", "ODI XML Logic", "Conclusion", "Recommended Action"]
    with path.open("w", encoding="utf-8") as f:
        f.write("# Full DRD vs ODI XML Rules -- Differences Only\n\n")
        f.write("This file contains only columns/rules where the static comparison found a structural, lookup, JOIN, CASE, condition, or expression difference.\n\n")
        if not rows:
            f.write("No differences found by static analysis.\n")
            return
        f.write("| " + " | ".join(fields) + " |\n")
        f.write("|" + "|".join(["---"] * len(fields)) + "|\n")
        for r in rows:
            f.write("| " + " | ".join(md_escape(short(r.get(field, ""), 900)) for field in fields) + " |\n")

def write_markdown_report(path: Path, detection: MappingDetection, scenario_rows: List[Dict[str, str]],
                          mapping_rows: List[Dict[str, str]], step_rows: List[Dict[str, str]],
                          column_diff_rows: List[Dict[str, str]], logic_rows: List[Dict[str, str]]) -> None:
    total_mapping = len({r["target_column"] for r in mapping_rows})
    in_both = sum(1 for r in column_diff_rows if r["status"] == "IN_BOTH")
    mapping_only = [r for r in column_diff_rows if r["status"] == "MAPPING_ONLY"]
    xml_only = [r for r in column_diff_rows if r["status"] == "XML_ONLY"]
    high_logic = [r for r in logic_rows if r["severity"] == "HIGH"]
    medium_logic = [r for r in logic_rows if r["severity"] == "MEDIUM"]

    with path.open("w", encoding="utf-8") as f:
        f.write("# DRD Excel vs ODI XML Comparison Report\n\n")

        f.write("## Auto-detected DRD layout\n\n")
        f.write("| Field | Value |\n|---|---|\n")
        for k, v in detection.as_human().items():
            f.write(f"| {md_escape(k)} | {md_escape(v)} |\n")
        f.write("\n")

        f.write("## Summary\n\n")
        f.write(f"- Mapping target columns extracted: **{total_mapping}**\n")
        f.write(f"- Columns found in both DRD and ODI lineage: **{in_both}**\n")
        f.write(f"- DRD-only columns: **{len(mapping_only)}**\n")
        f.write(f"- ODI-only columns/aliases: **{len(xml_only)}**\n")
        f.write(f"- High-risk join/case candidates: **{len(high_logic)}**\n")
        f.write(f"- Medium-risk join/case candidates: **{len(medium_logic)}**\n\n")

        if scenario_rows:
            f.write("## ODI scenario metadata\n\n")
            f.write("| Scenario | Version | Scenario No | First Date | Last Date |\n")
            f.write("|---|---:|---:|---|---|\n")
            for s in scenario_rows[:10]:
                f.write(
                    f"| {md_escape(s.get('scenario_name',''))} | {md_escape(s.get('scenario_version',''))} | "
                    f"{md_escape(s.get('scenario_no',''))} | {md_escape(s.get('first_date',''))} | {md_escape(s.get('last_date',''))} |\n"
                )
            f.write("\n")

        f.write("## ODI step summary\n\n")
        f.write("| Step | Name | Type | Tasks | SQL Tasks | Resource | OK Next |\n")
        f.write("|---:|---|---|---:|---:|---|---:|\n")
        for s in step_rows:
            f.write(
                f"| {md_escape(s.get('step_no',''))} | {md_escape(s.get('step_name',''))} | "
                f"{md_escape(s.get('step_type',''))} | {md_escape(s.get('task_count',''))} | "
                f"{md_escape(s.get('sql_task_count',''))} | {md_escape(s.get('resource_name',''))} | "
                f"{md_escape(s.get('ok_next_step',''))} |\n"
            )
        f.write("\n")

        f.write("## DRD-only columns\n\n")
        if mapping_only:
            f.write("| Target Column | DRD Source Tables | DRD Rule |\n")
            f.write("|---|---|---|\n")
            for r in mapping_only[:200]:
                f.write(
                    f"| `{md_escape(r.get('target_column',''))}` | {md_escape(short(r.get('mapping_source_tables',''), 200))} | "
                    f"{md_escape(short(r.get('drd_rule',''), 500))} |\n"
                )
        else:
            f.write("No DRD-only columns found.\n")
        f.write("\n\n")

        f.write("## ODI-only columns / aliases\n\n")
        if xml_only:
            f.write("| XML Column/Alias | XML Steps | XML Expression |\n")
            f.write("|---|---|---|\n")
            for r in xml_only[:200]:
                f.write(
                    f"| `{md_escape(r.get('target_column',''))}` | {md_escape(r.get('xml_steps',''))} | "
                    f"{md_escape(short(r.get('xml_expressions',''), 500))} |\n"
                )
        else:
            f.write("No ODI-only columns/aliases found.\n")
        f.write("\n\n")

        f.write("## High / medium join-case candidates\n\n")
        candidates = [r for r in logic_rows if r["severity"] in {"HIGH", "MEDIUM"}]
        if candidates:
            f.write("| Severity | Target Column | Conclusion | Missing DRD Conditions in XML | DRD Rule | ODI Expression Excerpt |\n")
            f.write("|---|---|---|---|---|---|\n")
            for r in candidates[:200]:
                f.write(
                    f"| **{md_escape(r.get('severity',''))}** | `{md_escape(r.get('target_column',''))}` | "
                    f"{md_escape(r.get('conclusion',''))} | {md_escape(r.get('missing_drd_conditions_in_xml',''))} | "
                    f"{md_escape(short(r.get('drd_rule',''), 500))} | {md_escape(short(r.get('xml_expression_excerpt',''), 500))} |\n"
                )
        else:
            f.write("No high/medium join-case candidates found by static analysis.\n")
        f.write("\n\n")

        f.write("## Review notes\n\n")
        f.write("- Auto-detection can be overridden with `--mapping-sheet`, `--target-col`, `--source-cols`, `--rule-col`, and `--header-row`.\n")
        f.write("- Static parsing flags potential mismatches. Validate high/medium rows with executed ODI SQL or database output.\n")
        f.write("- ODI variables and macros are normalized where possible, but runtime substitutions may still affect final SQL.\n")


# ======================================================================================
# Main
# ======================================================================================

def run(args: argparse.Namespace) -> Path:
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    xml_path = Path(args.xml).expanduser().resolve()
    out_dir = Path(args.out).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    if not xml_path.exists():
        raise FileNotFoundError(f"XML file not found: {xml_path}")

    objects = parse_odi_objects(xml_path)
    xml_targets = extract_target_resources_from_xml(objects)
    scenario_rows, step_rows, sql_blocks = extract_odi_summary(objects)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    detection = auto_detect_mapping(
        wb,
        xml_targets=xml_targets,
        target_table_override=args.target_table or "",
        mapping_sheet_override=args.mapping_sheet or "",
        header_row_override=args.header_row,
        target_col_override=args.target_col or "",
        source_cols_override=args.source_cols or "",
        rule_col_override=args.rule_col or "",
    )

    mapping_rows, notes_rows = extract_mapping_from_xlsx(xlsx_path, detection, notes_sheet=args.notes_sheet)
    lineage_rows = build_odi_lineage(sql_blocks)
    final_lineage_rows = select_final_target_lineage(lineage_rows)

    # Business column diff should compare DRD to final target load only.
    # Full ODI lineage is still written separately for forensic review.
    column_diff_rows = compare_columns(mapping_rows, final_lineage_rows)

    # Join/case logic still uses full lineage, because conditions may be implemented in prior stages.
    logic_rows = build_logic_diff_candidates(mapping_rows, lineage_rows, sql_blocks)

    # Outputs
    detection_json = detection.as_human()
    detection_json["requested_profile"] = args.profile
    # Actual profile is resolved later for report generation; this field records the user request.
    (out_dir / "detected_layout.json").write_text(json.dumps(detection_json, indent=2), encoding="utf-8")

    write_csv(
        out_dir / "mapping_extract.csv",
        mapping_rows,
        ["excel_row", "target_column_raw", "target_column", "source_1", "source_2", "source_3",
         "source_tables_combined", "drd_rule", "drd_notes"],
    )

    note_fields = sorted({k for r in notes_rows for k in r.keys()}) if notes_rows else ["excel_row", "row_text"]
    write_csv(out_dir / "etl_notes_extract.csv", notes_rows, note_fields)

    write_csv(
        out_dir / "odi_scenario_summary.csv",
        scenario_rows,
        ["scenario_name", "scenario_version", "scenario_no", "first_date", "last_date", "first_user", "last_user"],
    )

    write_csv(
        out_dir / "odi_step_summary.csv",
        step_rows,
        ["step_no", "step_name", "step_type", "ok_next_step", "ko_next_step", "task_count", "sql_task_count",
         "variable_name", "variable_value", "resource_name", "table_name"],
    )

    write_sql_blocks(out_dir / "odi_sql_blocks.sql", sql_blocks)

    write_csv(
        out_dir / "odi_column_lineage_best_effort.csv",
        lineage_rows,
        ["step_no", "task_no", "task_name_1", "task_name_2", "task_name_3", "field",
         "target_column", "expression", "alias", "source_tables", "lineage_type", "has_case", "has_join", "sql_excerpt", "xml_logic_full"],
    )

    column_diff_fields = [
        "target_column", "status", "mapping_excel_row", "mapping_source_tables", "drd_rule", "drd_notes",
        "xml_occurrences", "xml_steps", "xml_lineage_types", "xml_expressions", "xml_source_tables", "xml_logic_full"
    ]

    # Main business diff: DRD vs final target load only.
    write_csv(
        out_dir / "column_diff.csv",
        column_diff_rows,
        column_diff_fields,
    )

    # Backward-compatible business diff filename.
    write_csv(
        out_dir / "mapping_vs_xml_column_diff.csv",
        column_diff_rows,
        column_diff_fields,
    )

    # Forensic diff: DRD vs every ODI SQL alias/stage/control column.
    # This intentionally contains staging aliases and technical columns.
    all_lineage_column_diff_rows = compare_columns(mapping_rows, lineage_rows)
    write_csv(
        out_dir / "all_odi_lineage_column_diff.csv",
        all_lineage_column_diff_rows,
        column_diff_fields,
    )

    write_csv(
        out_dir / "join_case_diff_candidates.csv",
        logic_rows,
        ["target_column", "severity", "conclusion", "mapping_excel_row", "drd_source_tables", "drd_rule", "drd_notes",
         "xml_steps", "xml_expression_excerpt", "drd_conditions", "xml_conditions", "missing_drd_conditions_in_xml",
         "drd_tables", "xml_tables", "missing_drd_tables_in_xml", "drd_has_case", "xml_has_case",
         "drd_has_join_or_where_or_lookup", "xml_has_join"],
    )

    # Profile behavior:
    # - generic: no target-specific curated findings and no domain equivalence suppression.
    # - avy: force AVY curated review table when applicable.
    # - taxlot: use generic extraction + TaxLot equivalence/source-drift heuristics.
    # - auto: choose based on target metadata, not filenames.
    active_profile = args.profile
    if active_profile == "auto":
        target_blob = (detection.target_table_from_sheet + " " + " ".join(detection.target_resources_from_xml)).upper()
        if "AVY_FACT" in target_blob:
            active_profile = "avy"
        elif "TAX_LOT" in target_blob or "TAXLOTS" in target_blob:
            active_profile = "taxlot"
        else:
            active_profile = "generic"

    raw_full_rules_rows = []
    used_curated_avy = False
    if active_profile == "avy":
        raw_full_rules_rows = build_avy_review_rules_diff(column_diff_rows, logic_rows, detection)
        used_curated_avy = bool(raw_full_rules_rows)

    if not raw_full_rules_rows:
        raw_full_rules_rows = build_full_drd_vs_odi_xml_rules_diff(column_diff_rows, logic_rows)

    if active_profile == "generic" or used_curated_avy:
        # Generic mode keeps all extracted differences/review candidates and does not apply
        # TaxLot-specific equivalence suppression. AVY curated rows are already curated.
        full_rules_diff_rows = raw_full_rules_rows
        equivalent_rows = []
    else:
        full_rules_diff_rows, equivalent_rows = split_mismatch_and_equivalent_rows(raw_full_rules_rows)

    rules_fields = ["#", "Area / Columns", "Difference Type", "Mapping Logic", "ODI XML Logic", "Conclusion", "Recommended Action"]

    write_csv(
        out_dir / "full_drd_vs_odi_xml_rules_diff.csv",
        full_rules_diff_rows,
        rules_fields,
    )

    write_rules_table_markdown(
        out_dir / "full_drd_vs_odi_xml_rules_diff.md",
        full_rules_diff_rows,
        "Full DRD vs ODI XML Rules -- Mismatches Only",
        "This file contains only real mismatches / review-required differences after filtering equivalent implementations."
    )

    # Backward-compatible file names requested by earlier review flow.
    write_csv(
        out_dir / "full_drd_vs_odi_xml_rules.csv",
        full_rules_diff_rows,
        rules_fields,
    )

    write_rules_table_markdown(
        out_dir / "full_drd_vs_odi_xml_rules.md",
        full_rules_diff_rows,
        "Full DRD vs ODI XML Rules -- Mismatches Only",
        "This file contains only real mismatches / review-required differences after filtering equivalent implementations."
    )

    write_csv(
        out_dir / "matched_logic_equivalent.csv",
        equivalent_rows,
        rules_fields,
    )

    write_rules_table_markdown(
        out_dir / "matched_logic_equivalent.md",
        equivalent_rows,
        "Matched Logic Equivalent Rows",
        "Rows removed from the mismatch report because DRD and ODI logic appear equivalent after source-name / alias normalization."
    )

    write_markdown_report(out_dir / "comparison_report.md", detection, scenario_rows, mapping_rows, step_rows, column_diff_rows, logic_rows)

    if not args.quiet:
        print(f"Done. Output written to: {out_dir}")
        print("Detected layout:", json.dumps(detection_json, ensure_ascii=False))
        print(f"Mapping columns: {len({r['target_column'] for r in mapping_rows})}")
        print(f"ODI lineage rows: {len(lineage_rows)}")
        print(f"Column diff rows: {len(column_diff_rows)}")
        print(f"Join/case candidates: {len(logic_rows)}")
    return out_dir


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Universal DRD Excel vs ODI scenario XML comparison tool.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--xlsx", required=True, help="Path to DRD mapping Excel file (.xlsx)")
    p.add_argument("--xml", required=True, help="Path to ODI scenario XML export")
    p.add_argument("--out", default="odi_drd_compare_output", help="Output folder")

    p.add_argument("--mapping-sheet", default="", help="Optional mapping sheet override. Auto-detected when omitted.")
    p.add_argument("--notes-sheet", default="ETL Notes", help="ETL notes sheet name")
    p.add_argument("--target-table", default="", help="Optional target table override used to pick the best sheet.")
    p.add_argument("--target-col", default="", help="Optional Excel target column override, e.g. B or C")
    p.add_argument("--source-cols", default="", help="Optional comma-separated source columns, e.g. Y,Z,AA or V,W,X")
    p.add_argument("--rule-col", default="", help="Optional rule column override, e.g. AD or AA")
    p.add_argument("--header-row", type=int, default=None, help="Optional header row override")
    p.add_argument(
        "--profile",
        default="auto",
        choices=["auto", "generic", "avy", "taxlot"],
        help=(
            "Comparison profile. auto uses target metadata to enable known domain heuristics. "
            "generic disables target-specific curated/equivalence heuristics."
        ),
    )
    p.add_argument("--quiet", action="store_true", help="Suppress console summary")
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    try:
        run(args)
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2




# ---------------------------------------------------------------------------
# R2 vendor-in wrappers (added 2026-06-05). Generic-only by default: the
# `profile` arg below only affects the curated full-rules-diff output, never
# the column counts. Pipeline functions above are untouched.
# ---------------------------------------------------------------------------
from collections import Counter as _Counter


def compare_summary(xlsx_path, xml_path, *, profile: str = "generic") -> dict:
    """Run the v15 pipeline up to the DRD-vs-final-ODI column diff and return
    the count summary WITHOUT writing any files. Honest, generic, no heuristics."""
    xlsx_path = Path(xlsx_path).expanduser().resolve()
    xml_path = Path(xml_path).expanduser().resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    if not xml_path.exists():
        raise FileNotFoundError(f"XML file not found: {xml_path}")

    objects = parse_odi_objects(xml_path)
    xml_targets = extract_target_resources_from_xml(objects)
    _scn, _steps, sql_blocks = extract_odi_summary(objects)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    detection = auto_detect_mapping(wb, xml_targets=xml_targets)

    mapping_rows, _notes = extract_mapping_from_xlsx(xlsx_path, detection)
    lineage_rows = build_odi_lineage(sql_blocks)
    final_lineage_rows = select_final_target_lineage(lineage_rows)
    column_diff_rows = compare_columns(mapping_rows, final_lineage_rows)

    counts = _Counter(r["status"] for r in column_diff_rows)
    return {
        "profile": profile,
        "mapping_columns": len({r["target_column"] for r in mapping_rows if r.get("target_column")}),
        "in_both": counts.get("IN_BOTH", 0),
        "mapping_only": counts.get("MAPPING_ONLY", 0),
        "xml_only": counts.get("XML_ONLY", 0),
        "detection": detection.as_human(),
    }


def compare_to_dir(xlsx_path, xml_path, out_dir, *, profile: str = "generic") -> Path:
    """Full v15 run: writes all report artifacts to out_dir, returns the dir.
    profile defaults to 'generic' (no AVY/TaxLot curated heuristics)."""
    ns = argparse.Namespace(
        xlsx=str(xlsx_path), xml=str(xml_path), out=str(out_dir),
        mapping_sheet="", notes_sheet="ETL Notes", target_table="",
        target_col="", source_cols="", rule_col="", header_row=None,
        profile=profile, quiet=True,
    )
    return run(ns)

if __name__ == "__main__":
    raise SystemExit(main())
