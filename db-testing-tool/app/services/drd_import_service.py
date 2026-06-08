"""DRD (Data Requirements Document) import service.

Parses CSV/Excel files in enterprise DRD format (like Book3.csv, Book44.csv),
allows column filtering, and groups column-level mappings into mapping rules.
"""
import csv
import hashlib
import io
import json
import logging
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandera.pandas as pa
from rapidfuzz import fuzz, process as rfprocess
from app.services.schema_kb_service import load_schema_kb_payload
from app.services.sql_pattern_validation import split_valid_invalid_test_defs

logger = logging.getLogger(__name__)


# ── Pandera schema for post-parse DRD row validation ────────────────────────

import pandas as pd

_drd_row_schema = pa.DataFrameSchema(
    {
        "physical_name": pa.Column(nullable=True),
        "logical_name": pa.Column(nullable=True),
        "source_table": pa.Column(nullable=True),
        "source_attribute": pa.Column(nullable=True),
        "source_schema": pa.Column(nullable=True),
        "transformation": pa.Column(nullable=True),
    },
    # Accept rows that have at least physical_name OR logical_name
    checks=[
        pa.Check(lambda df: bool((df["physical_name"].notna() | df["logical_name"].notna()).all()),
                 error="Every row needs physical_name or logical_name"),
    ],
    strict=False,  # allow extra columns (raw_col_*, notes, etc.)
)


def _validate_parsed_rows(column_mappings: List[Dict[str, Any]]) -> List[str]:
    """Run pandera validation on parsed DRD rows, returning any error messages."""
    if not column_mappings:
        return []
    try:
        df = pd.DataFrame(column_mappings)
        # Fill missing expected columns to avoid KeyError
        for col in ("physical_name", "logical_name", "source_table", "source_attribute", "source_schema", "transformation"):
            if col not in df.columns:
                df[col] = None
        _drd_row_schema.validate(df, lazy=True)
        return []
    except pa.errors.SchemaErrors as exc:
        return [str(e) for e in exc.schema_errors][:20]
    except pa.errors.SchemaError as exc:
        return [str(exc)]
    except Exception as exc:
        return [f"Pandera validation error: {exc}"]


def _fq_name(schema: str, table: str) -> str:
    schema = (schema or "").strip()
    table = (table or "").strip()
    return f"{schema}.{table}" if schema else table


def _column_token_parts(expr: str) -> Tuple[str, str]:
    text = (expr or "").strip().strip('"')
    if not text:
        return "", ""
    match = re.search(r'(?:(?P<alias>[A-Z0-9_\$#]+)\.)?"?(?P<column>[A-Z0-9_\$#]+)"?$', text.upper())
    if not match:
        return "", ""
    return (match.group("alias") or "").upper(), (match.group("column") or "").upper()


def _parse_main_grain_pairs(main_grain: str) -> List[Tuple[str, str]]:
    if not main_grain:
        return []

    pairs: List[Tuple[str, str]] = []
    seen = set()
    raw_parts = re.split(r'\bAND\b|,|\n|;', main_grain, flags=re.IGNORECASE)
    for raw_part in raw_parts:
        part = (raw_part or "").strip()
        if not part:
            continue

        if "=" in part:
            left, right = [segment.strip() for segment in part.split("=", 1)]
            left_alias, left_col = _column_token_parts(left)
            right_alias, right_col = _column_token_parts(right)
            if not left_col or not right_col:
                continue

            target_col = left_col
            source_col = right_col
            if left_alias == "S" and right_alias == "T":
                target_col = right_col
                source_col = left_col
            elif left_alias != "T" and right_alias == "T":
                target_col = right_col
                source_col = left_col

            key = (target_col, source_col)
            if key not in seen:
                seen.add(key)
                pairs.append(key)
            continue

        _, col = _column_token_parts(part)
        if not col:
            continue
        key = (col, col)
        if key not in seen:
            seen.add(key)
            pairs.append(key)

    return pairs


def _build_kb_table_index(datasource_id: int) -> Dict[Tuple[str, str], Dict[str, Any]]:
    payload = load_schema_kb_payload(datasource_id)
    sources = payload.get("sources", []) if isinstance(payload, dict) else []
    table_index: Dict[Tuple[str, str], Dict[str, Any]] = {}

    for src in sources:
        pdm = (src or {}).get("pdm", {})
        for s in pdm.get("schemas", []) or []:
            schema_name = (s.get("schema") or "").strip()
            if not schema_name:
                continue
            for t in s.get("tables", []) or []:
                table_name = (t.get("name") or "").strip()
                if not table_name:
                    continue
                col_map = {}
                for c in t.get("columns", []) or []:
                    col_name = (c.get("name") or "").strip()
                    if col_name:
                        col_map[col_name.upper()] = col_name
                table_index[(schema_name.upper(), table_name.upper())] = {
                    "schema": schema_name,
                    "table": table_name,
                    "columns": col_map,
                }

    return table_index


# ── Fuzzy matching helpers (rapidfuzz) ───────────────────────────────────────

_FUZZY_SCORE_THRESHOLD = 78  # min score (0-100) to accept a fuzzy match


def _rfuzz_best(query: str, choices: List[str], threshold: int = _FUZZY_SCORE_THRESHOLD) -> Optional[Tuple[str, float]]:
    """Return (best_match, score) using rapidfuzz token_sort_ratio, or None."""
    if not choices:
        return None
    result = rfprocess.extractOne(query, choices, scorer=fuzz.token_sort_ratio, score_cutoff=threshold)
    if result is None:
        return None
    match_str, score, _idx = result
    return match_str, score


def _resolve_table_from_index(
    table_index: Dict[Tuple[str, str], Dict[str, Any]],
    schema: str,
    table: str,
) -> Optional[Dict[str, Any]]:
    schema_u = (schema or "").strip().upper()
    table_u = (table or "").strip().upper()
    if not table_u:
        return None

    if schema_u:
        direct = table_index.get((schema_u, table_u))
        if direct:
            return direct
        same_schema_keys = [k for k in table_index.keys() if k[0] == schema_u]
        hit = _rfuzz_best(table_u, [k[1] for k in same_schema_keys])
        if hit:
            return table_index.get((schema_u, hit[0]))

    global_candidates = [k for k in table_index.keys() if k[1] == table_u]
    if global_candidates:
        return table_index[global_candidates[0]]

    all_tables = [k[1] for k in table_index.keys()]
    hit_global = _rfuzz_best(table_u, all_tables)
    if hit_global:
        for key in table_index.keys():
            if key[1] == hit_global[0]:
                return table_index[key]
    return None


def _resolve_column_name(col_map: Dict[str, str], column_name: str) -> Optional[str]:
    col_u = (column_name or "").strip().upper()
    if not col_u:
        return None
    direct = col_map.get(col_u)
    if direct:
        return direct
    hit = _rfuzz_best(col_u, list(col_map.keys()))
    if hit:
        return col_map.get(hit[0])
    return None


def _resolve_column_with_confidence(col_map: Dict[str, str], column_name: str) -> Tuple[Optional[str], float]:
    """Like _resolve_column_name but also returns a 0-100 confidence score."""
    col_u = (column_name or "").strip().upper()
    if not col_u:
        return None, 0.0
    direct = col_map.get(col_u)
    if direct:
        return direct, 100.0
    hit = _rfuzz_best(col_u, list(col_map.keys()))
    if hit:
        return col_map.get(hit[0]), hit[1]
    return None, 0.0


def validate_column_mappings_with_kb(
    column_mappings: List[Dict[str, Any]],
    target_schema: str,
    target_table: str,
    source_datasource_id: int,
    target_datasource_id: int,
) -> Dict[str, Any]:
    """Normalize DRD mapping attributes against local KB/PDM tables and columns."""
    source_index = _build_kb_table_index(source_datasource_id)
    target_index = _build_kb_table_index(target_datasource_id)

    resolved_target = _resolve_table_from_index(target_index, target_schema, target_table)
    target_col_map = (resolved_target or {}).get("columns", {})

    fixed_rows: List[Dict[str, Any]] = []
    mismatches: List[str] = []
    confidence_log: List[Dict[str, Any]] = []
    validated = 0
    unresolved = 0

    for row in column_mappings or []:
        rec = dict(row or {})

        src_schema = (rec.get("source_schema") or "").strip()
        src_table = (rec.get("source_table") or "").strip()
        src_attr = (rec.get("source_attribute") or "").strip()
        tgt_attr = (rec.get("physical_name") or "").strip()
        logical = (rec.get("logical_name") or "").strip()

        src_tbl = _resolve_table_from_index(source_index, src_schema, src_table)
        if src_tbl:
            rec["source_schema"] = src_tbl.get("schema") or rec.get("source_schema")
            rec["source_table"] = src_tbl.get("table") or rec.get("source_table")

        src_col_map = (src_tbl or {}).get("columns", {})
        # If the DRD source_attribute already matches a column exactly, keep it
        # and skip fuzzy resolution to avoid mis-mapping similar column names.
        src_attr_u = src_attr.upper().strip()
        if src_col_map and src_attr_u and src_attr_u in src_col_map:
            src_resolved, src_conf = src_col_map[src_attr_u], 100.0
        else:
            src_resolved, src_conf = _resolve_column_with_confidence(src_col_map, src_attr) if src_tbl else (src_attr, 100.0 if src_attr else 0.0)
        if src_resolved:
            rec["source_attribute"] = src_resolved
            validated += 1
        else:
            unresolved += 1
            if src_table and src_attr:
                mismatches.append(f"Source attribute not found in PDM: {src_schema}.{src_table}.{src_attr}")
        confidence_log.append({"field": f"{src_schema}.{src_table}.{src_attr}", "resolved": src_resolved, "confidence": src_conf, "side": "source"})

        tgt_candidate = tgt_attr or (logical.upper().replace(" ", "_") if logical else "")
        # If the DRD target column already matches exactly, keep it — don't let
        # fuzzy matching reassign it to a different column with a similar name.
        tgt_candidate_u = (tgt_candidate or "").upper().strip()
        if target_col_map and tgt_candidate_u and tgt_candidate_u in target_col_map:
            tgt_resolved, tgt_conf = target_col_map[tgt_candidate_u], 100.0
        else:
            tgt_resolved, tgt_conf = _resolve_column_with_confidence(target_col_map, tgt_candidate) if target_col_map else (tgt_candidate, 100.0 if tgt_candidate else 0.0)
        if tgt_resolved:
            rec["physical_name"] = tgt_resolved
            validated += 1
        else:
            unresolved += 1
            if tgt_candidate:
                mismatches.append(f"Target attribute not found in PDM: {target_schema}.{target_table}.{tgt_candidate}")
        confidence_log.append({"field": f"{target_schema}.{target_table}.{tgt_candidate}", "resolved": tgt_resolved, "confidence": tgt_conf, "side": "target"})

        fixed_rows.append(rec)

    # Log low-confidence matches for diagnostics
    low_conf = [c for c in confidence_log if 0 < c["confidence"] < 100]
    if low_conf:
        logger.info("Fuzzy-matched %d columns with <100%% confidence: %s",
                    len(low_conf), json.dumps(low_conf[:20], default=str))

    return {
        "column_mappings": fixed_rows,
        "mismatch_highlights": mismatches[:200],
        "validated_count": validated,
        "unresolved_count": unresolved,
        "confidence_details": low_conf[:50],
    }


def _kb_columns_for_table(table_index: Dict[Tuple[str, str], Dict[str, Any]], schema: str, table: str) -> Dict[str, str]:
    row = _resolve_table_from_index(table_index, schema, table)
    if not row:
        return {}
    return row.get("columns", {}) or {}


def _identifiers_from_text(text: str) -> List[str]:
    if not text:
        return []
    tokens = re.findall(r'\b[A-Z][A-Z0-9_]*\b', text.upper())
    stop = {
        "CASE", "WHEN", "THEN", "ELSE", "END", "AND", "OR", "NULL", "IS", "NOT",
        "LEFT", "RIGHT", "JOIN", "OUTER", "ON", "WHERE", "AS", "FROM", "USE", "LOOKUP",
    }
    return [t for t in tokens if t not in stop]


def _extract_lookup_filters(transformation: str) -> str:
    upper = (transformation or "").upper()
    extra_filter = ""

    for scm_pat in [r'CL_SCM_(?:ID|CD)\s*=\s*(\d+)', r'SCM_ID\s*=\s*(\d+)']:
        scm_match = re.search(scm_pat, upper)
        if scm_match:
            extra_filter += f" AND LK.CL_SCM_ID = {scm_match.group(1)}"
            break

    if "ACTV_F" in upper and ("'Y'" in upper or '"Y"' in upper or "= Y" in upper):
        extra_filter += " AND LK.ACTV_F = 'Y'"

    where_match = re.search(r'WHERE\s+([^;]+?)(?:$|ORDER|GROUP)', upper)
    if where_match:
        extra_cond = where_match.group(1).strip()
        cond_pairs = re.findall(r'([A-Z0-9_\.]+)\s*=\s*([A-Z0-9_\.\'\-]+)', extra_cond)
        for left, right in cond_pairs[:3]:
            lcol = left.split('.')[-1]
            rhs = right.strip()
            if "ACTV_F" in lcol or "CL_SCM" in lcol or lcol == "SCM_ID":
                continue
            if rhs.startswith("'") and rhs.endswith("'"):
                extra_filter += f" AND LK.{lcol} = {rhs}"
            elif re.fullmatch(r'\d+(?:\.\d+)?', rhs):
                extra_filter += f" AND LK.{lcol} = {rhs}"

    return extra_filter


def _infer_lookup_pair_from_text(transformation: str, src_attr_u: str) -> Tuple[str, str]:
    upper = (transformation or "").upper()
    source_lookup_col = src_attr_u
    lookup_join_col = src_attr_u

    use_match = re.search(r'USE\s+([A-Z0-9_]+)\s+(?:IN|UNDER)\s+([A-Z0-9_]+)', upper)
    if use_match:
        source_lookup_col = use_match.group(1)
        lookup_join_col = use_match.group(2)
    elif not use_match:
        # Fallback: handle "X UNDER Y" without USE prefix
        under_match = re.search(r'\b([A-Z0-9_]+)\s+UNDER\s+([A-Z0-9_]+)', upper)
        if under_match:
            source_lookup_col = under_match.group(1)
            lookup_join_col = under_match.group(2)

    eq_matches = re.findall(r'([A-Z0-9_\.]+)\s*=\s*([A-Z0-9_\.]+)', upper)
    scored: List[Tuple[int, str, str]] = []
    for left, right in eq_matches:
        _, lcol = _column_token_parts(left)
        _, rcol = _column_token_parts(right)
        if not lcol or not rcol:
            continue
        if lcol == src_attr_u or rcol == src_attr_u:
            continue

        # Skip literal values (e.g. CL_SCM_ID = 86 — 86 is not a column)
        if re.fullmatch(r'\d+(?:\.\d+)?', lcol) or re.fullmatch(r'\d+(?:\.\d+)?', rcol):
            continue
        score = 10
        if lcol == rcol:
            score += 100
        if lcol.endswith("_ID") or rcol.endswith("_ID"):
            score += 20
        if lcol.endswith("_DT") or rcol.endswith("_DT"):
            score += 10
        if any(alias in upper for alias in ["SUB_LOT_MSTR", "SOURCE", "SRC", " S."]):
            score += 10
        scored.append((score, lcol, rcol))

    if scored:
        _, lcol, rcol = max(scored, key=lambda item: item[0])
        if lcol == rcol:
            return lcol, rcol
        return rcol, lcol

    return source_lookup_col, lookup_join_col


def _extract_literal_expression(source_attribute: str, transformation: str) -> Optional[str]:
    source_text = (source_attribute or "").strip()
    transform_text = (transformation or "").strip()
    combined = f"{source_text} {transform_text}".strip()
    upper = combined.upper()
    if not combined:
        return None

    if source_text.upper() == "NULL" or upper == "NULL":
        return "NULL"

    default_match = re.search(r'(?:DEFAULT\s+TO|POPULATE\s+AS|AS)\s*-?\s*([A-Z0-9_]+)', upper)
    if default_match:
        literal = default_match.group(1)
        if re.fullmatch(r'\d+(?:\.\d+)?', literal):
            return literal
        return f"'{literal}'"

    return None


def _build_constant_mismatch_sql(target_fq: str, target_col: str, literal_expr: str) -> str:
    tgt_bare = target_fq.split(".")[-1]
    if literal_expr == "NULL":
        return (
            "SELECT /*+ PARALLEL(8) */\n"
            "COUNT(*) AS cnt\n"
            f"FROM {target_fq}\n"
            f"WHERE {tgt_bare}.{target_col} IS NOT NULL"
        )
    return (
        "SELECT /*+ PARALLEL(8) */\n"
        "COUNT(*) AS cnt\n"
        f"FROM {target_fq}\n"
        f"WHERE {_oracle_null_safe_neq(f'{tgt_bare}.{target_col}', literal_expr)}"
    )


def _pick_existing_column(candidates: List[str], col_map: Dict[str, str], preferred: str = "") -> str:
    if not col_map:
        return preferred or ""
    cmap = {k.upper(): v for k, v in col_map.items()}
    for c in candidates:
        if c.upper() in cmap:
            return cmap[c.upper()]
    if preferred and preferred.upper() in cmap:
        return cmap[preferred.upper()]
    return ""


def _case_expr_refs_only_known_columns(expr: str, source_alias: str, col_map: Dict[str, str]) -> bool:
    if not expr or not col_map:
        return False

    upper_expr = expr.upper()
    cmap = {k.upper() for k in col_map.keys()}
    alias_u = (source_alias or "").upper()

    # Reject references to any external alias/table (e.g. TAX_LOT_OPN_MSTR.COL).
    qualified_refs = re.findall(r'\b([A-Z0-9_]+)\.([A-Z0-9_]+)\b', upper_expr)
    for alias, col in qualified_refs:
        if alias != alias_u:
            return False
        if col not in cmap:
            return False

    # Validate bare identifiers as source columns unless they are SQL keywords/functions.
    scrubbed = re.sub(r"'[^']*'", " ", upper_expr)
    tokens = re.findall(r'\b[A-Z][A-Z0-9_]*\b', scrubbed)
    allowed = {
        "CASE", "WHEN", "THEN", "ELSE", "END", "AND", "OR", "NOT", "NULL", "IS",
        "NVL", "TO_CHAR", "CAST", "DECODE", "TRIM", "REPLACE", "SUBSTR", "COALESCE",
    }
    for tok in tokens:
        if tok in allowed or tok == alias_u:
            continue
        if tok in cmap:
            continue
        # ignore numeric-like tokens accidentally captured
        if tok.isdigit():
            continue
        return False

    return True

# Known DRD header patterns (case-insensitive partial match)
DRD_HEADER_PATTERNS = {
    "logical_name": [
        "logical name of attribute",
        "logical name",
    ],
    "physical_name": [
        "physical name of attribute",
        "physical name of atribute",
        "physical name",
        "target attribute name",
        "target atribute name",
        "target column name",
        "target column",
        "target field",
        "target attribute",
        "target atribute",
        "column name",
    ],
    "target_datatype_oracle": [
        "data type in oracle",
        "target data type",
        "target datatype",
        "oracle data type",
        "oracle datatype",
    ],
    "target_nullable_oracle": [
        "nullable in table and in oracle",
        "nullable in table",
        "target nullable",
    ],
    "target_datatype_redshift": [
        "data type in redshift",
        "redshift data type",
        "redshift datatype",
    ],
    "target_nullable_redshift": [
        "nullable in redshift",
        "nullable in view in redshift",
    ],
    "business_definition": [
        "business definition",
        "description",
        "business description",
    ],
    "sample_data": [
        "sample data values",
        "sample data",
    ],
    "action": [
        "action on attribute",
    ],
    "action_version": [
        "action version",
    ],
    "pbi_number": [
        "table pbi number",
        "pbi number",
        "pbi#",
    ],
    "functional_pbi": [
        "list functional pbi",
    ],
    "include_in_view": [
        "include  in view",
        "include in view",
    ],
    "is_on_data_model": [
        "is on data model",
    ],
    "source_schema": [
        "source schema",
        "src schema",
    ],
    "source_table": [
        "source table name",
        "source table",
        "source entity",
        "src table",
    ],
    "source_attribute": [
        "source attribute name",
        "source atribute name",
        "source attribute",
        "source atribute",
        "source column name",
        "source column",
        "source field",
        "src attribute",
        "src atribute",
        "src column",
    ],
    "source_datatype": [
        "source data type",
        "source datatype",
        "data type",   # after source_attribute
    ],
    "source_nullable": [
        "null/not null",
        "nullable?",
        "nullable",
    ],
    "transformation": [
        "transformation /business rules /join conditions",
        "transformation/business rules/join conditions",
        "transformation",
        "business rules",
    ],
    "notes": [
        "notes / comments",
        "notes/comments",
        "notes",
        "comments on when it was added",
    ],
    "indicator": [
        "indicator",
    ],
    "domain": [
        "domain",
    ],
    "cross_domain": [
        "cross-domain",
        "cross domain",
    ],
}


def preview_file(file_bytes: bytes, filename: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """Parse file headers and first few rows for preview.

    Returns:
        {
            "headers": [...],              # raw header strings
            "mapped_headers": {...},       # field_name -> column_index
            "sample_rows": [...],          # first 5 data rows
            "total_rows": int,
            "is_drd_format": bool,
            "suggested_columns": [...],    # columns recommended for import
            "metadata": {...},             # DRD header metadata (table_name, schemas, grain, etc.)
            "sheets": [...],               # sheet names for Excel files
        }
    """
    rows = _read_raw_rows(file_bytes, filename, sheet_name=sheet_name)
    if not rows:
        return {"headers": [], "mapped_headers": {}, "sample_rows": [], "total_rows": 0,
                "is_drd_format": False, "suggested_columns": []}

    # Find the actual header row (skip grouping rows)
    header_row_idx, headers = _find_header_row(rows)
    data_rows = rows[header_row_idx + 1:]

    # Map headers to known DRD fields
    mapped = _map_drd_headers(headers)
    is_drd = len(mapped) >= 3  # at least logical/physical name + source table

    # Determine suggested columns (useful ones for mapping)
    suggested = []
    used_indexes = set()
    important_fields = [
        "logical_name", "physical_name", "source_schema", "source_table",
        "source_attribute", "transformation", "notes",
        "target_datatype_oracle", "target_datatype_redshift",
        "business_definition", "action", "pbi_number",
        "target_nullable_oracle", "target_nullable_redshift",
        "source_datatype", "source_nullable",
    ]
    for field in important_fields:
        if field in mapped:
            col_idx = mapped[field]
            used_indexes.add(col_idx)
            suggested.append({
                "field": field,
                "header": headers[col_idx] if col_idx < len(headers) else "",
                "column_index": col_idx,
                "selected": field in (
                    "logical_name", "physical_name", "source_schema", "source_table",
                    "source_attribute", "transformation", "notes",
                    "target_datatype_oracle", "target_datatype_redshift",
                    "business_definition",
                ),
            })

    # Include all remaining headers so "Select Columns to Import" shows complete list
    for col_idx, header in enumerate(headers):
        if col_idx in used_indexes:
            continue
        if not str(header or "").strip():
            continue
        suggested.append({
            "field": f"raw_col_{col_idx}",
            "header": str(header),
            "column_index": col_idx,
            "selected": True,
        })

    # Build sample rows (first 5 with non-empty data)
    sample_rows = []
    for row in data_rows[:5]:
        sample_rows.append([str(cell) if cell else "" for cell in row])

    return {
        "headers": headers,
        "mapped_headers": mapped,
        "sample_rows": sample_rows,
        "total_rows": len(data_rows),
        "is_drd_format": is_drd,
        "suggested_columns": suggested,
        "metadata": extract_drd_metadata(file_bytes, filename, sheet_name=sheet_name),
    }


def parse_drd_file(
    file_bytes: bytes,
    filename: str,
    selected_fields: List[str],
    target_schema: str,
    target_table: str,
    source_datasource_id: int = 1,
    target_datasource_id: int = 1,
    default_source_table: str = "",
    sheet_name: Optional[str] = None,
    header_row_override: Optional[int] = None,
    exclude_strikethrough: bool = False,
) -> Dict[str, Any]:
    """Parse a DRD file and create mapping rules grouped by source table.

    Args:
        file_bytes: raw file content
        filename: original filename
        selected_fields: list of DRD field names to include
        target_schema: target schema name (user-provided)
        target_table: target table name (user-provided)
        source_datasource_id: datasource ID for source
        target_datasource_id: datasource ID for target
        default_source_table: fallback table name if not in file (optional)
        exclude_strikethrough: when True, skip struck-through rows (de-scoped DRD entries)

    Returns:
        {
            "rules": [...],           # list of mapping rule dicts
            "column_mappings": [...],  # individual column-level mappings for tests
            "stats": {...},
            "errors": [...],
        }
    """
    rows = _read_raw_rows(file_bytes, filename, sheet_name=sheet_name, exclude_strikethrough=exclude_strikethrough)
    if not rows:
        return {"rules": [], "column_mappings": [], "stats": {}, "errors": ["Empty file"]}

    header_row_idx, headers = _find_header_row(rows)
    if header_row_override is not None:
        # Override is 1-based Excel row index from canonical v16 detection.
        try:
            forced_row = int(header_row_override)
        except (TypeError, ValueError):
            forced_row = -1
        if forced_row >= 1:
            forced_idx = forced_row - 1
            if forced_idx < len(rows):
                header_row_idx = forced_idx
                forced_header = rows[header_row_idx] if header_row_idx < len(rows) else []
                headers = [str(h).strip() if h is not None else "" for h in forced_header]
        else:
            errors.append(
                f"Invalid header_row_override={header_row_override!r}; using auto-detected header row"
            )
    data_rows = rows[header_row_idx + 1:]
    mapped = _map_drd_headers(headers)

    # Extract metadata from DRD preamble if present
    metadata = extract_drd_metadata(file_bytes, filename, sheet_name=sheet_name)

    column_mappings = []
    errors = []

    for row_idx, row in enumerate(data_rows, start=header_row_idx + 2):
        if not any(cell for cell in row):
            continue

        try:
            record = _extract_record(row, mapped, selected_fields)
            if not record.get("physical_name") and not record.get("logical_name"):
                continue

            # Best-effort 1:1 fallback when source attribute is omitted.
            if not (record.get("source_attribute") or "").strip():
                if (record.get("physical_name") or "").strip():
                    record["source_attribute"] = (record.get("physical_name") or "").strip()
                elif (record.get("logical_name") or "").strip():
                    record["source_attribute"] = (record.get("logical_name") or "").strip().upper().replace(" ", "_")

            # Clean source_table: when a DRD cell contains multiple table names
            # (newline- or comma-separated), use only the first one as the primary source.
            _st_raw = (record.get("source_table") or "").strip()
            if _st_raw and ("\n" in _st_raw or "," in _st_raw):
                record["source_table"] = _st_raw.split("\n")[0].split(",")[0].strip()

            # Apply default source table if not specified in record
            if not record.get("source_table") and default_source_table:
                record["source_table"] = default_source_table

            column_mappings.append(record)
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    # Pandera structural validation on parsed rows
    pandera_warnings = _validate_parsed_rows(column_mappings)
    if pandera_warnings:
        errors.extend([f"[validation] {w}" for w in pandera_warnings])

    # Group by source table to create mapping rules
    rules, grouping_errors = _group_into_rules(
        column_mappings, target_schema, target_table,
        source_datasource_id, target_datasource_id,
    )
    errors.extend(grouping_errors)

    return {
        "rules": rules,
        "column_mappings": column_mappings,
        "stats": {
            "total_rows": len(data_rows),
            "parsed_columns": len(column_mappings),
            "generated_rules": len(rules),
            "errors": len(errors),
        },
        "errors": errors,
        "metadata": metadata,
    }


def generate_e2e_scripts(
    column_mappings: List[Dict[str, Any]],
    target_schema: str,
    target_table: str,
    test_rows: int = 10,
    pdm_filter: bool = False,
    target_datasource_id: int = 1,
) -> Dict[str, Any]:
    """Generate CREATE TABLE, INSERT, and validation SQL scripts from DRD column mappings.

    Uses physical_name + target_datatype_oracle + target_nullable_oracle from each record.
    All scripts are offline-generated (no DB connection required).

    Returns:
        {
            "create_table_sql": str,
            "insert_sql": str,
            "validation_sqls": [{"label": str, "sql": str}],
            "column_count": int,
            "columns": [{"name", "data_type", "nullable"}],
        }
    """
    import random
    import string

    fq_table = f"{target_schema.upper()}.{target_table.upper()}" if target_schema else target_table.upper()

    # ── Collect distinct columns by physical_name ──────────────────────────
    seen: set[str] = set()
    columns: list[dict] = []
    for rec in column_mappings:
        pname = (rec.get("physical_name") or "").strip().upper()
        if not pname or pname in seen:
            continue
        seen.add(pname)
        raw_type = (rec.get("target_datatype_oracle") or "VARCHAR2(255)").strip()
        nullable_raw = (rec.get("target_nullable_oracle") or "NULL").strip().upper()
        nullable = (nullable_raw != "NOT NULL")
        columns.append({"name": pname, "data_type": raw_type, "nullable": nullable})

    # ── PDM filter: keep only columns that exist in the PDM target table ─────
    pdm_cols: set = set()
    pdm_extra_cols: list = []
    if pdm_filter:
        try:
            tbl_index = _build_kb_table_index(target_datasource_id)
            schema_u = target_schema.strip().upper()
            table_u = target_table.strip().upper()
            tbl_entry = tbl_index.get((schema_u, table_u))
            if tbl_entry is None:
                for (s, t), entry in tbl_index.items():
                    if t == table_u:
                        tbl_entry = entry
                        break
            if tbl_entry:
                pdm_cols = set(tbl_entry.get("columns", {}).keys())
        except Exception as exc:
            logger.warning("PDM filter lookup failed for %s.%s: %s", target_schema, target_table, exc)

    if pdm_filter and pdm_cols:
        pdm_extra_cols = [c["name"] for c in columns if c["name"].upper() not in pdm_cols]
        columns = [c for c in columns if c["name"].upper() in pdm_cols]

    if not columns:
        return {
            "create_table_sql": "",
            "insert_sql": "",
            "validation_sqls": [],
            "column_count": 0,
            "columns": [],
            "pdm_extra_cols": pdm_extra_cols,
            "pdm_filtered": pdm_filter,
        }

    # ── CREATE TABLE ──────────────────────────────────────────────────────
    col_defs: list[str] = []
    for c in columns:
        null_clause = "" if c["nullable"] else " NOT NULL"
        col_defs.append(f"  {c['name']} {c['data_type']}{null_clause}")
    create_sql = (
        f"CREATE TABLE {fq_table} (\n" + ",\n".join(col_defs) + "\n)"
    )

    # ── INSERT with synthetic literals ────────────────────────────────────
    def _fake(col_name: str, data_type: str, idx: int) -> str:
        dt = data_type.upper().split("(")[0].strip()
        if dt in ("NUMBER", "INTEGER", "FLOAT", "BINARY_FLOAT", "BINARY_DOUBLE"):
            return str(idx + 1)
        if dt == "DATE":
            y = 2020 + (idx % 5); m = (idx % 12) + 1
            return f"DATE '{y:04d}-{m:02d}-01'"
        if dt.startswith("TIMESTAMP"):
            y = 2020 + (idx % 5); m = (idx % 12) + 1
            return f"TIMESTAMP '{y:04d}-{m:02d}-01 00:00:00'"
        if dt == "CHAR":
            return f"'C{idx:03d}'"
        suffix = "".join(random.choices(string.ascii_uppercase, k=4))
        prefix = col_name[:6]
        return f"'T_{prefix}_{suffix}'"

    col_names_csv = ", ".join(c["name"] for c in columns)
    insert_parts: list[str] = []
    for i in range(test_rows):
        vals = ", ".join(_fake(c["name"], c["data_type"], i) for c in columns)
        insert_parts.append(
            f"INSERT INTO {fq_table} ({col_names_csv})\nVALUES ({vals});"
        )
    insert_sql = "\n".join(insert_parts)

    # ── Validation queries ────────────────────────────────────────────────
    validation_sqls: list[dict] = []
    # 1. Row count
    validation_sqls.append({
        "label": "row_count",
        "sql": f"SELECT COUNT(*) AS row_count FROM {fq_table};",
    })
    # 2. NOT NULL checks (columns declared NOT NULL must have zero NULLs)
    for c in columns:
        if not c["nullable"]:
            validation_sqls.append({
                "label": f"null_check_{c['name']}",
                "sql": (
                    f"SELECT COUNT(*) AS null_violations FROM {fq_table}\n"
                    f"WHERE {c['name']} IS NULL; -- expect 0"
                ),
            })
    # 3. Spot check per column: distinct count
    for c in columns:
        validation_sqls.append({
            "label": f"distinct_{c['name']}",
            "sql": f"SELECT COUNT(DISTINCT {c['name']}) AS distinct_vals FROM {fq_table};",
        })

    return {
        "create_table_sql": create_sql,
        "insert_sql": insert_sql,
        "validation_sqls": validation_sqls,
        "column_count": len(columns),
        "columns": columns,
        "pdm_extra_cols": pdm_extra_cols,
        "pdm_filtered": pdm_filter,
    }


# ── Internal helpers ────────────────────────────────────────────────────────

def _read_raw_rows(
    file_bytes: bytes,
    filename: str,
    sheet_name: Optional[str] = None,
    exclude_strikethrough: bool = False,
) -> List[List[Any]]:
    """Read rows from CSV or Excel file.  For xlsx, *sheet_name* selects a specific tab."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext == "csv":
        return _read_csv(file_bytes)
    elif ext in ("xlsx", "xls"):
        return _read_excel(file_bytes, sheet_name=sheet_name, exclude_strikethrough=exclude_strikethrough)
    else:
        # Try CSV first, then Excel
        try:
            return _read_csv(file_bytes)
        except Exception:
            return _read_excel(file_bytes, sheet_name=sheet_name, exclude_strikethrough=exclude_strikethrough)


def _read_csv(file_bytes: bytes) -> List[List[Any]]:
    """Read CSV file, handling various encodings."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = file_bytes.decode(encoding)
            reader = csv.reader(io.StringIO(text))
            return [row for row in reader]
        except (UnicodeDecodeError, csv.Error):
            continue
    raise ValueError("Unable to decode CSV file")


# ── Workbook cache (one parse per upload) ────────────────────────────────────
_workbook_cache: Dict[str, Any] = {}  # md5(file_bytes) → openpyxl.Workbook


def _get_workbook(file_bytes: bytes) -> Any:
    """Return a workbook, caching by content hash to avoid re-parsing."""
    key = hashlib.md5(file_bytes).hexdigest()
    wb = _workbook_cache.get(key)
    if wb is None:
        # Keep workbook in normal mode so font/style metadata (strike-through) is available.
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        _workbook_cache[key] = wb
        # Keep cache bounded (LRU-ish: evict oldest if >5)
        if len(_workbook_cache) > 5:
            oldest_key = next(iter(_workbook_cache))
            try:
                _workbook_cache[oldest_key].close()
            except Exception:
                pass
            _workbook_cache.pop(oldest_key, None)
    return wb


# Source mapping columns in the DRD fixed layout:
#   Y(24) = Source Schema, Z(25) = Source Table, AA(26) = Source Attribute
# Only rows struck in THESE columns are deprecated/de-scoped entries.
# Columns 0-23 and 27+ contain reference/formula cells that are commonly struck
# for non-structural reasons (formula errors, dated notes, secondary refs) and
# must NOT trigger row exclusion.
_SOURCE_COL_INDICES = (24, 25, 26)


def _is_row_strikethrough(row) -> bool:
    """Return True if a source-mapping DRD column (Y/Z/AA) has strike-through font."""
    cells = list(row)
    for idx in _SOURCE_COL_INDICES:
        if idx >= len(cells):
            continue
        cell = cells[idx]
        if cell.value is None:
            continue
        font = getattr(cell, "font", None)
        if font is not None and getattr(font, "strike", False):
            return True
    return False


def _read_excel(
    file_bytes: bytes,
    sheet_name: Optional[str] = None,
    exclude_strikethrough: bool = False,
) -> List[List[Any]]:
    """Read Excel file in read-only mode with caching.

    When *sheet_name* is given only that sheet is read, otherwise the best
    DRD data-sheet is auto-detected.

    When *exclude_strikethrough* is True, any row where at least one non-empty
    cell has strike-through font is skipped (de-scoped DRD rows).
    """
    wb = _get_workbook(file_bytes)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = _pick_best_drd_sheet(wb)

    rows = []
    for row in ws.iter_rows():
        if exclude_strikethrough and _is_row_strikethrough(row):
            continue
        rows.append([cell.value for cell in row])
    return rows


def _pick_best_drd_sheet(wb) -> Any:
    """Pick the best DRD data sheet from an openpyxl Workbook.

     Priority:
     1. Prefer ``Table-View (2)`` style tabs when present
     2. Among ``Table-View*`` tabs, pick the one with strongest header match
         and the highest non-empty physical-name coverage in column B below header
     3. Otherwise pick the sheet with strongest DRD header indicators
     4. Fallback to active (first) sheet
    """
    _header_indicators = [
        "logical name", "physical name", "target column", "target attribute",
        "source attribute", "source column", "column name", "source table",
        "transformation", "data type",
    ]
    def _norm_name(name: str) -> str:
        return re.sub(r"[\s_\-]+", "", (name or "").lower())

    def _header_score(ws) -> int:
        best = 0
        for row_vals in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            cells = [str(c).strip().lower() if c else "" for c in row_vals]
            score = sum(1 for ind in _header_indicators if any(ind in c for c in cells))
            if score > best:
                best = score
        return best

    def _col_b_non_empty_below_header(ws) -> int:
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        if not rows:
            return 0
        header_idx, _headers = _find_header_row(rows)
        count = 0
        for row in rows[header_idx + 1:]:
            # Column B contains physical target name in enterprise DRD template.
            val = row[1] if len(row) > 1 else None
            if val is not None and str(val).strip():
                count += 1
        return count

    candidates = [
        sn for sn in wb.sheetnames
        if _norm_name(sn).startswith("tableview")
    ]

    if candidates:
        scored = []
        for sn in candidates:
            ws = wb[sn]
            ns = _norm_name(sn)
            # Highest priority: exact "tableview" match (i.e. "Table-View") — operator locked
            exact_match = 1 if ns == "tableview" else 0
            prefers_tab2 = 1 if "(2)" in sn or ns.endswith("2") else 0
            scored.append((
                exact_match,
                prefers_tab2,
                _header_score(ws),
                _col_b_non_empty_below_header(ws),
                ws,
            ))
        scored.sort(key=lambda t: (t[0], t[1], t[2], t[3]), reverse=True)
        return scored[0][4]

    # Fallback: score each sheet by header indicator matches.
    best_sheet = None
    best_score = -1
    for sn in wb.sheetnames:
        ws = wb[sn]
        score = _header_score(ws)
        if score > best_score:
            best_score = score
            best_sheet = ws
            if score >= 3:
                return ws
    return best_sheet or wb.active


def read_excel_all_sheets(file_bytes: bytes) -> Dict[str, List[List[Any]]]:
    """Read ALL sheets from an Excel file.  Returns ``{sheet_name: rows}``."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result: Dict[str, List[List[Any]]] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        result[sn] = rows
    return result


def extract_drd_metadata(file_bytes: bytes, filename: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """Extract metadata block from DRD Excel/CSV files.

    Many enterprise DRD files have a metadata preamble in rows 1-8 that contains:
    - View Name, Table Name, Table Schema Name, Source Schema, Filter Criteria, Grain
    These are typically in (label, value) pairs in columns A-B.
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in ("xlsx", "xls"):
        rows = _read_excel(file_bytes, sheet_name=sheet_name)
    elif ext == "csv":
        rows = _read_csv(file_bytes)
    else:
        # Phase 7.16 silent-failure round 2 fix: was bare except -> a
        # malformed CSV was silently re-read as Excel (openpyxl parses
        # binary as cell data, producing garbage rows downstream).  Now
        # only specific CSV-parse exceptions trigger the Excel fallback.
        try:
            rows = _read_csv(file_bytes)
        except (csv.Error, UnicodeDecodeError) as _csv_exc:
            import logging
            logging.getLogger(__name__).info(
                "Ambiguous-extension file failed CSV parse (%s); trying Excel.",
                _csv_exc,
            )
            rows = _read_excel(file_bytes, sheet_name=sheet_name)

    meta: Dict[str, Any] = {"sheets": []}

    # If xlsx, also list all sheet names
    if ext in ("xlsx", "xls"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            meta["sheets"] = wb.sheetnames
        except Exception as _xlsx_exc:
            # Phase 7.16 silent-failure round 2 fix: was bare pass.  Sheet
            # names dropping silently looks identical to "single-sheet CSV"
            # to callers downstream.  Now surface the error in meta.
            import logging
            logging.getLogger(__name__).warning(
                "Failed to read xlsx sheet names: %s", _xlsx_exc,
            )
            meta["sheets_error"] = f"{type(_xlsx_exc).__name__}: {_xlsx_exc}"

    # Scan first 12 rows for label-value metadata pairs
    _meta_keys = {
        "view name": "view_name",
        "table name": "table_name",
        "table schema name": "table_schema",
        "table schema": "table_schema",
        "schema name": "table_schema",
        "target schema": "table_schema",
        "source schema": "source_schema",
        "source table": "source_table",
        "filter criteria": "filter_criteria",
        "grain": "grain",
        "main grain": "grain",
        "description": "description",
    }
    for row in rows[:12]:
        if not row or len(row) < 2:
            continue
        label = str(row[0] or "").strip().lower()
        value = str(row[1] or "").strip() if row[1] else ""
        for pattern, key in _meta_keys.items():
            if pattern in label and value:
                meta[key] = value
                if key == "table_name":
                    # A "Table Name (From DA Team)" row may carry BOTH the
                    # logical/DA-team name and the physical table name in
                    # adjacent cells (e.g. "cls_tax_lots_fact_rjt |
                    # CLS_TAX_LOTS_NON_BKR_FACT").  Capture every cell that
                    # looks like a table identifier as a resolution candidate
                    # so the control-table flow can fall back to the physical
                    # name when the logical one is absent from the PDM.
                    cands: List[str] = []
                    for cell in row[1:]:
                        cv = str(cell or "").strip()
                        if cv and re.match(r"^[A-Za-z][A-Za-z0-9_$#]*(\.[A-Za-z][A-Za-z0-9_$#]*)?$", cv) \
                                and cv.upper() not in {x.upper() for x in cands}:
                            cands.append(cv)
                    if len(cands) > 1:
                        meta["table_name_candidates"] = cands
                break
    return meta


def _find_header_row(rows: List[List[Any]]) -> Tuple[int, List[str]]:
    """Find the actual header row in DRD files (skip grouping rows like 'Lighthouse Target').

    Recognizes header rows containing typical DRD/mapping column names:
    logical name, physical name, source attribute, target column, column name, etc.
    """
    _header_indicators = [
        "logical name", "physical name", "target column", "target attribute",
        "source attribute", "source column", "column name", "source table",
        "transformation", "data type",
    ]
    for idx, row in enumerate(rows[:15]):
        cells = [str(c).strip().lower() if c else "" for c in row]
        matches = sum(1 for ind in _header_indicators if any(ind in c for c in cells))
        if matches >= 2:
            return idx, [str(c).strip() if c else "" for c in row]

    # Fallback: first non-empty row with 3+ non-empty cells
    for idx, row in enumerate(rows[:10]):
        non_empty = sum(1 for c in row if c and str(c).strip())
        if non_empty >= 3:
            return idx, [str(c).strip() if c else "" for c in row]

    return 0, [str(c).strip() if c else "" for c in rows[0]] if rows else (0, [])


def _map_drd_headers(headers: List[str]) -> Dict[str, int]:
    """Map raw headers to known DRD field names. Returns field_name -> column_index."""
    mapped = {}

    def _normalize(value: str) -> str:
        value = (value or "").lower().strip()
        value = re.sub(r"\s+", " ", value)
        return value

    headers_lower = [_normalize(h) for h in headers]

    # Track which columns are already assigned to avoid double-mapping
    assigned_cols = set()

    # Map in priority order (more specific patterns first)
    for field, patterns in DRD_HEADER_PATTERNS.items():
        for pattern in patterns:
            normalized_pattern = _normalize(pattern)
            for col_idx, header in enumerate(headers_lower):
                if col_idx in assigned_cols:
                    continue
                if normalized_pattern in header:
                    mapped[field] = col_idx
                    assigned_cols.add(col_idx)
                    break
            if field in mapped:
                break

    # Heuristic fallback: support non-template DRD headers when intent is clear.
    def _contains_all(header: str, tokens: List[str]) -> bool:
        return all(tok in header for tok in tokens)

    fallback_checks = {
        "source_schema": lambda h: _contains_all(h, ["source", "schema"]),
        "source_table": lambda h: "source" in h and any(tok in h for tok in ("table", "view", "entity")),
        "source_attribute": lambda h: "source" in h and any(tok in h for tok in ("attribute", "column", "field", "col")),
        "physical_name": lambda h: ("target" in h and any(tok in h for tok in ("column", "attribute", "field", "col"))) or ("physical" in h and "name" in h),
        "logical_name": lambda h: "logical" in h and "name" in h,
        "transformation": lambda h: any(tok in h for tok in ("transform", "business rule", "rule", "logic", "calculation", "expression")),
        "source_datatype": lambda h: "source" in h and "type" in h,
        "target_datatype_oracle": lambda h: ("target" in h and "type" in h and "redshift" not in h) or ("oracle" in h and "type" in h),
        "target_datatype_redshift": lambda h: "redshift" in h and "type" in h,
        "notes": lambda h: any(tok in h for tok in ("note", "comment", "remark")),
    }
    for field, predicate in fallback_checks.items():
        if field in mapped:
            continue
        for col_idx, header in enumerate(headers_lower):
            if col_idx in assigned_cols:
                continue
            if predicate(header):
                mapped[field] = col_idx
                assigned_cols.add(col_idx)
                break

    return mapped


_NULLABLE_FLAG_TOKENS = {
    "", "Y", "N", "YES", "NO", "TRUE", "FALSE", "NULL", "NOT NULL",
    "NULLABLE", "NOT NULLABLE", "0", "1", "NA", "N/A", "NONE",
}


def _is_misplaced_transformation_prose(text: str) -> bool:
    """True when a "Nullable?" cell actually holds a transformation rule (prose),
    not a Y/N nullable flag.

    Some DRD authors type the lookup/transformation rule into the "Nullable?"
    column (C25) instead of the "Transformation" column (C26) -- a recurring
    authoring slip on lookup columns (e.g. the CL_VAL scheme rules in the CLOSE
    taxlot DRD).  A real nullable flag is a short Y/N-style token; anything with
    whitespace or lookup/scheme keywords is a misplaced rule.  (operator 2026-06-04)
    """
    t = (text or "").strip()
    if not t or t.upper() in _NULLABLE_FLAG_TOKENS:
        return False
    return (" " in t) or bool(re.search(r"CL_VAL|CL_SCM_ID|LOOKUP|\bUNDER\b", t, re.IGNORECASE))


def _extract_record(
    row: List[Any], mapped: Dict[str, int], selected_fields: List[str]
) -> Dict[str, Any]:
    """Extract a record from a row using the mapped header positions."""
    record = {}
    for field in selected_fields:
        if field.startswith("raw_col_"):
            try:
                idx = int(field.split("_")[-1])
                val = row[idx] if idx < len(row) else None
                record[field] = str(val).strip() if val is not None else None
            except Exception:
                record[field] = None
            continue
        if field in mapped:
            idx = mapped[field]
            val = row[idx] if idx < len(row) else None
            record[field] = str(val).strip() if val is not None else None
        else:
            record[field] = None

    # Recover a transformation rule misplaced in the "Nullable?" column: when the
    # mapped Transformation cell is empty but the source_nullable cell holds prose
    # (not a Y/N flag), fold it into transformation so the builder can honor the
    # lookup.  Without this, DRD CL_VAL scheme rules ("... where CL_SCM_ID = 86 and
    # pick CL_VAL_NM") are silently dropped.  (operator 2026-06-04)
    if "transformation" in selected_fields and not (record.get("transformation") or "").strip():
        nidx = mapped.get("source_nullable")
        if nidx is not None and nidx < len(row):
            nval = row[nidx]
            ntext = str(nval).strip() if nval is not None else ""
            if _is_misplaced_transformation_prose(ntext):
                record["transformation"] = ntext
    return record


def _group_into_rules(
    column_mappings: List[Dict[str, Any]],
    target_schema: str,
    target_table: str,
    source_datasource_id: int,
    target_datasource_id: int,
) -> Tuple[List[Dict], List[str]]:
    """Group column-level mappings into mapping rules by source table.

    Creates one MappingRule per unique source table, aggregating all columns
    and transformations.
    """
    rules = []
    errors = []

    # Group by source_schema + source_table
    groups: Dict[str, List[Dict]] = {}
    no_source = []

    for cm in column_mappings:
        src_schema = cm.get("source_schema") or ""
        src_table = cm.get("source_table") or ""

        if not src_table:
            no_source.append(cm)
            continue

        key = f"{src_schema}.{src_table}" if src_schema else src_table
        groups.setdefault(key, []).append(cm)

    # Also create a consolidated "all columns" rule
    all_source_cols = []
    all_target_cols = []
    all_transformations = []
    all_joins = []
    all_descriptions = []

    for key, columns in groups.items():
        parts = key.split(".", 1) if "." in key else ("", key)
        src_schema = parts[0] if len(parts) > 1 else ""
        src_table = parts[1] if len(parts) > 1 else parts[0]

        source_cols = []
        target_cols = []
        transformations = []
        join_parts = []
        desc_parts = []

        for cm in columns:
            src_attr = cm.get("source_attribute") or ""
            physical = cm.get("physical_name") or ""
            logical = cm.get("logical_name") or ""
            transform = cm.get("transformation") or ""
            notes = cm.get("notes") or ""

            if src_attr:
                source_cols.append(src_attr)
            if physical:
                target_cols.append(physical)

            # Extract join conditions from transformation text
            if transform:
                transform_lower = transform.lower()
                if "join" in transform_lower or "left join" in transform_lower:
                    join_parts.append(transform)
                transformations.append(f"{physical}: {transform}" if physical else transform)

            if notes:
                desc_parts.append(f"{physical}: {notes}" if physical else notes)

            # Track for consolidated rule
            all_source_cols.append({"source_schema": src_schema, "source_table": src_table,
                                    "source_attribute": src_attr, "physical_name": physical,
                                    "logical_name": logical, "transformation": transform})

        rule = {
            "name": f"{src_table} → {target_table}",
            "rule_type": "lookup" if join_parts else "direct",
            "source_datasource_id": source_datasource_id,
            "source_schema": src_schema,
            "source_table": src_table,
            "source_columns": json.dumps(source_cols) if source_cols else None,
            "target_datasource_id": target_datasource_id,
            "target_schema": target_schema,
            "target_table": target_table,
            "target_columns": json.dumps(target_cols) if target_cols else None,
            "transformation_sql": "\n".join(transformations) if transformations else None,
            "join_condition": "\n".join(join_parts) if join_parts else None,
            "filter_condition": None,
            "description": "\n".join(desc_parts) if desc_parts else f"Mapping from {key} to {target_schema}.{target_table}",
        }
        rules.append(rule)

    # Handle columns without source tables (e.g., surrogate keys, derived columns)
    if no_source:
        target_cols = []
        desc_parts = []
        for cm in no_source:
            physical = cm.get("physical_name") or ""
            logical = cm.get("logical_name") or ""
            if physical:
                target_cols.append(physical)
            desc = cm.get("business_definition") or cm.get("notes") or ""
            if desc and physical:
                desc_parts.append(f"{physical}: {desc}")

        if target_cols:
            rules.append({
                "name": f"Derived/Surrogate → {target_table}",
                "rule_type": "custom",
                "source_datasource_id": source_datasource_id,
                "source_schema": "",
                "source_table": target_table,
                "source_columns": None,
                "target_datasource_id": target_datasource_id,
                "target_schema": target_schema,
                "target_table": target_table,
                "target_columns": json.dumps(target_cols),
                "transformation_sql": None,
                "join_condition": None,
                "filter_condition": None,
                "description": "\n".join(desc_parts) if desc_parts else "Derived/surrogate columns",
            })

    return rules, errors


def generate_drd_tests(
    column_mappings: List[Dict[str, Any]],
    target_schema: str,
    target_table: str,
    source_datasource_id: int = 1,
    target_datasource_id: int = 1,
    max_row_count_tests: int = 0,
    main_grain: str = "",
    single_db_testing: bool = True,
    cross_db_optional: bool = True,
    include_diagnostics: bool = False,
    default_source_table: str = "",
) -> List[Dict[str, Any]] | Dict[str, Any]:
    """Generate test case definitions from DRD column mappings.

    Creates two types of tests:
    1. Record count tests: source table COUNT(*) vs target table COUNT(*)
    2. Transformation/mapping validation: verify each column transformation
    
    Args:
        default_source_table: fallback source table if not specified per row (optional)
    """
    tests = []
    skipped_rows = []
    lookup_alias_counts: Dict[str, int] = {}  # Track lookup table alias numbering (CL_VAL_1, CL_VAL_2, ...)

    normalized_rows: List[Dict[str, Any]] = [dict(cm or {}) for cm in (column_mappings or [])]

    # Infer source table when file omits it row-by-row.
    inferred_source_table = (default_source_table or "").strip()
    if not inferred_source_table:
        table_freq: Dict[str, int] = {}
        for cm in normalized_rows:
            tbl = (cm.get("source_table") or "").strip()
            if tbl:
                table_freq[tbl] = table_freq.get(tbl, 0) + 1
        if table_freq:
            inferred_source_table = max(table_freq.items(), key=lambda item: item[1])[0]

    for cm in normalized_rows:
        if not (cm.get("source_table") or "").strip() and inferred_source_table:
            cm["source_table"] = inferred_source_table
        if not (cm.get("source_attribute") or "").strip():
            fallback_attr = (cm.get("physical_name") or "").strip()
            if not fallback_attr:
                fallback_attr = (cm.get("logical_name") or "").strip().upper().replace(" ", "_")
            if fallback_attr:
                cm["source_attribute"] = fallback_attr

    # Group by source table for join-focused tests
    source_tables: Dict[str, List[Dict]] = {}
    for cm in normalized_rows:
        src_schema = cm.get("source_schema") or ""
        src_table = (cm.get("source_table") or default_source_table or "").strip()
        # Split FQ name to avoid double-schema keys
        if src_table and "." in src_table:
            _parts = src_table.split(".", 1)
            if not src_schema:
                src_schema = _parts[0].strip()
            src_table = _parts[1].strip()
        if src_table:
            key = f"{src_schema}.{src_table}" if src_schema else src_table
            source_tables.setdefault(key, []).append(cm)

    tgt_fq = _fq_name(target_schema, target_table)
    source_table_index = _build_kb_table_index(source_datasource_id)

    # Pre-compute reusable join keys per source table
    explicit_join_pairs = _parse_main_grain_pairs(main_grain)
    join_keys_by_source = {}
    for key, rows_for_source in source_tables.items():
        derived = _derive_join_key_pairs(rows_for_source)
        if explicit_join_pairs:
            # Build a target→source column map from the DRD rows so that
            # single-name grain entries (e.g. "TAX_LOT_EXT_REFR_KEY_PRIM")
            # resolve to the correct source column (e.g. "SUB_LOT_EXT_REFR_KEY_PRIM").
            derived_map = {tgt: src for tgt, src in derived}
            mapping_by_target: Dict[str, str] = {}
            for row in rows_for_source:
                src_a = (row.get("source_attribute") or "").strip().upper()
                tgt_a = (row.get("physical_name") or "").strip().upper()
                if src_a and tgt_a and tgt_a not in mapping_by_target:
                    mapping_by_target[tgt_a] = src_a
            resolved: List[Tuple[str, str]] = []
            for tgt_col, src_col in explicit_join_pairs:
                if tgt_col == src_col:
                    # Try DRD mapping first, then derived pairs, then keep as-is
                    real_src = mapping_by_target.get(tgt_col) or derived_map.get(tgt_col) or src_col
                    resolved.append((tgt_col, real_src))
                else:
                    resolved.append((tgt_col, src_col))
            join_keys_by_source[key] = resolved
        else:
            join_keys_by_source[key] = derived
    base_source_key = _choose_base_source_key(source_tables)
    cross_db_mode = (not bool(single_db_testing)) and int(source_datasource_id or 0) != int(target_datasource_id or 0)

    def _column_profile_sql(table_fq: str, column_name: str) -> str:
        col = (column_name or "").strip().upper()
        return (
            "SELECT /*+ PARALLEL(8) */\n"
            "COUNT(*) AS total_cnt,\n"
            f"COUNT(DISTINCT {col}) AS distinct_cnt,\n"
            f"SUM(CASE WHEN {col} IS NULL THEN 1 ELSE 0 END) AS null_cnt\n"
            f"FROM {table_fq}"
        )

    # Optional: at most one record-count test by default (disabled for join-focused generation)
    if max_row_count_tests > 0 and source_tables:
        first_src_key = next(iter(source_tables.keys()))
        parts = first_src_key.split(".", 1) if "." in first_src_key else ("", first_src_key)
        src_schema = parts[0] if len(parts) > 1 else ""
        src_table = parts[1] if len(parts) > 1 else parts[0]
        src_fq = _fq_name(src_schema, src_table)
        tests.append({
            "name": f"Record Count (single): {src_table} → {target_table}",
            "test_type": "row_count",
            "source_datasource_id": source_datasource_id,
            "target_datasource_id": target_datasource_id,
            "source_query": f"SELECT /*+ PARALLEL(DEFAULT) ENABLE_PARALLEL_QUERY*/ COUNT(*) AS cnt FROM {src_fq}",
            "target_query": f"SELECT /*+ PARALLEL(DEFAULT) ENABLE_PARALLEL_QUERY*/ COUNT(*) AS cnt FROM {tgt_fq}",
            "severity": "medium",
            "description": f"Single source-to-target volume check: {first_src_key} vs {target_schema}.{target_table}",
        })

    # 2) Join-based transformation/mapping validation tests per column
    for cm in normalized_rows:
        src_schema = cm.get("source_schema") or ""
        src_table = (cm.get("source_table") or default_source_table or "").strip()
        src_attr = cm.get("source_attribute") or ""
        physical = cm.get("physical_name") or ""
        logical = cm.get("logical_name") or ""
        transformation = cm.get("transformation") or ""
        literal_expr = _extract_literal_expression(src_attr, transformation)
        target_col = (physical or logical.upper().replace(" ", "_")).upper()

        # Normalize: if src_table is already FQ (contains dot), split schema from table.
        if src_table and "." in src_table:
            _parts = src_table.split(".", 1)
            if not src_schema:
                src_schema = _parts[0].strip()
            src_table = _parts[1].strip()
        # Back-fill missing source attribute: if target column exists, assume same name (direct 1:1).
        if not src_attr and target_col:
            src_attr = target_col

        if (not src_table or not src_attr) and literal_expr:
            tests.append({
                "name": f"Constant Mapping: {target_table}.{target_col}",
                "test_type": "custom_sql",
                "source_datasource_id": source_datasource_id,
                "target_datasource_id": target_datasource_id,
                "source_query": _build_constant_mismatch_sql(tgt_fq, target_col, literal_expr),
                "target_query": None,
                "expected_result": json.dumps({"cnt": 0}),
                "severity": "medium",
                "description": f"Validate constant/default assignment for {target_col}.",
                "source_field": src_attr,
                "target_field": target_col,
                "mapping_type": "constant",
                "transformation_rule": transformation or src_attr or "Constant mapping",
            })
            continue

        if not src_table or not src_attr:
            skipped_rows.append({
                "target_field": physical or logical or "",
                "source_table": src_table,
                "source_field": src_attr,
                "reason": "Missing source table or source attribute",
            })
            continue

        src_fq = _fq_name(src_schema, src_table)
        src_attr_u = src_attr.upper()

        src_key = f"{src_schema}.{src_table}" if src_schema else src_table
        join_pairs = join_keys_by_source.get(src_key) or []
        lookup_spec = _extract_lookup_spec(transformation, src_attr_u, target_col, src_schema, src_table)
        if lookup_spec and "." not in lookup_spec.get("lookup_table", "") and src_schema and _is_lookup_table_name(src_table):
            lookup_spec["lookup_table"] = f"{src_schema}.{lookup_spec['lookup_table']}"

        effective_src_key = src_key
        if lookup_spec and _is_lookup_table_name(src_table) and base_source_key:
            effective_src_key = base_source_key

        effective_src_schema, effective_src_table = _split_source_key(effective_src_key)
        effective_src_fq = _fq_name(effective_src_schema, effective_src_table)
        effective_src_cols = _kb_columns_for_table(source_table_index, effective_src_schema, effective_src_table)

        is_complex = bool(transformation and transformation.strip())
        mapping_type = "complex" if is_complex else "direct"

        if cross_db_mode:
            tests.append({
                "name": f"Cross-DB Profile Mapping: {effective_src_table}.{src_attr_u} → {target_table}.{target_col}",
                "test_type": "value_match",
                "source_datasource_id": source_datasource_id,
                "target_datasource_id": target_datasource_id,
                "source_query": _column_profile_sql(effective_src_fq, src_attr_u),
                "target_query": _column_profile_sql(tgt_fq, target_col),
                "severity": "high" if mapping_type == "complex" else "medium",
                "description": (
                    f"Cross-DB profile comparison for {target_col}. "
                    f"Runs source and target metrics separately to avoid cross-database joins in a single query."
                ),
                "source_field": src_attr,
                "target_field": target_col,
                "mapping_type": "cross_db_profile",
                "transformation_rule": transformation or "Direct mapping",
            })
            continue

        # Use bare table names (no aliases) for source and target
        effective_join_on = _build_join_on_sql(
            join_keys_by_source.get(effective_src_key) or [],
            target_table=target_table,
            source_table=effective_src_table,
        )
        effective_join_on_pretty = re.sub(r'\s+AND\s+', '\nAND ', effective_join_on, flags=re.IGNORECASE)

        # Resolve lookup join columns against KB to avoid invalid identifiers from free-text notes.
        if lookup_spec:
            lookup_schema, lookup_table_name = _split_source_key(lookup_spec.get("lookup_table", ""))
            lookup_cols = _kb_columns_for_table(source_table_index, lookup_schema, lookup_table_name)
            candidates = _identifiers_from_text(transformation)

            src_lookup_literal = lookup_spec.get("source_lookup_literal", "")
            src_lookup = lookup_spec.get("source_lookup_col", "")
            if not src_lookup_literal:
                resolved_src_lookup = _pick_existing_column([src_lookup, src_attr_u, *candidates], effective_src_cols, src_attr_u)
                if resolved_src_lookup:
                    lookup_spec["source_lookup_col"] = resolved_src_lookup.upper()

            lk_join = lookup_spec.get("lookup_join_col", "")
            resolved_lk_join = _pick_existing_column([lk_join, *candidates], lookup_cols, lk_join) if lookup_cols else lk_join
            if resolved_lk_join:
                lookup_spec["lookup_join_col"] = resolved_lk_join.upper()

            lk_value = lookup_spec.get("lookup_value_col", "")
            resolved_lk_value = _pick_existing_column([lk_value, target_col, *candidates], lookup_cols, lk_value) if lookup_cols else lk_value
            if resolved_lk_value:
                lookup_spec["lookup_value_col"] = resolved_lk_value.upper()

            # If lookup metadata is still not valid after KB resolution, disable lookup test.
            src_ok = bool(src_lookup_literal) or lookup_spec.get("source_lookup_col", "").upper() in {k.upper() for k in effective_src_cols.keys()}
            lk_join_ok = (not lookup_cols) or lookup_spec.get("lookup_join_col", "").upper() in {k.upper() for k in lookup_cols.keys()}
            lk_val_ok = (not lookup_cols) or lookup_spec.get("lookup_value_col", "").upper() in {k.upper() for k in lookup_cols.keys()}
            if not (src_ok and lk_join_ok and lk_val_ok):
                if not lookup_spec.get("explicit_on_clause"):
                    lookup_spec = None

        case_expr = _extract_case_expression(transformation)
        # Use bare table name instead of alias S
        source_from_clause = f"{effective_src_fq}"
        compare_rhs_expr = f"{effective_src_table}.{src_attr_u}"
        if case_expr and not lookup_spec:
            normalized_case_expr = _normalize_case_expression(case_expr, effective_src_table)
            # Only trust parsed CASE when all referenced source columns exist in KB.
            if _case_expr_refs_only_known_columns(normalized_case_expr, effective_src_table, effective_src_cols):
                source_from_clause = (
                    f"(SELECT {effective_src_table}.*, {normalized_case_expr} AS EXPECTEDVAL "
                    f"FROM {effective_src_fq}) {effective_src_table}"
                )
                compare_rhs_expr = f"{effective_src_table}.EXPECTEDVAL"

        direct_mismatch_sql = (
            "SELECT /*+ PARALLEL(8) */\n"
            "COUNT(*) AS cnt\n"
            f"FROM {source_from_clause}\n"
            f"LEFT JOIN {tgt_fq}\n"
            f"ON {effective_join_on_pretty}\n"
            f"WHERE {_oracle_null_safe_neq(f'{target_table}.{target_col}', compare_rhs_expr) }"
        )

        # Core mapping validation test (skip generic direct check for lookup-based transformations)
        # If this row originates from a lookup/dimension table but no valid lookup spec could be built,
        # skip direct test generation to avoid invalid source-column references on the base staging table.
        if not lookup_spec and _is_lookup_table_name(src_table):
            if literal_expr:
                tests.append({
                    "name": f"Constant Mapping: {target_table}.{target_col}",
                    "test_type": "custom_sql",
                    "source_datasource_id": source_datasource_id,
                    "target_datasource_id": target_datasource_id,
                    "source_query": _build_constant_mismatch_sql(tgt_fq, target_col, literal_expr),
                    "target_query": None,
                    "expected_result": json.dumps({"cnt": 0}),
                    "severity": "medium",
                    "description": f"Validate constant/default assignment for {target_col}.",
                    "source_field": src_attr,
                    "target_field": target_col,
                    "mapping_type": "constant",
                    "transformation_rule": transformation or src_attr or "Constant mapping",
                })
            else:
                skipped_rows.append({
                    "target_field": target_col,
                    "source_table": src_table,
                    "source_field": src_attr,
                    "reason": "Lookup mapping could not derive a valid join key from DRD text or KB",
                })
            continue

        if not lookup_spec:
            tests.append({
                "name": f"{mapping_type.title()} Join Mapping: {effective_src_table}.{src_attr_u} → {target_table}.{target_col}",
                "test_type": "custom_sql",
                "source_datasource_id": source_datasource_id,
                "target_datasource_id": target_datasource_id,
                "source_query": direct_mismatch_sql,
                "target_query": None,
                "expected_result": json.dumps({"cnt": 0}),
                "severity": "high",
                "description": (
                    f"Join-based validation from staging to target. "
                    f"Compares {target_table}.{target_col} vs {effective_src_table}.{src_attr_u} using business keys. "
                    f"{'Transformation/lookup rule: ' + transformation if transformation else 'Direct column mapping.'}"
                ),
                "source_field": src_attr,
                "target_field": target_col,
                "mapping_type": mapping_type,
                "transformation_rule": transformation or "Direct mapping",
            })

        # If there's a lookup/join transformation, generate lookup vs target mismatch validation
        if lookup_spec:
            # Build lookup alias from the bare table name (e.g. CL_VAL_1, ACG_TP_DIM_1)
            lk_full_table = lookup_spec.get("lookup_table", "")
            lk_bare_table = lk_full_table.split(".")[-1]
            lookup_alias_counts[lk_bare_table] = lookup_alias_counts.get(lk_bare_table, 0) + 1
            lk_alias = f"{lk_bare_table}_{lookup_alias_counts[lk_bare_table]}"

            _src_literal = lookup_spec.get("source_lookup_literal", "")
            if _src_literal:
                _lk_join_core = f"{lk_alias}.{lookup_spec['lookup_join_col']} = {_src_literal}"
            else:
                _lk_join_core = _oracle_null_safe_eq(
                    f"{lk_alias}.{lookup_spec['lookup_join_col']}",
                    f"{effective_src_table}.{lookup_spec['source_lookup_col']}",
                )
            # Replace LK. references in extra_filter with the actual alias
            extra_filter = lookup_spec["extra_filter"].replace("LK.", f"{lk_alias}.")
            lookup_join = _lk_join_core + extra_filter
            lookup_join_pretty = re.sub(r'\s+AND\s+', '\nAND ', lookup_join, flags=re.IGNORECASE)
            lookup_value_col = lookup_spec["lookup_value_col"]
            # Determine correct JOIN type for this lookup table
            lookup_join_type = _get_lookup_join_type(lk_full_table)
            complex_source_sql = (
                "SELECT /*+ PARALLEL(8) */\n"
                "COUNT(*) AS cnt\n"
                f"FROM {effective_src_fq}\n"
                f"LEFT JOIN {tgt_fq}\n"
                f"ON {effective_join_on_pretty}\n"
                f"{lookup_join_type} {lk_full_table} {lk_alias}\n"
                f"ON {lookup_join_pretty}\n"
                f"WHERE {_oracle_null_safe_neq(f'{target_table}.{target_col}', f'{lk_alias}.{lookup_value_col}') }"
            )
            tests.append({
                "name": f"Lookup Join Validation: {target_col} ({logical or src_attr_u})",
                "test_type": "custom_sql",
                "source_datasource_id": source_datasource_id,
                "target_datasource_id": target_datasource_id,
                "source_query": complex_source_sql,
                "target_query": None,
                "expected_result": json.dumps({"cnt": 0}),
                "severity": "medium",
                "description": (
                    f"Validate lookup/dimension join output against target value for {target_col}. "
                    f"Transform rule: {transformation}"
                ),
                "source_field": src_attr,
                "target_field": target_col,
                "mapping_type": "complex",
                "transformation_rule": transformation,
            })

    # ── Data type validation tests ──────────────────────────────────────────
    # When source and target data type metadata are both present, generate tests
    # that validate the column data types match expectations.
    target_table_index = _build_kb_table_index(target_datasource_id)
    tgt_cols_kb = _kb_columns_for_table(target_table_index, target_schema, target_table)
    for cm in normalized_rows:
        src_dtype = (cm.get("source_datatype") or cm.get("raw_col_3") or "").strip().upper()
        tgt_dtype = (cm.get("target_datatype_oracle") or "").strip().upper()
        if not src_dtype or not tgt_dtype:
            continue
        physical = cm.get("physical_name") or ""
        target_col = (physical or (cm.get("logical_name") or "").upper().replace(" ", "_")).upper()
        if not target_col:
            continue
        # Only generate test when types are meaningfully different (ignore size differences)
        src_base = re.sub(r'\(.*\)', '', src_dtype).strip()
        tgt_base = re.sub(r'\(.*\)', '', tgt_dtype).strip()
        if src_base == tgt_base:
            continue
        # Build a type-appropriate regex pattern for Oracle REGEXP_LIKE.
        # For numeric target types: count values that do NOT look like numbers.
        # For date/timestamp types: count values that do NOT match a date-like pattern.
        # For all other types: metadata-only — no SQL execution (avoids always-PASS SQL).
        _NUMERIC_BASES = {"NUMBER", "INTEGER", "INT", "FLOAT", "DECIMAL",
                          "NUMERIC", "DOUBLE", "REAL", "BINARY_FLOAT", "BINARY_DOUBLE"}
        _DATE_BASES = {"DATE", "TIMESTAMP"}
        tgt_upper = tgt_base.upper()

        if any(tgt_upper == t or tgt_upper.startswith(t) for t in _NUMERIC_BASES):
            # Oracle POSIX: not a valid decimal number (including scientific notation)
            bad_val_pattern = r"^[+-]?[[:digit:]]+(\\.[[:digit:]]+)?([Ee][+-]?[[:digit:]]+)?$"
            sql_check = (
                f"-- Data type validation: source={src_dtype}, target={tgt_dtype}\n"
                f"SELECT /*+ PARALLEL(8) */\n"
                f"COUNT(*) AS cnt\n"
                f"FROM {tgt_fq}\n"
                f"WHERE {target_table}.{target_col} IS NOT NULL\n"
                f"AND NOT REGEXP_LIKE(TO_CHAR({target_table}.{target_col}), '{bad_val_pattern}')"
            )
            expected = json.dumps({"cnt": 0})
        elif any(tgt_upper == t or tgt_upper.startswith(t) for t in _DATE_BASES):
            # Oracle default date format DD-MON-YY or YYYY-MM-DD
            bad_val_pattern = r"^[0-9]{{2}}-[A-Z]{{3}}-[0-9]{{2,4}}$|^[0-9]{{4}}-[0-9]{{2}}-[0-9]{{2}}"
            sql_check = (
                f"-- Data type validation: source={src_dtype}, target={tgt_dtype}\n"
                f"SELECT /*+ PARALLEL(8) */\n"
                f"COUNT(*) AS cnt\n"
                f"FROM {tgt_fq}\n"
                f"WHERE {target_table}.{target_col} IS NOT NULL\n"
                f"AND NOT REGEXP_LIKE(TO_CHAR({target_table}.{target_col}), '{bad_val_pattern}')"
            )
            expected = json.dumps({"cnt": 0})
        else:
            # Non-numeric/non-date type mismatch: metadata warning only, no SQL
            tests.append({
                "name": f"Data Type Check: {target_table}.{target_col} ({src_dtype} -> {tgt_dtype})",
                "test_type": "info",
                "source_datasource_id": source_datasource_id,
                "target_datasource_id": target_datasource_id,
                "source_query": None,
                "target_query": None,
                "expected_result": None,
                "severity": "low",
                "description": (
                    f"Data type mismatch (metadata only): source column has {src_dtype}, "
                    f"target column {target_col} has {tgt_dtype}. Review if implicit conversion is intended."
                ),
                "source_field": cm.get("source_attribute") or "",
                "target_field": target_col,
                "mapping_type": "datatype_check",
            })
            continue

        tests.append({
            "name": f"Data Type Check: {target_table}.{target_col} ({src_dtype} -> {tgt_dtype})",
            "test_type": "custom_sql",
            "source_datasource_id": source_datasource_id,
            "target_datasource_id": target_datasource_id,
            "source_query": sql_check,
            "target_query": None,
            "expected_result": expected,
            "severity": "low",
            "description": (
                f"Data type mismatch detected: source column has {src_dtype}, "
                f"target column {target_col} has {tgt_dtype}. Review if implicit conversion is intended."
            ),
            "source_field": cm.get("source_attribute") or "",
            "target_field": target_col,
            "mapping_type": "datatype_check",
        })

    valid_tests, invalid_tests = split_valid_invalid_test_defs(tests)
    if invalid_tests:
        for bad in invalid_tests:
            skipped_rows.append({
                "target_field": bad.get("target_field") or bad.get("name") or "",
                "source_table": bad.get("source_field") or "",
                "source_field": bad.get("source_field") or "",
                "reason": (
                    "SQL pattern validation failed: "
                    + "; ".join((bad.get("pattern_errors", {}).get("source") or []) + (bad.get("pattern_errors", {}).get("target") or []))
                ),
            })

    if include_diagnostics:
        return {
            "tests": valid_tests,
            "skipped_rows": skipped_rows,
            "invalid_sql_tests": invalid_tests,
        }
    return valid_tests


def _oracle_null_safe_neq(left_expr: str, right_expr: str) -> str:
    return f"NVL(TO_CHAR({left_expr}), '-999') <> NVL(TO_CHAR({right_expr}), '-999')"


def _oracle_null_safe_eq(left_expr: str, right_expr: str) -> str:
    return f"NVL(TO_CHAR({left_expr}), '-999') = NVL(TO_CHAR({right_expr}), '-999')"


def _extract_case_expression(transformation: str) -> Optional[str]:
    text = (transformation or "").strip()
    if not text:
        return None
    m = re.search(r'(CASE[\s\S]*?END)', text, flags=re.IGNORECASE)
    if not m:
        return None
    return m.group(1).strip()


def _normalize_case_expression(case_expr: str, source_alias: str) -> str:
    """Normalize alias references in CASE expression to use source_alias (table name)."""
    expr = (case_expr or "").strip()
    if not expr:
        return expr
    expr = re.sub(r'\b(?:SRC|SOURCE|S|S0)\.', f'{source_alias}.', expr, flags=re.IGNORECASE)
    return expr


def _build_join_on_sql(join_pairs: List[Tuple[str, str]], target_table: str = "T", source_table: str = "S") -> str:
    if not join_pairs:
        return "1=1"
    return " AND ".join([_oracle_null_safe_eq(f"{target_table}.{tgt}", f"{source_table}.{src}") for tgt, src in join_pairs])


def _derive_join_key_pairs(rows_for_source: List[Dict[str, Any]]) -> List[Tuple[str, str]]:
    if not rows_for_source:
        return []

    mapping_by_target: Dict[str, str] = {}
    source_attrs = set()
    for row in rows_for_source:
        src_attr = (row.get("source_attribute") or "").strip().upper()
        tgt_attr = (row.get("physical_name") or "").strip().upper()
        if src_attr:
            source_attrs.add(src_attr)
        if src_attr and tgt_attr and tgt_attr not in mapping_by_target:
            mapping_by_target[tgt_attr] = src_attr

    preferred_keys = [
        "AR_ID", "CCAL_PD_ID", "ACQ_DT", "TXN_INTT_DT", "TRD_NUM", "SD",
        "TAX_LOT_EXT_REFR_KEY_PRIM", "SRC_STM_ID",
    ]
    aliases = {
        "ACQ_DT": ["ACQ_DT", "TD"],
        "TXN_INTT_DT": ["TXN_INTT_DT", "TXN_INIT_DT"],
        "TAX_LOT_EXT_REFR_KEY_PRIM": ["TAX_LOT_EXT_REFR_KEY_PRIM", "SUB_LOT_EXT_REFR_KEY_PRIM"],
        "SRC_STM_ID": ["SRC_STM_ID", "SRC_STM_CD"],
    }

    join_pairs: List[Tuple[str, str]] = []
    for key in preferred_keys:
        src_col = mapping_by_target.get(key)
        if not src_col:
            for cand in aliases.get(key, []):
                if cand in source_attrs:
                    src_col = cand
                    break
        if src_col:
            join_pairs.append((key, src_col))

    # fallback: use same-name columns that look key-like
    if not join_pairs:
        for tgt_col, src_col in mapping_by_target.items():
            if tgt_col == src_col and (tgt_col.endswith("_ID") or tgt_col.endswith("_DT")):
                join_pairs.append((tgt_col, src_col))
            if len(join_pairs) >= 4:
                break

    return join_pairs


def _split_source_key(source_key: str) -> Tuple[str, str]:
    if "." in source_key:
        parts = source_key.split(".", 1)
        return parts[0], parts[1]
    return "", source_key


def _is_lookup_table_name(table_name: str) -> bool:
    t = (table_name or "").upper()
    return any(marker in t for marker in ["_DIM", "_MAP", "CL_VAL"])


def _get_lookup_join_type(lookup_table: str) -> str:
    """Determine if lookup should use JOIN or LEFT JOIN based on table type.
    
    CL_VAL and system code tables: JOIN (required match in dimension table)
    DIM and MAP tables: LEFT JOIN (optional dimension may not exist)
    """
    t = (lookup_table or "").upper()
    # System code lookups require a match
    if "CL_VAL" in t or "CL_LOOKUP" in t or "LOOKUP_VAL" in t:
        return "JOIN"
    # Dimension and mapping tables are optional
    if "_DIM" in t or "_MAP" in t:
        return "LEFT JOIN"
    # Default to LEFT JOIN for safety
    return "LEFT JOIN"


def _choose_base_source_key(source_tables: Dict[str, List[Dict[str, Any]]]) -> str:
    if not source_tables:
        return ""
    ranked = sorted(
        source_tables.items(),
        key=lambda kv: (
            0 if _is_lookup_table_name(kv[0].split(".")[-1]) else 1,
            len(kv[1]),
        ),
        reverse=True,
    )
    return ranked[0][0]


def _extract_lookup_spec(
    transformation: str,
    src_attr_u: str,
    target_col_u: str,
    src_schema: str = "",
    src_table: str = "",
) -> Optional[Dict[str, str]]:
    if not transformation:
        return None

    text = transformation.strip()
    upper = text.upper()
    if not any(k in upper for k in ["LOOKUP", "LOOK UP", "JOIN", "CL_VAL", "DIM", "MAP"]):
        # Allow through if src_table itself is a lookup/dimension (e.g. "Use X as N" patterns)
        if not (src_table and _is_lookup_table_name(src_table)):
            return None

    def _looks_like_lookup_table(token: str) -> bool:
        t = (token or "").strip().upper()
        if not t:
            return False
        if t.endswith("_ID") and t not in {"CL_VAL"}:
            return False
        if "." in t:
            return True
        return (
            t.endswith("_DIM")
            or t.endswith("_MAP")
            or t.endswith("_LKUP")
            or t.endswith("_LKP")
            or t.endswith("CL_VAL")
            or t == "CL_VAL"
        )

    def _normalize_lookup_table_token(token: str) -> str:
        t = (token or "").strip().upper()
        if "." in t:
            return t
        if src_schema:
            return f"{src_schema.upper()}.{t}"
        return t

    # Pattern 0: Explicit LEFT [OUTER] JOIN schema.table alias ON col1 = col2
    # Include $ in table/column patterns for Oracle identifiers like J$TXN
    m0 = re.search(
        r'LEFT\s+(?:OUTER\s+)?JOIN\s+(?:TO\s+)?'
        r'([\w\.\$#]+)\s+([A-Z_][A-Z0-9_\$#]*)(?:\s+TABLE)?\s+ON\s+'
        r'([\w\.\$#]+)\s*=\s*([\w\.\$#]+)',
        upper,
    )
    if m0:
        lk_tbl = m0.group(1)
        lk_alias = m0.group(2)
        left_token = m0.group(3)
        right_token = m0.group(4)
        left_prefix, left_col = _column_token_parts(left_token)
        right_prefix, right_col = _column_token_parts(right_token)

        # Determine which side of = references the lookup table by alias prefix.
        lk_bare = lk_tbl.split(".")[-1].upper() if "." in lk_tbl else lk_tbl.upper()
        left_is_lookup = (left_prefix.upper() in {lk_alias.upper(), lk_bare, lk_tbl.upper()})
        right_is_lookup = (right_prefix.upper() in {lk_alias.upper(), lk_bare, lk_tbl.upper()})

        if left_is_lookup and not right_is_lookup:
            lk_join_col = left_col
            src_col = right_col
            src_alias = right_prefix
        elif right_is_lookup and not left_is_lookup:
            lk_join_col = right_col
            src_col = left_col
            src_alias = left_prefix
        else:
            # Ambiguous or both match - use positional default (left=lookup)
            lk_join_col = left_col
            src_col = right_col
            src_alias = right_prefix

        extra_filter = _extract_lookup_filters(transformation)
        _tail = upper[m0.end():]
        _tail_m = re.match(r'\s*(AND\b[\s\S]+)$', _tail)
        if _tail_m:
            _tail_filter = (_tail_m.group(1) or "").strip()
            if extra_filter:
                if _tail_filter.upper() != extra_filter.upper():
                    extra_filter = f"{extra_filter} {_tail_filter}"
            else:
                extra_filter = _tail_filter

        # Pattern 0 is explicit JOIN syntax from DRD, so trust the lookup table token.
        if lk_tbl and lk_join_col and src_col:
            return {
                "lookup_table": lk_tbl,
                "source_lookup_col": src_col,
                "lookup_join_col": lk_join_col,
                "lookup_value_col": target_col_u,
                "extra_filter": extra_filter,
                "explicit_on_clause": True,
                "source_alias_hint": src_alias,
            }

    # Pattern 0b: "LOOK UP USING A.B = C.D"
    m0b = re.search(r'(?:LOOK\s*UP|LOOKUP)\s+USING\s+([A-Z0-9_\.]+)\s*=\s*([A-Z0-9_\.]+)', upper)
    if m0b:
        left_tok = m0b.group(1)
        right_tok = m0b.group(2)
        left_tbl, left_col = _column_token_parts(left_tok)
        right_tbl, right_col = _column_token_parts(right_tok)

        src_table_u = (src_table or "").strip().upper().split(".")[-1]
        src_schema_u = (src_schema or "").strip().upper()

        def _matches_src(tbl: str) -> bool:
            t = (tbl or "").strip().upper()
            if not t:
                return False
            if t == src_table_u:
                return True
            if src_schema_u and t == f"{src_schema_u}.{src_table_u}":
                return True
            return False

        # If one side references the row's source table, treat the other side as driving source.
        if _matches_src(right_tbl):
            lk_tbl = right_tbl
            lk_join_col = right_col
            src_col = left_col
        elif _matches_src(left_tbl):
            lk_tbl = left_tbl
            lk_join_col = left_col
            src_col = right_col
        else:
            # Fallback: keep non-TXN side as lookup table when possible.
            left_is_txn = (left_tbl or "").strip().upper().endswith(".TXN") or (left_tbl or "").strip().upper() == "TXN"
            right_is_txn = (right_tbl or "").strip().upper().endswith(".TXN") or (right_tbl or "").strip().upper() == "TXN"
            if left_is_txn and not right_is_txn:
                lk_tbl = right_tbl
                lk_join_col = right_col
                src_col = left_col
            elif right_is_txn and not left_is_txn:
                lk_tbl = left_tbl
                lk_join_col = left_col
                src_col = right_col
            else:
                lk_tbl = right_tbl or left_tbl
                lk_join_col = right_col or left_col
                src_col = left_col or src_attr_u

        lk_tbl = _normalize_lookup_table_token(lk_tbl)
        if _looks_like_lookup_table(lk_tbl) and lk_join_col and src_col:
            return {
                "lookup_table": lk_tbl,
                "source_lookup_col": src_col,
                "lookup_join_col": lk_join_col,
                "lookup_value_col": target_col_u,
                "extra_filter": _extract_lookup_filters(transformation),
                "explicit_on_clause": True,
            }

    # Try multiple patterns to extract the lookup table name
    lookup_table = None

    # Pattern 1: "LOOK UP ON SCHEMA.TABLE" or explicit qualified names
    for pat in [r'(?:LOOK\s*UP|LOOKUP)\s+ON\s+([A-Z0-9_\.]+)', r'(?:LOOK\s*UP|LOOKUP)\s+(?:ON\s+)?([A-Z0-9_\.]+)']:
        m = re.search(pat, upper)
        if m and _looks_like_lookup_table(m.group(1)):
            lookup_table = m.group(1)
            break
    
    # Pattern 2: "Use X in Y in TABLE" (extract TABLE)
    if not lookup_table:
        m = re.search(r'(?:USE|IN)\s+[A-Z0-9_]+\s+(?:IN|UNDER)\s+([A-Z0-9_\.]+)', upper)
        if m and _looks_like_lookup_table(m.group(1)):
            lookup_table = m.group(1)

    # Pattern 2b: "X UNDER CL_VAL_ID" or "X UNDER CL_VAL" (handles both USE-prefixed and plain UNDER forms)
    # Return early to preserve the PICK/GET value column and mark as explicit for KB gate bypass.
    if not lookup_table:
        m2b = re.search(r'([A-Z0-9_]+)\s+UNDER\s+(CL_VAL(?:_ID)?)\b', upper)
        if m2b:
            src_col = m2b.group(1)
            val_col = target_col_u
            pick_m = re.search(r'(?:PICK|GET|RETURN)\s+([A-Z0-9_]+)', upper)
            if pick_m:
                val_col = pick_m.group(1)
            return {
                "lookup_table": "CL_VAL",
                "source_lookup_col": src_col,
                "lookup_join_col": "CL_VAL_ID",
                "lookup_value_col": val_col,
                "extra_filter": _extract_lookup_filters(transformation),
                "explicit_on_clause": True,
            }

    # Pattern 2c: "Use FIELD as CL_VAL_ID [and get CL_VAL_NM/CL_VAL_CD]"
    # Covers: "Use TAX_LOT_TXN_TP_ID as CL_Val_id and get CL_VAL_CD where CL_SCM_ID = 84"
    if not lookup_table:
        m2c = re.search(r'\bUSE\s+([A-Z0-9_]+)\s+AS\s+CL_VAL(?:_ID)?\b', upper)
        if m2c:
            src_col = m2c.group(1)
            val_col = "CL_VAL_NM"  # default
            get_m = re.search(r'\bGET\s+(CL_VAL_(?:NM|CD|CODE|NAME))\b', upper)
            if get_m:
                val_col = get_m.group(1)
            return {
                "lookup_table": "CL_VAL",
                "source_lookup_col": src_col,
                "lookup_join_col": "CL_VAL_ID",
                "lookup_value_col": val_col,
                "extra_filter": _extract_lookup_filters(transformation),
                "explicit_on_clause": True,
            }

    # Pattern 3: "IN TABLE" or "JOIN TABLE"
    if not lookup_table:
        for pat in [r'(?:IN|JOIN|LOOKUP)\s+([A-Z0-9_\.]+(?: *(?:_DIM|_MAP|CL_VAL)))', r'(?:IN|JOIN)\s+([A-Z0-9_\.]+)']:
            m = re.search(pat, upper)
            if m and _looks_like_lookup_table(m.group(1).strip()):
                lookup_table = m.group(1).strip()
                break

    # Pattern 4: "Use X as N [and get Y]" — constant-key lookup against src_table DIM
    if not lookup_table and src_table and _is_lookup_table_name(src_table):
        use_as_m = re.search(r'USE\s+([A-Z0-9_]+)\s+AS\s+(\d+)', upper)
        if use_as_m:
            lk_join_col_raw = use_as_m.group(1)
            src_literal = use_as_m.group(2)
            lk_join_col = lk_join_col_raw
            table_base = re.sub(r'_DIM$|_MAP$', '', src_table.upper())
            if table_base and not lk_join_col.startswith(table_base):
                lk_join_col = f"{table_base}_ID"
            val_col = target_col_u
            get_m = re.search(r'\bGET\s+(CODE|NAME|NM|CD|ID)\b', upper)
            if get_m:
                suffix_map = {"CODE": "CD", "NAME": "NM", "NM": "NM", "CD": "CD", "ID": "ID"}
                suffix = suffix_map.get(get_m.group(1), get_m.group(1))
                val_col = f"{table_base}_{suffix}"
            lk_table = f"{src_schema}.{src_table}" if src_schema else src_table
            return {
                "lookup_table": lk_table,
                "source_lookup_col": lk_join_col,
                "lookup_join_col": lk_join_col,
                "lookup_value_col": val_col,
                "extra_filter": "",
                "source_lookup_literal": src_literal,
                "explicit_on_clause": True,
            }

    # Pattern 5: "Use TABLE.COL to get VALUE_COL" when src_table itself is a DIM/MAP.
    # Covers: "Use OPN_TAX_LOTS_NONBKR_TGT.ACG_TP_CODE to get ACG_TP_ID"
    if not lookup_table and src_table and _is_lookup_table_name(src_table):
        use_to_get_m = re.search(r'\bUSE\s+(?:[A-Z0-9_]+\.)?([A-Z0-9_]+)\s+TO\s+GET\s+([A-Z0-9_]+)', upper)
        if use_to_get_m:
            source_join_col = use_to_get_m.group(1)
            get_col = use_to_get_m.group(2)
            # Normalize staging column name: ACG_TP_CODE → ACG_TP_CD (DIM join convention)
            dim_join_col = re.sub(r'_CODE$', '_CD', source_join_col.upper())
            lk_table = f"{src_schema}.{src_table}" if src_schema else src_table
            return {
                "lookup_table": lk_table,
                "source_lookup_col": source_join_col,
                "lookup_join_col": dim_join_col,
                "lookup_value_col": get_col,
                "extra_filter": "",
                "explicit_on_clause": True,
            }

    if not lookup_table:
        return None

    if lookup_table.endswith("_ID") and "CL_VAL" in lookup_table:
        lookup_table = "CL_VAL"
    
    source_lookup_col, lookup_join_col = _infer_lookup_pair_from_text(transformation, src_attr_u)
    
    # Extract target value column from lookup reference
    value_col = target_col_u
    
    # Explicitly mentioned "pick" columns take priority
    pick_match = re.search(r'(?:PICK|GET|GET VALUE|RETURN)\s+([A-Z0-9_]+)', upper)
    if pick_match:
        value_col = pick_match.group(1)
    
    # Common lookup value columns by marker
    for marker in ["CL_VAL_NM", "CL_VAL_CD", "ACG_TP_NM", " ACG_TP_ID", "ACG_TP_CD", "SRC_STM_NM", "SRC_STM_CD", "CIRD_PD_ID", "CIRD_PD_NM"]:
        if marker in upper:
            value_col = marker.strip()
            break
    
    extra_filter = _extract_lookup_filters(transformation)

    return {
        "lookup_table": lookup_table,
        "source_lookup_col": source_lookup_col,
        "lookup_join_col": lookup_join_col,
        "lookup_value_col": value_col,
        "extra_filter": extra_filter,
    }


def _build_transformation_query(
    src_schema: str,
    src_table: str,
    src_attr: str,
    transformation: str,
    record: Dict[str, Any],
) -> Tuple[str, Optional[str]]:
    """Build SQL query based on transformation rules from DRD.

    Returns (basic_query, join_query) where join_query includes the full join.
    """
    src_fq = _fq_name(src_schema, src_table)

    # Basic query: just count distinct/total from source
    basic_query = (
        f"SELECT COUNT(DISTINCT {src_attr}) AS distinct_cnt, "
        f"COUNT(*) AS total_cnt, "
        f"SUM(CASE WHEN {src_attr} IS NULL THEN 1 ELSE 0 END) AS null_cnt "
        f"FROM {src_fq}"
    )

    if not transformation:
        return basic_query, None

    # Try to build a join query from transformation text
    join_query = None
    transform_lower = transformation.lower().strip()
    transform_upper = transformation.upper().strip()

    def _extract_lookup_table(text: str) -> str:
        patterns = [
            r'lookup\s+on\s+([a-z0-9_\.]+)',
            r'look\s*up\s+on\s+([a-z0-9_\.]+)',
            r'join\s+([a-z0-9_\.]+)',
            r'\bin\s+([a-z0-9_\.]+)',
        ]
        for pattern in patterns:
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                return m.group(1).upper()
        return ""

    def _extract_join_parts(text: str) -> Tuple[str, str, str]:
        source_col = src_attr.upper()
        lookup_col = src_attr.upper()
        extra_filter = ""

        # Pattern: "SUB_LOT_TXN_TP_ID under CL_VAL_ID"
        under_match = re.search(r'([A-Z0-9_]+)\s+UNDER\s+([A-Z0-9_]+)', text.upper())
        if under_match:
            source_col = under_match.group(1)
            lookup_col = under_match.group(2)

        # Pattern: explicit equality
        eq_matches = re.findall(r'([A-Z0-9_\.]+)\s*=\s*([A-Z0-9_\.]+)', text.upper())
        if eq_matches:
            left, right = eq_matches[0]
            left_col = left.split('.')[-1]
            right_col = right.split('.')[-1]
            if source_col in (left_col, right_col):
                lookup_col = right_col if left_col == source_col else left_col
            else:
                source_col = left_col
                lookup_col = right_col

        scm_match = re.search(r'CL_SCM_ID\s*=\s*(\d+)', text.upper())
        if scm_match:
            extra_filter = f" AND lk.CL_SCM_ID = {scm_match.group(1)}"

        condition = f"{src_table}.{source_col} = lk.{lookup_col}"
        return condition, lookup_col, extra_filter

    # Pattern: "LEFT JOIN schema.table ON condition"
    join_match = re.search(
        r'(?:left\s+)?join\s+([\w.]+)\s+(\w+)?\s*(?:on\s+(.+?))?(?:$|and\s+|where\s+)',
        transformation,
        re.IGNORECASE | re.DOTALL,
    )
    if join_match:
        join_table = join_match.group(1)
        alias = join_match.group(2) or (join_table.split('.')[-1].upper() if join_table else "LK")
        derived_condition, lookup_key_col, extra_filter = _extract_join_parts(transformation)
        condition = join_match.group(3) or derived_condition
        # Replace placeholder 'lk.' with actual alias
        condition = re.sub(r'\blk\.', f'{alias}.', condition, flags=re.IGNORECASE)
        extra_filter = re.sub(r'\blk\.', f'{alias}.', extra_filter, flags=re.IGNORECASE)

        where_filter = ""
        where_match = re.search(r'where\s+(.+)$', transformation, re.IGNORECASE)
        if where_match:
            where_filter = f" AND ({where_match.group(1).strip()})"

        join_query = (
            f"SELECT COUNT(*) AS total_cnt, "
            f"SUM(CASE WHEN {alias}.{lookup_key_col} IS NULL THEN 1 ELSE 0 END) AS unmatched_lookup_cnt "
            f"FROM {src_fq} "
            f"LEFT JOIN {join_table} {alias} ON {condition}{extra_filter}{where_filter}"
        )

    # Pattern: "Look up using TABLE.COL = TABLE.COL"
    lookup_match = re.search(r'look\s*up\s+using\s+(.+)', transform_lower)
    if lookup_match:
        join_table = _extract_lookup_table(transform_lower)
        condition, lookup_key_col, extra_filter = _extract_join_parts(lookup_match.group(1))
        if join_table:
            _lk_bare = join_table.split('.')[-1].upper() if join_table else 'LK'
            condition = re.sub(r'\blk\.', f'{_lk_bare}.', condition, flags=re.IGNORECASE)
            extra_filter = re.sub(r'\blk\.', f'{_lk_bare}.', extra_filter, flags=re.IGNORECASE)
            join_query = (
                f"SELECT COUNT(*) AS total_cnt, "
                f"SUM(CASE WHEN {_lk_bare}.{lookup_key_col} IS NULL THEN 1 ELSE 0 END) AS unmatched_lookup_cnt "
                f"FROM {src_fq} "
                f"LEFT JOIN {join_table} {_lk_bare} ON {condition}{extra_filter}"
            )

    if not join_query and ("lookup" in transform_lower or "look up" in transform_lower):
        join_table = _extract_lookup_table(transform_lower)
        if join_table:
            condition, lookup_key_col, extra_filter = _extract_join_parts(transform_upper)
            _lk_bare2 = join_table.split('.')[-1].upper() if join_table else 'LK'
            condition = re.sub(r'\blk\.', f'{_lk_bare2}.', condition, flags=re.IGNORECASE)
            extra_filter = re.sub(r'\blk\.', f'{_lk_bare2}.', extra_filter, flags=re.IGNORECASE)
            join_query = (
                f"SELECT COUNT(*) AS total_cnt, "
                f"SUM(CASE WHEN {_lk_bare2}.{lookup_key_col} IS NULL THEN 1 ELSE 0 END) AS unmatched_lookup_cnt "
                f"FROM {src_fq} "
                f"LEFT JOIN {join_table} {_lk_bare2} ON {condition}{extra_filter}"
            )

    return basic_query, join_query
