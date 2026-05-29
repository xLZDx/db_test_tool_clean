#!/usr/bin/env python3
"""Extract column definitions from DRD Excel file."""
import openpyxl
from pathlib import Path

# Try to find the DRD Excel file
possible_locations = [
    "DRD_Activity_Fact.xlsx",
    "../DRD_Activity_Fact.xlsx",
    "c:/Users/ikorostelev/Downloads/DRD_Activity_Fact.xlsx",
]

excel_file = None
for loc in possible_locations:
    if Path(loc).exists():
        excel_file = loc
        print(f"Found: {loc}")
        break

if not excel_file:
    # List files in current directory
    print("Files in current dir:")
    for f in Path(".").glob("*.xlsx"):
        print(f"  {f}")
    print("\nFiles in parent:")
    for f in Path("..").glob("*.xlsx"):
        print(f"  {f}")
    exit(1)

# Load workbook
wb = openpyxl.load_workbook(excel_file)
print(f"\nSheet names: {wb.sheetnames}")

# Try to find the table view (view tab2, or similar)
target_sheet = None
for sheet_name in wb.sheetnames:
    if 'view' in sheet_name.lower() or 'tab' in sheet_name.lower() or 'table' in sheet_name.lower():
        target_sheet = sheet_name
        print(f"Trying sheet: {sheet_name}")
        break

if not target_sheet:
    target_sheet = wb.sheetnames[0]  # default to first sheet

print(f"\nUsing sheet: {target_sheet}")
ws = wb[target_sheet]

print(f"Max columns: {ws.max_column}, Max rows: {ws.max_row}")

# Extract columns from first row
columns = []
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    if cell.value:
        col_name = str(cell.value).strip()
        columns.append(col_name)

print(f"\nTotal columns: {len(columns)}")
print("\nColumn list:")
for i, col in enumerate(columns, 1):
    print(f"{i:3d}. {col}")

# Save to JSON for later use
import json
with open('drd_columns.json', 'w') as f:
    json.dump(columns, f, indent=2)
print(f"\nSaved to drd_columns.json")
