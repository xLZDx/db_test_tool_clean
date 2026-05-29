#!/usr/bin/env python3
"""Extract columns from Table-View (2) sheet - likely has one column per row."""
import openpyxl
import json

excel_file = "c:/Users/ikorostelev/Downloads/DRD_Activity_Fact.xlsx"
wb = openpyxl.load_workbook(excel_file)

# Try the sheet that has 385 rows - likely the column list
ws = wb['Table-View (2)']
print(f"Sheet: Table-View (2)")
print(f"Columns: {ws.max_column}, Rows: {ws.max_row}\n")

# Read column A (first column) which likely has the column names
columns = []
for row in range(2, ws.max_row + 1):  # Skip header
    cell = ws.cell(row=row, column=1)
    if cell.value:
        col_name = str(cell.value).strip()
        if col_name and not col_name.startswith('View Name'):
            columns.append(col_name)

print(f"Total columns extracted: {len(columns)}\n")
print("First 30 columns:")
for i, col in enumerate(columns[:30], 1):
    print(f"  {i:3d}. {col}")

if len(columns) > 30:
    print(f"\n... ({len(columns)-30} more) ...\n")
    print("Last 20 columns:")
    for i, col in enumerate(columns[-20:], len(columns)-19):
        print(f"  {i:3d}. {col}")

# Save to JSON
with open('drd_columns.json', 'w') as f:
    json.dump(columns, f, indent=2)

print(f"\nSaved {len(columns)} columns to drd_columns.json")
